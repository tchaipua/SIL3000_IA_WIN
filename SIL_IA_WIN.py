import customtkinter as ctk
from tkinter import ttk
import tkinter as tk
from tkinter import filedialog
import pyodbc
import sys
import os
import time
import requests
import urllib.parse
from unidecode import unidecode
from tkinter import messagebox
from PIL import Image
from datetime import datetime, date

import pandas as pd
import json
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import mplcursors

class ToolTip(object):
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x = y = 0
        try:
            x, y, cx, cy = self.widget.bbox("insert")
        except:
            pass
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)
        label = tk.Label(tw, text=self.text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         font=("Arial", "9", "normal"))
        label.pack(ipadx=2, ipady=1)

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

class BaseWindow(ctk.CTkFrame):
    def __init__(self, parent, title_str, id_str):
        self.id_str = id_str
        self.ultima_query_bruta = "" # Sistema de Auditoria Global (Audit-Ready)
        super().__init__(parent, fg_color="#FFFFFF", corner_radius=0)
        self.grid_rowconfigure(1, weight=1); self.grid_columnconfigure(0, weight=1)

        # Top Bar (Cabeçalho Chumbo)
        self.top_frame = ctk.CTkFrame(self, fg_color="#D1D5DB", corner_radius=0)
        self.top_frame.grid(row=0, column=0, sticky="ew")
        
        # --- NOVO: Linha de Título Superior ---
        self.title_bar = ctk.CTkFrame(self.top_frame, fg_color="transparent", height=25)
        self.title_bar.pack(side="top", fill="x", padx=20, pady=(0, 0))


        
        # Título da Tela à Esquerda (Azul)
        self.lbl_titulo = ctk.CTkLabel(self.title_bar, text=title_str.upper(), font=("Arial", 16, "bold"), text_color="#1565C0")
        self.lbl_titulo.pack(side="left", pady=0)



        # Grid frame
        self.grid_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=0)
        self.grid_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 10))


        style = ttk.Style()
        # --- PADRÃO PREMIUM: ATIVAR TEMA CLAM PARA CORES CABEÇALHO NO WINDOWS ---
        try:
            style.theme_use('clam')
        except:
            pass # Caso o sistema não suporte clam (raro)

        # Configurar Estilo Base para este Grid
        style.configure(id_str + ".Treeview", 
                        background="#FFFFFF", fieldbackground="#FFFFFF", foreground="black", 
                        rowheight=32, font=("Arial", 11), borderwidth=0)
        
        # Cabeçalho Reformulado: Azul Total (#0D47A1), Fonte 13 Bold
        style.configure(id_str + ".Treeview.Heading", 
                        background="#0D47A1", foreground="white", 
                        font=("Arial", 13, "bold"), borderwidth=1, relief="flat")
        
        style.map(id_str + ".Treeview.Heading", 
                  background=[('active', '#1565C0'), ('pressed', '#0D47A1')],
                  foreground=[('active', 'white')])

        self.tree = ttk.Treeview(self.grid_frame, show="headings", selectmode="browse", style=id_str + ".Treeview")
        self.tree.pack(fill="both", expand=True, side="left")
        self.tree.tag_configure('even', background='#FFFFFF')
        self.tree.tag_configure('odd', background='#F1F5F9') # Azul bem clarinho para alternar

        # Rodape de Botões (Posição final row=3)
        self.bottom_bar = ctk.CTkFrame(self, fg_color="transparent", height=40, corner_radius=0)
        self.bottom_bar.grid(row=3, column=0, sticky="ew")
        
        btn_sair = ctk.CTkButton(self.bottom_bar, text="✖  Fechar Tela", command=self.fechar_tela, fg_color="transparent", hover_color="#E0E0E0", text_color="black", font=("Arial", 12, "bold"), border_width=2, border_color="black", width=130)
        btn_sair.pack(side="left", padx=20, pady=5)

        # NOVO: Botão de Auditoria SQL Global (?)
        self.btn_debug_sql = ctk.CTkButton(self.bottom_bar, text="?", width=28, height=28, corner_radius=14, fg_color="#C62828", hover_color="#B71C1C", text_color="white", font=("Arial", 12, "bold"), command=self.mostrar_sql_bruto)
        self.btn_debug_sql.pack(side="left", padx=(0, 10), pady=0)
        ToolTip(self.btn_debug_sql, "VER COMANDO SQL BRUTO (Audit)")

        # NOVO PADRÃO: Contador de Registros no Canto Esquerdo
        self.lbl_total_regs = ctk.CTkLabel(self.bottom_bar, text="Total: 0 registros", font=("Arial", 11, "bold"), text_color="#475569")
        self.lbl_total_regs.pack(side="left", padx=10)

        # Botões de Exportação
        self.btn_export_pdf = ctk.CTkButton(self.bottom_bar, text="📄 Exportar PDF", width=120, height=32, fg_color="#E53935", hover_color="#B71C1C", command=self.exportar_pdf)
        self.btn_export_pdf.pack(side="right", padx=5)
        
        self.btn_export_excel = ctk.CTkButton(self.bottom_bar, text="📊 Exportar Excel", width=120, height=32, fg_color="#43A047", hover_color="#2E7D32", command=self.exportar_excel)
        self.btn_export_excel.pack(side="right", padx=5)

        from tkinter import messagebox
        btn_copy_id = ctk.CTkButton(self.bottom_bar, text="📋", width=25, height=25, fg_color="transparent", hover_color="#E0E0E0", text_color="black", font=("Arial", 11, "bold"), command=self.acao_copiar_id)
        btn_copy_id.pack(side="right", padx=(5, 20))

        self.lbl_id = ctk.CTkLabel(self.bottom_bar, text=f"Tela: {id_str}", font=("Arial", 11, "bold"), text_color="gray")
        self.lbl_id.pack(side="right", padx=5)
        ToolTip(btn_copy_id, "COPIAR NOME TELA")

    def fechar_tela(self):
        if hasattr(self, 'hub'):
            if getattr(self, 'is_posto_table', False):
                self.hub.abrir_posto()
                return
            self.pack_forget()
            if hasattr(self.hub, 'modules_frame'):
                self.hub.modules_frame.pack(fill="both", expand=True, padx=30, pady=30)
        else:
            self.pack_forget()
        self.destroy()
    def acao_copiar_id(self):
        """Copia apenas o identificador da tela para o clipboard."""
        self.clipboard_clear()
        self.clipboard_append(self.id_str)
        self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def abrir_calendario(self, entry_target):
        """Implementação de um seletor de data dinâmico (Pop-up). Todos os módulos herdam este seletor."""
        pop = ctk.CTkToplevel(self)
        pop.title("📅 Selecionar Data")
        pop.geometry("320x400")
        pop.attributes("-topmost", True)
        pop.grab_set()
        
        # Centralizar pop-up
        w, h = 320, 400
        x = (pop.winfo_screenwidth() // 2) - (w // 2)
        y = (pop.winfo_screenheight() // 2) - (h // 2)
        pop.geometry(f"{w}x{h}+{x}+{y}")

        from datetime import date
        import calendar
        hoje = date.today()
        cal_state = {"mes": hoje.month, "ano": hoje.year}

        main_f = ctk.CTkFrame(pop, fg_color="white")
        main_f.pack(fill="both", expand=True, padx=10, pady=10)

        header_f = ctk.CTkFrame(main_f, fg_color="transparent")
        header_f.pack(fill="x", pady=5)
        
        lbl_mes_ano = ctk.CTkLabel(header_f, text="", font=("Arial", 14, "bold"), text_color="#1E88E5")
        
        def update_view():
            mes, ano = cal_state["mes"], cal_state["ano"]
            nome_mes = ["?", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][mes]
            lbl_mes_ano.configure(text=f"{nome_mes} {ano}")
            for widget in grid_f.winfo_children():
                if int(widget.grid_info()["row"]) > 0: widget.destroy()
            cal = calendar.monthcalendar(ano, mes)
            for r, week in enumerate(cal):
                for c, day in enumerate(week):
                    if day != 0:
                        btn = ctk.CTkButton(grid_f, text=str(day), width=35, height=35, fg_color="#F1F5F9", text_color="black", hover_color="#CBD5E1", 
                                          command=lambda d=day: set_date(d))
                        btn.grid(row=r+1, column=c, padx=2, pady=2)

        def mudar_mes(inc):
            cal_state["mes"] += inc
            if cal_state["mes"] > 12: cal_state["mes"] = 1; cal_state["ano"] += 1
            elif cal_state["mes"] < 1: cal_state["mes"] = 12; cal_state["ano"] -= 1
            update_view()

        ctk.CTkButton(header_f, text="◀", width=30, height=30, fg_color="#94A3B8", command=lambda: mudar_mes(-1)).pack(side="left", padx=10)
        lbl_mes_ano.pack(side="left", expand=True)
        ctk.CTkButton(header_f, text="▶", width=30, height=30, fg_color="#94A3B8", command=lambda: mudar_mes(1)).pack(side="right", padx=10)

        grid_f = ctk.CTkFrame(main_f, fg_color="transparent")
        grid_f.pack(fill="both", expand=True)
        dias_sem = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
        for i, ds in enumerate(dias_sem):
            ctk.CTkLabel(grid_f, text=ds, font=("Arial", 10, "bold"), text_color="gray").grid(row=0, column=i, pady=5)

        def set_date(dia):
            entry_target.delete(0, "end")
            entry_target.insert(0, f"{dia:02d}/{cal_state['mes']:02d}/{cal_state['ano']}")
            pop.destroy()

        update_view()

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL Pronto/Executado seguindo o padrão premium visual."""
        resumo = self.get_sql_summary() if hasattr(self, 'get_sql_summary') else ""
        query = self.get_current_query() if hasattr(self, 'get_current_query') else getattr(self, 'ultima_query_bruta', "")
        
        if not query:
            messagebox.showinfo("SQL Bruto", "Aguardando carregamento de dados...")
            return
            
        conteudo_completo = f"{resumo}\n\n{'='*50}\n\n{query}" if resumo else query
            
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", conteudo_completo)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def copiar_sql_debug(self, sql, window):
        """Copia para o clipboard com feedback."""
        self.clipboard_clear(); self.clipboard_append(sql); self.update()
        messagebox.showinfo("Clipboard", "SQL Copiado!", parent=window)
        window.destroy()

    def configuring_grid(self, columns, headings, widths, aligns):
        """Método de compatibilidade (v1)"""
        self.configurar_grid(columns, headings, widths, aligns)

    def configurar_grid(self, columns, headings, widths, aligns):
        """Método centralizado para configurar colunas, cabeçalhos e o rodapé de totais."""
        self.tree["columns"] = columns
        self.headers_dict = {col: head for col, head in zip(columns, headings)}
        
        for i, col in enumerate(columns):
            self.tree.heading(col, text=headings[i] + " ↕", anchor=aligns[i], command=lambda c=col: self.ordenar_por(c, False))
            self.tree.column(col, width=widths[i], anchor=aligns[i])
            
        # Criar Treeview de Totais (Novo Padrão Edge-to-Edge: Barra Azul Total)
        if not hasattr(self, "tree_totais"):
            # Frame de Totais ocupando largura total na row=2
            self.frame_totais_container = ctk.CTkFrame(self, fg_color="#0D47A1", height=30, corner_radius=0)
            self.frame_totais_container.grid(row=2, column=0, sticky="ew", padx=0)
            
            style = ttk.Style()
            # Estilo Premium para o rodapé (Azul #0D47A1)
            style_name = "ExcelTotal.Treeview"
            style.configure(style_name, 
                            background="#0D47A1", fieldbackground="#0D47A1", 
                            foreground="white", font=("Arial", 11, "bold"), rowheight=30,
                            borderwidth=0, highlightthickness=0)
            
            # Estilo para os headers não aparecerem (caso o tema clam force)
            style.configure("ExcelTotal.Treeview.Heading", background="#0D47A1", foreground="#0D47A1", font=("Arial", 1))

            self.tree_totais = ttk.Treeview(self.frame_totais_container, show="", height=1, columns=columns, style=style_name)
            self.tree_totais.pack(fill="x", side="left", expand=True) # expand=True para forçar stretch
            
            # --- NOVO: Compensador de Scrollbar ---
            # Adicionar um pequeno espaço na direita para compensar a scrollbar do grid principal
            self.scrollbar_spacer = ctk.CTkFrame(self.frame_totais_container, width=16, height=30, fg_color="#0D47A1", corner_radius=0)
            self.scrollbar_spacer.pack(side="right")

            for i, col in enumerate(columns):
                self.tree_totais.column(col, width=widths[i], anchor=aligns[i])

    def ordenar_por(self, col, reverse):
        """Lógica de ordenação genérica para as janelas que utilizam BaseWindow."""
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        def clean_val(v):
            if not v: return ""
            s = str(v).strip().upper()
            # 1. Moedas R$ 1.234,56
            if "R$" in s:
                try: return float(s.replace("R$", "").replace(".", "").replace(",", ".").replace(" ", "").strip())
                except: return 0.0
            # 2. Datas DD/MM/AAAA
            if "/" in s and len(s) == 10 and s[2] == "/" and s[5] == "/":
                try:
                    from datetime import datetime
                    return datetime.strptime(s, "%d/%m/%Y")
                except: pass
            # 3. Números Genéricos
            if s.replace(".", "").replace("-", "").isdigit():
                try: return float(s)
                except: return s
            return s
            
        data.sort(key=lambda t: clean_val(t[0]), reverse=reverse)
        for index, (val, k) in enumerate(data):
            self.tree.move(k, "", index)
        
        # Inverter seta no cabeçalho
        for c in self.tree["columns"]:
            anc = self.tree.column(c, "anchor")
            if c == col:
                seta = " ▼" if reverse else " ▲"
                self.tree.heading(c, text=self.headers_dict[c] + seta, anchor=anc, command=lambda _c=c: self.ordenar_por(_c, not reverse))
            else:
                self.tree.heading(c, text=self.headers_dict.get(c, c) + " ↕", anchor=anc, command=lambda _c=c: self.ordenar_por(_c, False))
        
        self.re_zebrar()

    def re_zebrar(self):
        """Atualiza o zebrado (Padrão: even/odd) no Treeview e conta os registros."""
        items = self.tree.get_children()
        for i, item in enumerate(items):
            tag = 'even' if i % 2 == 0 else 'odd'
            self.tree.item(item, tags=(tag,))
        
        # Atualiza o contador de registros global
        if hasattr(self, "lbl_total_regs"):
            self.lbl_total_regs.configure(text=f"Total: {len(items)} registros")

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: BASE ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def exportar_excel(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        try:
             columns = self.tree["columns"]; cols_titles = [self.tree.heading(c)["text"] for c in columns]; data = [self.tree.item(k)['values'] for k in items]
             import pandas as pd, tempfile, os
             df = pd.DataFrame(data, columns=cols_titles)
             filepath = os.path.join(tempfile.gettempdir(), "Relatorio_Dados.xlsx")
             
             writer = pd.ExcelWriter(filepath, engine='openpyxl')
             df.to_excel(writer, index=False, sheet_name='Sheet1')
             worksheet = writer.sheets['Sheet1']
             for i, col in enumerate(df.columns):
                  max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                  from openpyxl.utils import get_column_letter
                  worksheet.column_dimensions[get_column_letter(i+1)].width = max_len
             writer.close()
             
             os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro Excel", str(e))

    def exportar_pdf(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        try:
             columns = self.tree["columns"]; cols_titles = [self.tree.heading(c)["text"] for c in columns]; data = [self.tree.item(k)['values'] for k in items]
             import tempfile, os
             filepath = os.path.join(tempfile.gettempdir(), "Relatorio_Dados.pdf")
             import sys, subprocess
             try: import fpdf
             except: subprocess.run([sys.executable, "-m", "pip", "install", "fpdf"], capture_output=True); import fpdf
             pdf = fpdf.FPDF(); pdf.add_page(); pdf.set_font("Arial", 'B', 12); pdf.cell(190, 10, "RELATORIO DE DADOS", 0, 1, 'C'); pdf.ln(5)
             pdf.set_font("Arial", 'B', 8); num_cols = len(cols_titles); w_col = 190 / num_cols if num_cols > 0 else 30
             for c in cols_titles: 
                  c_str = str(c).encode('latin-1', 'replace').decode('latin-1')
                  pdf.cell(w_col, 8, c_str, 1, 0, 'C')
             pdf.ln(8); pdf.set_font("Arial", '', 7)
             for r in data:
                  for val in r: 
                       v_str = str(val)[:20].encode('latin-1', 'replace').decode('latin-1')
                       pdf.cell(w_col, 7, v_str, 1, 0, 'C')
                  pdf.ln(7)
             if hasattr(self, 'lbl_rodape_total'):
                  pdf.ln(5); pdf.set_font("Arial", 'B', 10)
                  txt_rodape = self.lbl_rodape_total.cget("text").encode('latin-1', 'replace').decode('latin-1')
                  pdf.cell(190, 10, txt_rodape, 0, 1, 'R')
             pdf.ln(5); pdf.output(filepath); os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro PDF", str(e))


class TabelaConfigWindow(ctk.CTkToplevel):
    def __init__(self, parent, config, update_callback):
        super().__init__(parent)
        self.grab_set()  # Trava a janela pai
        self.title("Mapeamento da Tabela")
        
        # Centraliza a tela
        w, h = 450, 600
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        self.transient(parent)
        self.update_callback = update_callback
        
        # 1. Barra de Identificação de Tela (Absolute Bottom)
        self.bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_bar.pack(side="bottom", fill="x", padx=20, pady=(0, 10))
        
        self.frame_id = ctk.CTkFrame(self.bottom_bar, fg_color="transparent")
        self.frame_id.pack(side="right")
        
        self.lbl_tela_id = ctk.CTkLabel(self.frame_id, text="Tela: CFG_TABELA_CEP", font=ctk.CTkFont(size=11, weight="bold"), text_color="gray")
        self.lbl_tela_id.pack(side="left")
        
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=20, height=20, fg_color="transparent", hover_color="#E0E0E0", text_color="black", command=lambda: self.copiar_id("CFG_TABELA_CEP"))
        self.btn_copy_id.pack(side="left", padx=5)
        ToolTip(self.btn_copy_id, "COPIAR NOME TELA")

        # 2. Barra de Botões (Logo acima da Identificação)
        self.buttons_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.buttons_frame.pack(side="bottom", fill="x", padx=30, pady=(15, 5))
        
        # Botão Retornar à Esquerda
        self.btn_sair = ctk.CTkButton(self.buttons_frame, text="❌ Retornar", width=140, height=32, fg_color="#C62828", text_color="white", hover_color="#B12020", command=self.destroy)
        self.btn_sair.pack(side="left")

        # Botão Salvar à Direita
        self.btn_salvar = ctk.CTkButton(self.buttons_frame, text="💾 Salvar", width=140, height=32, command=self.salvar, fg_color="#43A047", hover_color="#2E7D32")
        self.btn_salvar.pack(side="right")

        # --- Corpo da Tela (Formulário) ---
        self.lbl_title = ctk.CTkLabel(self, text="⚙️ MAPEAMENTO DA TABELA", font=ctk.CTkFont(size=18, weight="bold"), text_color="#1565C0", anchor="w")
        self.lbl_title.pack(pady=(20, 10), padx=30, fill="x")

        self.e_tabela = self.criar_input("Nome da Tabela", config["tabela"])
        self.e_id = self.criar_input("Coluna do ID/Código", config["col_id"])
        self.e_nome = self.criar_input("Coluna do Nome", config["col_nome"])
        self.e_cep = self.criar_input("Coluna do CEP", config["col_cep"])
        self.e_rua = self.criar_input("Coluna da Rua", config["col_rua"])
        self.e_log = self.criar_input("Coluna do Logradouro", config["col_logradouro"])
        self.e_bairro = self.criar_input("Coluna do Bairro", config["col_bairro"])
        self.e_uf = self.criar_input("Estado sempre (Valor Fixo)", config["col_uf"])

    def criar_input(self, label, valor, show=None):
        ctk.CTkLabel(self, text=label, anchor="w").pack(fill="x", padx=30, pady=(10, 0))
        entry = ctk.CTkEntry(self, width=340, show=show); entry.pack(fill="x", padx=30, pady=(0, 5))
        if valor: entry.insert(0, valor)
        return entry

    def copiar_id(self, texto):
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def salvar(self):
        new_cfg = {
            "tabela": self.e_tabela.get(),
            "col_id": self.e_id.get(),
            "col_nome": self.e_nome.get(),
            "col_cep": self.e_cep.get(),
            "col_rua": self.e_rua.get(),
            "col_logradouro": self.e_log.get(),
            "col_bairro": self.e_bairro.get(),
            "col_uf": self.e_uf.get()
        }
        self.update_callback(new_cfg)
        self.destroy()

class ConfigWindow(ctk.CTkToplevel):
    def __init__(self, parent, config, update_callback):
        super().__init__(parent)
        self.grab_set()
        self.title("⚙️ Configurações de Acesso")
        
        # Centraliza a tela
        w, h = 480, 580
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        self.transient(parent)
        self.update_callback = update_callback

        # --- Título Superior ---
        self.lbl_title = ctk.CTkLabel(self, text="🛠️ CONFIGURAÇÕES TÉCNICAS", font=ctk.CTkFont(size=20, weight="bold"), text_color="#1565C0")
        self.lbl_title.pack(pady=(25, 20))

        # --- Inputs ---
        self.e_nome = self.criar_input("👤 Nome do Operador", config.get("nome_usuario", "Operador"))
        self.e_servidor = self.criar_input("🖥️ Servidor SQL (SERVER\\INSTANCE)", config.get("servidor", ""))
        self.e_banco = self.criar_input("🗄️ Nome do Banco de Dados", config.get("banco", ""))
        self.e_user = self.criar_input("👤 Usuário do Banco (UID)", config.get("usuario_bd", "sa"))
        self.e_senha = self.criar_input("🔑 Senha do Banco (PWD)", config.get("senha_bd", ""), show="*")

        # --- Barra de Identificação de Tela ---
        self.id_str = "CFG_ACESSO"
        self.bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_bar.pack(side="bottom", fill="x", padx=20, pady=(0, 10))
        
        self.frame_id = ctk.CTkFrame(self.bottom_bar, fg_color="transparent")
        self.frame_id.pack(side="right")
        self.lbl_tela_id = ctk.CTkLabel(self.frame_id, text=f"Tela: {self.id_str}", font=ctk.CTkFont(size=11, weight="bold"), text_color="gray")
        self.lbl_tela_id.pack(side="left")
        
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=20, height=20, fg_color="transparent", hover_color="#E0E0E0", text_color="black", command=self.acao_copiar_id)
        self.btn_copy_id.pack(side="left", padx=5)
        ToolTip(self.btn_copy_id, "COPIAR NOME TELA")

        # --- Barra de Botões ---
        self.buttons_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.buttons_frame.pack(side="bottom", fill="x", padx=30, pady=(20, 15))
        
        ctk.CTkButton(self.buttons_frame, text="✖ Cancelar", width=140, height=35, fg_color="#C62828", hover_color="#B12020", command=self.destroy).pack(side="left")
        ctk.CTkButton(self.buttons_frame, text="💾 Salvar Configurações", width=220, height=35, fg_color="#2E7D32", hover_color="#1B5E20", command=self.salvar).pack(side="right")

    def criar_input(self, label, valor, show=None):
        f = ctk.CTkFrame(self, fg_color="transparent")
        f.pack(fill="x", padx=40, pady=5)
        ctk.CTkLabel(f, text=label, font=("Arial", 12, "bold")).pack(anchor="w")
        e = ctk.CTkEntry(f, width=380, height=35, show=show)
        e.pack(pady=(2, 8)); e.insert(0, valor)
        return e

    def acao_copiar_id(self):
        self.clipboard_clear()
        self.clipboard_append(self.id_str)
        self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def salvar(self):
        new_config = {
            "nome_usuario": self.e_nome.get(),
            "servidor": self.e_servidor.get(),
            "banco": self.e_banco.get(),
            "usuario_bd": self.e_user.get(),
            "senha_bd": self.e_senha.get()
        }
        self.update_callback(new_config)
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "Configurações atualizadas com sucesso!")
        self.destroy()

class App(ctk.CTkFrame):
    def __init__(self, parent, db_config):
        super().__init__(parent, fg_color="#FFFFFF", corner_radius=0)

        # Configurações DB via CLI
        self.cfg_server = db_config["servidor"]
        self.cfg_db = db_config["banco"]
        self.cfg_user = db_config["usuario_bd"]
        self.cfg_pass = db_config["senha_bd"]

        # Configurações Iniciais da Tabela
        self.tabela_config = {
            "tabela": "crcli",
            "col_id": "crclicod",
            "col_nome": "crclides",
            "col_cep": "cmcepcod",
            "col_rua": "crcliend",
            "col_logradouro": "crclilog",
            "col_bairro": "crclibai",
            "col_uf": "SP"
        }

        self.mostrar_somente_invalidos = False
        self.ceps_invalidos_cache = set()
        self.lista_updates = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.ultima_query_bruta = "" # Auditoria SQL

        # --- Top Frame (Menu Superior Expandido) Cor Chumbo ---
        self.top_frame = ctk.CTkFrame(self, fg_color="#D1D5DB", corner_radius=0)
        self.top_frame.grid(row=0, column=0, padx=0, pady=0, sticky="ew")
        self.top_frame.grid_columnconfigure(6, weight=1) # Espaçador

        # --- NOVO: Linha de Título Superior ---
        self.title_bar = ctk.CTkFrame(self.top_frame, fg_color="transparent", height=35)
        self.title_bar.grid(row=0, column=0, columnspan=7, sticky="ew", padx=20, pady=(5, 0))
        
        self.lbl_titulo = ctk.CTkLabel(self.title_bar, text="📍 ATUALIZAÇÃO DE CEP", font=("Arial", 16, "bold"), text_color="#1565C0")
        self.lbl_titulo.pack(side="left")

        # Componentes do Top Frame (Agora na Row 1)
        self.btn_config = ctk.CTkButton(self.top_frame, text="⚙️ Mapear", width=100, fg_color="#455A64", hover_color="#37474F", command=self.abrir_config)
        self.btn_config.grid(row=1, column=0, padx=10, pady=(5, 15))

        self.search_entry = ctk.CTkEntry(self.top_frame, placeholder_text="Buscar cliente (Deixe vazio para TODOS)...", width=260)
        self.search_entry.grid(row=1, column=1, padx=10, pady=(5, 15))

        self.btn_listar = ctk.CTkButton(self.top_frame, text="🔍 Simular Atualização", fg_color="#1E88E5", hover_color="#1565C0", command=self.simular_click)
        self.btn_listar.grid(row=1, column=2, padx=10, pady=(5, 15))

        self.btn_filtrar = ctk.CTkButton(self.top_frame, text="👁️ Mostrar: TODOS", width=130, fg_color="#7B1FA2", hover_color="#4A148C", command=self.alternar_filtro_click)
        self.btn_filtrar.grid(row=1, column=3, padx=10, pady=(5, 15))

        self.btn_gravar = ctk.CTkButton(self.top_frame, text="💾 Gravar no BD", width=120, fg_color="#43A047", hover_color="#2E7D32", state="disabled", command=self.gravar_click)
        self.btn_gravar.grid(row=1, column=4, padx=10, pady=(5, 15))
        
        self.lbl_stats = ctk.CTkLabel(self.top_frame, text="Simulação pendente...", font=ctk.CTkFont(slant="italic"))
        self.lbl_stats.grid(row=1, column=5, padx=10, pady=(5, 15))


        # Botão Fechar de topo removido para usar apenas o do rodapé

        # --- Frames Interiores (Grid e Log) ---
        self.main_content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_content_frame.grid(row=1, column=0, padx=20, pady=(10, 10), sticky="nsew")
        self.main_content_frame.grid_rowconfigure(0, weight=3)
        self.main_content_frame.grid_rowconfigure(1, weight=1)
        self.main_content_frame.grid_columnconfigure(0, weight=1)

        # Treeview
        self.table_frame = ctk.CTkFrame(self.main_content_frame, corner_radius=10)
        self.table_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)

        columns = ("id", "nome", "cidade", "cep_antigo", "cep_novo", "rua")
        style = ttk.Style()
        # Ajuste de cores da grade (Treeview) para o tema claro
        style.configure("Treeview", background="#FFFFFF", foreground="black", rowheight=30, fieldbackground="#FFFFFF", borderwidth=0)
        style.map('Treeview', background=[('selected', '#1E88E5')])
        style.configure("Treeview.Heading", background="#E0E0E0", foreground="black", borderwidth=0, font=('Arial', 10, 'bold'))

        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings", selectmode="browse")
        self.tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        self.tree.tag_configure('correto', background='#2E7D32', foreground='white') 
        self.tree.tag_configure('invalido', background='#C62828', foreground='white') 
        self.tree.tag_configure('desconhecido', background='#EF6C00', foreground='white') 

        self.tree.heading("id", text="ID")
        self.tree.heading("nome", text="Nome")
        self.tree.heading("cidade", text="Cidade")
        self.tree.heading("cep_antigo", text="CEP Antigo")
        self.tree.heading("cep_novo", text="CEP Novo")
        self.tree.heading("rua", text="Rua")

        self.tree.column("id", width=60, anchor="center")
        self.tree.column("nome", width=180)
        self.tree.column("cidade", width=120)
        self.tree.column("cep_antigo", width=90, anchor="center")
        self.tree.column("cep_novo", width=100, anchor="center")
        self.tree.column("rua", width=250)

        scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # Log
        self.log_frame = ctk.CTkFrame(self.main_content_frame, corner_radius=10)
        self.log_frame.grid(row=1, column=0, sticky="nsew")
        self.log_frame.grid_rowconfigure(0, weight=1)
        self.log_frame.grid_columnconfigure(0, weight=1)

        self.log_text = ctk.CTkTextbox(self.log_frame, wrap="word", font=('Consolas', 10))
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.log_text.configure(state="disabled")

        # --- Barra de Status Inferior ---
        self.bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_frame.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="ew")
        
        # Botão de retorno único
        try:
            img_fechar_pil = Image.open(os.path.join(os.path.dirname(__file__), "btn_fechar.png"))
            self.img_fechar_at = ctk.CTkImage(img_fechar_pil, size=(22, 22))
        except:
            self.img_fechar_at = None

        self.btn_sair = ctk.CTkButton(
            self.bottom_frame, 
            text="Fechar Tela", 
            image=self.img_fechar_at, 
            compound="left", 
            width=130, 
            height=32, 
            fg_color="transparent", 
            text_color="black", 
            hover_color="#E0E0E0", 
            border_width=2,
            border_color="black",
            command=self.fechar_tela
        )
        self.btn_sair.pack(side="left", padx=(0, 5))

        # Botão de Auditoria SQL (?) - Padronizado
        self.btn_debug_sql = ctk.CTkButton(self.bottom_frame, text="?", width=28, height=28, corner_radius=14, fg_color="#C62828", hover_color="#B12020", text_color="white", font=("Arial", 12, "bold"), command=self.mostrar_sql_bruto)
        self.btn_debug_sql.pack(side="left", padx=(0, 15))
        ToolTip(self.btn_debug_sql, "VER COMANDO SQL BRUTO (Audit)")

        self.lbl_status = ctk.CTkLabel(self.bottom_frame, text="💡 Status: Pronto. Preencha a Busca e clique em Simular.", font=ctk.CTkFont(size=12))
        self.lbl_status.pack(side="left")

        # Botões de Exportação CEP
        self.btn_export_pdf = ctk.CTkButton(self.bottom_frame, text="📄 Exportar PDF", width=120, height=32, fg_color="#E53935", hover_color="#B71C1C", command=self.exportar_pdf)
        self.btn_export_pdf.pack(side="right", padx=5)
        
        self.btn_export_excel = ctk.CTkButton(self.bottom_frame, text="📊 Exportar Excel", width=120, height=32, fg_color="#43A047", hover_color="#2E7D32", command=self.exportar_excel)
        self.btn_export_excel.pack(side="right", padx=5)

        self.frame_id = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        self.frame_id.pack(side="right")
        self.lbl_tela_id = ctk.CTkLabel(self.frame_id, text="Tela: ATUALIZA_CEP", font=ctk.CTkFont(size=11, weight="bold"), text_color="gray")
        self.lbl_tela_id.pack(side="left")
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=20, height=20, fg_color="transparent", hover_color="#E0E0E0", text_color="black", command=lambda: self.copiar_id("ATUALIZA_CEP"))
        self.btn_copy_id.pack(side="left", padx=5)
        ToolTip(self.btn_copy_id, "COPIAR NOME TELA")

    def abrir_config(self):
        TabelaConfigWindow(self, self.tabela_config, self.salvar_config)

    def salvar_config(self, config):
        self.tabela_config = config
        self.log("⚙️ Mapeamento de Tabela atualizado!")

    def copiar_id(self, texto):
        self.clipboard_clear(); self.clipboard_append(texto); self.update()
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def get_current_query(self):
        """Monta a query SQL dinamicamente com base nas configurações e busca atual."""
        c = self.tabela_config
        tabela = c["tabela"]
        c_id = c["col_id"]
        c_nome = c["col_nome"]
        c_cep = c["col_cep"]
        c_rua = c["col_rua"]
        c_logradouro = c["col_logradouro"]
        
        busca = self.search_entry.get().strip()
        where_clause = ""
        if busca:
            where_clause = f" WHERE {c_nome} LIKE '%{busca}%' OR {c_id} LIKE '%{busca}%'"

        return f"""SELECT 
    c.{c_id}, c.{c_nome}, c.{c_cep}, c.{c_rua}, 
    c.{c['col_bairro']}, c.{c_logradouro},
    m.CMUFCod as UF_Dinamica, m.CMCEPDes as Cidade_Dinamica, c.CMCEPCod  
FROM {tabela} c
LEFT JOIN CMCEP m ON c.CMCEPCod = m.CMCEPCod
{where_clause}"""

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL Pronto/Executado seguindo o padrão premium visual."""
        query = self.get_current_query()
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL (ATUALIZA_CEP)")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", query)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def copiar_sql_debug(self, sql, window):
        self.clipboard_clear(); self.clipboard_append(sql); self.update()
        messagebox.showinfo("Clipboard", "SQL Copiado!", parent=window)
        window.destroy()

    def log(self, mensagem):
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, mensagem + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")
        self.update()

    def limpar_texto(self, texto):
        if not texto: return ""
        return unidecode(str(texto)).strip().upper()

    def buscar_cep_por_endereco(self, uf, cidade, logradouro):
        uf = uf.strip().upper()
        cidade = cidade.strip()
        logradouro = logradouro.strip()
        if len(logradouro) < 3: return []
        cidade_encoded = urllib.parse.quote(cidade)
        logradouro_encoded = urllib.parse.quote(logradouro)
        url_viacep = f"https://viacep.com.br/ws/{uf}/{cidade_encoded}/{logradouro_encoded}/json/"
        try:
            response = requests.get(url_viacep, timeout=3)
            if response.status_code == 200:
                dados = response.json()
                if isinstance(dados, list) and len(dados) > 0:
                    return dados
        except: pass
        return []

    def simular_click(self):
        server = self.cfg_server
        database = self.cfg_db
        username = self.cfg_user
        password = self.cfg_pass

        c = self.tabela_config
        tabela = c["tabela"]
        c_id = c["col_id"]
        c_nome = c["col_nome"]
        c_cep = c["col_cep"]
        c_rua = c["col_rua"]
        c_logradouro = c["col_logradouro"]
        c_bairro = c["col_bairro"]
        valor_uf = c["col_uf"]
        
        busca = self.search_entry.get().strip()

        for item in self.tree.get_children(): self.tree.delete(item)

        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state="disabled")

        self.log("🚀 Iniciando Simulação...")
        self.lbl_status.configure(text="⏳ Conectando ao SQL Server...")
        self.update()

        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
        try:
            conn = pyodbc.connect(conn_str)
        except:
            try:
                conn_str_alt = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
                conn = pyodbc.connect(conn_str_alt)
            except Exception as e:
                messagebox.showerror("Erro de Conexão", "Não foi possível conectar ao Banco de Dados com as credenciais fornecidas.\n\nVerifique os dados de acesso (Servidor, Banco, Usuário ou Senha) nas configurações e tente novamente.")
                # Abre a tela ConfigWindow (CFG_ACESSO)
                ConfigWindow(self, self.config, lambda c: [setattr(self, 'config', c)])
                return

        cursor = conn.cursor()
        query = self.get_current_query()
        self.ultima_query_bruta = query # Rastro Técnico SQL (?)
        
        try:
            cursor.execute(query)
            clientes = cursor.fetchall()
        except Exception as e:
             messagebox.showerror("Erro de Query", f"Não foi possível ler a tabela '{tabela}' com as colunas definidas.\n\nDetalhe: {e}")
             cursor.close()
             conn.close()
             return
            
        self.log(f"📊 {len(clientes)} clientes filtrados no banco.")
        
        if len(clientes) == 0:
            self.lbl_status.configure(text=f"Nenhum cliente encontrado com a busca '{busca}'")
            return

        invalidos = 0
        total_analisado = 0
        self.lista_updates = [] 

        for cli in clientes:
            total_analisado += 1
            time.sleep(0.05) 
            self.lbl_status.configure(text=f"📊 Analisando Cliente {total_analisado} de {len(clientes)} ...")

            cod = str(cli[0])
            nome = str(cli[1])
            campo_cep = str(cli[2]).strip() if cli[2] else ""
            campo_logradouro = str(cli[5]).strip() if cli[5] else "" 
            cep_atual = (campo_cep + campo_logradouro).replace("-", "").replace(".", "").strip()
            # Máscara xxxxx-xxx para visualização na Treeview
            cep_antigo_formatado = f"{cep_atual[:5]}-{cep_atual[5:]}" if len(cep_atual) == 8 else cep_atual
            rua_orig = str(cli[3]) if cli[3] else "" 
            bairro_orig = str(cli[4]) if cli[4] else ""
            
            valor_uf = str(cli[6]).strip() if cli[6] else "SP" 
            cidade_busca = str(cli[7]).strip() if cli[7] else "Ipuã"
            chave_cidade = cli[8] 

            self.log(f"🔎 Cliente {cod} ({nome}) -> CEP Atual: '{cep_atual}'")

            if cep_atual in self.ceps_invalidos_cache:
                 cep_valido = False
            else:
                cep_valido = False
                if len(cep_atual) == 8:
                    try:
                        r_cep = requests.get(f"https://viacep.com.br/ws/{cep_atual}/json/", timeout=2)
                        if r_cep.status_code == 200 and "erro" not in r_cep.text:
                            cep_valido = True 
                        else:
                             self.ceps_invalidos_cache.add(cep_atual)
                    except: pass

            if cep_valido:
                if not self.mostrar_somente_invalidos:
                    self.tree.insert("", tk.END, values=(cod, nome, cidade_busca, cep_antigo_formatado, "✓ OK", rua_orig), tags=('correto',))
                continue 

            if campo_logradouro.isdigit() or campo_logradouro == "000":
                 logradouro_completo = rua_orig.strip()
            else:
                 logradouro_completo = f"{campo_logradouro} {rua_orig}".strip() 
            
            if not logradouro_completo.replace("None", "").strip():
                continue

            resultados = self.buscar_cep_por_endereco(valor_uf, cidade_busca, logradouro_completo)
            if not resultados:
                resultados = self.buscar_cep_por_endereco(valor_uf, cidade_busca, rua_orig)

            if resultados:
                cep_encontrado = None
                bairro_local_limpo = self.limpar_texto(bairro_orig)
                for res in resultados:
                    if self.limpar_texto(res.get("bairro", "")) == bairro_local_limpo:
                        cep_encontrado = res.get("cep")
                        res_rua = res.get("logradouro", logradouro_completo)
                        res_bairro = res.get("bairro", bairro_orig)
                        break

                if not cep_encontrado and len(resultados) == 1:
                    cep_encontrado = resultados[0].get("cep")
                    res_rua = resultados[0].get("logradouro", logradouro_completo)
                    res_bairro = resultados[0].get("bairro", bairro_orig)

                if cep_encontrado:
                     invalidos += 1
                     self.tree.insert("", tk.END, values=(cod, nome, cidade_busca, cep_antigo_formatado, cep_encontrado, logradouro_completo), tags=('invalido',))
                     self.tree.see(self.tree.get_children()[-1])
                     self.update() 
                     
                     self.lista_updates.append({
                         "cod": cod,
                         "cep": cep_encontrado.replace("-", ""),
                         "chave_cidade": chave_cidade, 
                         "nome_cidade": cidade_busca,
                         "nome_uf": valor_uf,
                         "rua": res_rua,
                         "bairro": res_bairro
                     })
                else:
                    if not self.mostrar_somente_invalidos:
                         self.tree.insert("", tk.END, values=(cod, nome, cidade_busca, cep_antigo_formatado, "⚠️ Sem CEP", logradouro_completo), tags=('desconhecido',))

        cursor.close()
        conn.close()

        self.btn_gravar.configure(state="normal" if self.lista_updates else "disabled")
        self.lbl_stats.configure(text=f"✅ {invalidos} Clientes precisam de correção.")
        self.lbl_status.configure(text=f"💡 Verifique a lista. {invalidos} CEPs para atualizar. Clique em 'Gravar' se concordar.")

    def alternar_filtro_click(self):
        self.mostrar_somente_invalidos = not self.mostrar_somente_invalidos
        if self.mostrar_somente_invalidos:
            self.btn_filtrar.configure(text="👁️ Mostrar: SÓ ERROS", fg_color="#C62828")
            for item in self.tree.get_children():
                tags = self.tree.item(item, "tags")
                if 'correto' in tags or 'desconhecido' in tags:
                    self.tree.detach(item) 
        else:
            self.btn_filtrar.configure(text="👁️ Mostrar: TODOS", fg_color="#7B1FA2")
            messagebox.showinfo("Filtro", "Para voltar a ver os corretos na busca atual, clique em 'Simular Atualização' novamente.")

    def gravar_click(self):
        if not hasattr(self, 'lista_updates') or not self.lista_updates:
            return

        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.cfg_server};DATABASE={self.cfg_db};UID={self.cfg_user};PWD={self.cfg_pass};"
        c = self.tabela_config
        
        try:
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            
            sucessos_cliente = 0
            insercoes_cmcep = 0
            insercoes_cmcepl = 0

            for up in self.lista_updates:
                cod_cliente = up["cod"]
                cep_limpo = up["cep"] 

                cm_cep_cod = int(cep_limpo[:5]) 
                cm_cep_log = str(cep_limpo[5:]).strip() 

                nome_cidade = up["nome_cidade"]
                nome_uf = up["nome_uf"]
                nome_rua = up["rua"]
                nome_bairro = up["bairro"]

                cursor.execute("SELECT CMCEPCod FROM CMCEP WHERE CMCEPCod = ?", cm_cep_cod)
                if not cursor.fetchone():
                    cursor.execute("""
                        INSERT INTO CMCEP (CMCEPCod, CMUFCod, CMCEPDes, CMCEPCUsu, CMCEPCHor, CMCEPCPrg)
                        VALUES (?, ?, ?, 'IA', '00:00:00', 'UPDATE')
                    """, cm_cep_cod, nome_uf, nome_cidade)
                    insercoes_cmcep += 1
                
                cursor.execute("SELECT CmCepLCod FROM cmcepl WHERE CmCepLCod = ? AND CmCepLLog = ?", cm_cep_cod, cm_cep_log)
                if not cursor.fetchone():
                     cursor.execute("""
                        INSERT INTO cmcepl (CmCepLCod, CmCepLLog, CmCepLDes, CmCepLBai, CmCepLCid, CmCepLUf, CmCepLPat)
                        VALUES (?, ?, ?, ?, ?, ?, '')
                     """, cm_cep_cod, cm_cep_log, nome_rua, nome_bairro, nome_cidade, nome_uf)
                     insercoes_cmcepl += 1

                cursor.execute(f"UPDATE {c['tabela']} SET {c['col_cep']} = ?, {c['col_logradouro']} = ? WHERE {c['col_id']} = ?", cm_cep_cod, cm_cep_log, cod_cliente)
                sucessos_cliente += 1

            conn.commit()
            cursor.close()
            conn.close()

            messagebox.showinfo("Sucesso", f"📊 Atualização Concluída!\n• {sucessos_cliente} Clientes Atualizados ✅\n• {insercoes_cmcep} Cidades (CMCEP) 🏙️\n• {insercoes_cmcepl} Ruas (cmcepl) 🛣️")
            self.lbl_status.configure(text="✅ Tudo atualizado! Simule novamente para conferir.")
            self.btn_gravar.configure(state="disabled")
            for item in self.tree.get_children(): self.tree.delete(item)

        except Exception as e:
            messagebox.showerror("Erro SQL", f"Erro ao gravar:\n{e}")

    def exportar_excel(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        columns = self.tree["columns"]; cols_titles = [self.tree.heading(c)["text"] for c in columns]; data = [self.tree.item(k)['values'] for k in items]
        try:
             import pandas as pd, tempfile, os
             df = pd.DataFrame(data, columns=cols_titles)
             filepath = os.path.join(tempfile.gettempdir(), "Relatorio_CEPs.xlsx")
             
             writer = pd.ExcelWriter(filepath, engine='openpyxl')
             df.to_excel(writer, index=False, sheet_name='Sheet1')
             worksheet = writer.sheets['Sheet1']
             for i, col in enumerate(df.columns):
                  max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                  from openpyxl.utils import get_column_letter
                  worksheet.column_dimensions[get_column_letter(i+1)].width = max_len
             writer.close()
             
             os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro Excel", str(e))

    def exportar_pdf(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        columns = self.tree["columns"]; cols_titles = [self.tree.heading(c)["text"] for c in columns]; data = [self.tree.item(k)['values'] for k in items]
        try:
             import tempfile, os
             filepath = os.path.join(tempfile.gettempdir(), "Relatorio_CEPs.pdf")
             import sys, subprocess
             try: import fpdf
             except: subprocess.run([sys.executable, "-m", "pip", "install", "fpdf"], capture_output=True); import fpdf
             pdf = fpdf.FPDF(); pdf.add_page(); pdf.set_font("Arial", 'B', 12); pdf.cell(190, 10, "RELATORIO DE CEPs", 0, 1, 'C'); pdf.ln(5)
             pdf.set_font("Arial", 'B', 8); num_cols = len(cols_titles); w_col = 190 / num_cols if num_cols > 0 else 30
             for c in cols_titles: pdf.cell(w_col, 8, str(c), 1, 0, 'C')
             pdf.ln(8); pdf.set_font("Arial", '', 7)
             for r in data:
                  for val in r: pdf.cell(w_col, 7, str(val)[:20], 1, 0, 'C')
                  pdf.ln(7)
             pdf.ln(5); pdf.output(filepath); os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro PDF", str(e))

    def fechar_tela(self):
         if hasattr(self, 'hub'):
              self.hub.fechar_modulo_atual()
         else:
              self.destroy()


class AnaliseVendasWindow(ctk.CTkFrame):
    def __init__(self, parent, config):
        super().__init__(parent, fg_color="#FFFFFF", corner_radius=0)
        self.config = config
        self.ultima_query_bruta = "" # Auditoria SQL
        
        # 1. Barra de Identificação de Tela
        self.bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_bar.pack(side="bottom", fill="x", padx=20, pady=(0, 10))
        
        self.frame_id = ctk.CTkFrame(self.bottom_bar, fg_color="transparent")
        self.frame_id.pack(side="right")
        self.lbl_tela_id = ctk.CTkLabel(self.frame_id, text="Tela: ANALISE_VENDAS", font=ctk.CTkFont(size=11, weight="bold"), text_color="gray")
        self.lbl_tela_id.pack(side="left")
        self.id_str = "ANALISE_VENDAS"
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=20, height=20, fg_color="transparent", hover_color="#E0E0E0", text_color="black", command=self.acao_copiar_id)
        self.btn_copy_id.pack(side="left", padx=5)
        ToolTip(self.btn_copy_id, "COPIAR NOME TELA")

        # 2. Barra de Botões (Rodapé)
        self.buttons_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.buttons_frame.pack(side="bottom", fill="x", padx=30, pady=(15, 5))
        
        try:
            img_fechar_pil = Image.open(os.path.join(os.path.dirname(__file__), "btn_fechar.png"))
            self.img_fechar = ctk.CTkImage(img_fechar_pil, size=(22, 22))
        except:
            self.img_fechar = None

        self.btn_sair = ctk.CTkButton(
            self.buttons_frame, 
            text="Fechar Tela", 
            image=self.img_fechar, 
            compound="left", 
            width=140, 
            height=32, 
            fg_color="transparent", 
            text_color="black", 
            hover_color="#E0E0E0", 
            border_width=2,
            border_color="black",
            command=self.fechar_tela
        )
        self.btn_sair.pack(side="left", padx=(0, 5))

        self.btn_debug_sql = ctk.CTkButton(self.buttons_frame, text="?", width=28, height=28, corner_radius=14, fg_color="#C62828", hover_color="#B12020", text_color="white", font=("Arial", 12, "bold"), command=self.mostrar_sql_bruto)
        self.btn_debug_sql.pack(side="left", padx=(0, 15))
        ToolTip(self.btn_debug_sql, "VER COMANDO SQL BRUTO (Audit)")

        # Exportações
        self.btn_export_pdf = ctk.CTkButton(self.buttons_frame, text="📄 Exportar PDF", width=140, height=32, fg_color="#E53935", hover_color="#B71C1C", command=self.exportar_pdf)
        self.btn_export_pdf.pack(side="right", padx=5)
        
        self.btn_export_excel = ctk.CTkButton(self.buttons_frame, text="📊 Exportar Excel", width=140, height=32, fg_color="#43A047", hover_color="#2E7D32", command=self.exportar_excel)
        self.btn_export_excel.pack(side="right", padx=5)

        # 3. Topo: Controles (Cabeçalho Chumbo Ocupando Extensão)
        self.top_frame = ctk.CTkFrame(self, fg_color="#D1D5DB", corner_radius=0)
        self.top_frame.pack(side="top", fill="x", padx=0, pady=0)      

        # --- NOVO: Linha de Título Superior ---
        self.title_bar = ctk.CTkFrame(self.top_frame, fg_color="transparent", height=35)
        self.title_bar.pack(side="top", fill="x", padx=20, pady=(5, 0))
        
        self.lbl_titulo = ctk.CTkLabel(self.title_bar, text="📊 ANÁLISE DE VENDAS POR HORA", font=("Arial", 16, "bold"), text_color="#1565C0")
        self.lbl_titulo.pack(side="left")

        # Sub-container interno para margens (Transparente para manter o fundo chumbo)
        self.top_grid_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.top_grid_frame.pack(side="top", fill="x", padx=10, pady=(5, 10))

        self.top_grid_frame.grid_columnconfigure((1, 3, 5), weight=1)

        # LINHA 0
        ctk.CTkLabel(self.top_grid_frame, text="Ano:").grid(row=0, column=0, padx=(15, 5), pady=8, sticky="e")
        self.combo_ano = ctk.CTkComboBox(self.top_grid_frame, values=["Todos"], width=100)
        self.combo_ano.grid(row=0, column=1, padx=5, pady=8, sticky="ew")
        self.combo_ano.set("Todos")

        ctk.CTkLabel(self.top_grid_frame, text="Mês:").grid(row=0, column=2, padx=(15, 5), pady=8, sticky="e")
        self.meses = ["Todos", "01 - Janeiro", "02 - Fevereiro", "03 - Março", "04 - Abril", "05 - Maio", "06 - Junho", "07 - Julho", "08 - Agosto", "09 - Setembro", "10 - Outubro", "11 - Novembro", "12 - Dezembro"]
        self.combo_mes = ctk.CTkComboBox(self.top_grid_frame, values=self.meses, width=140)
        self.combo_mes.grid(row=0, column=3, padx=5, pady=8, sticky="ew")
        self.combo_mes.set("Todos")

        ctk.CTkLabel(self.top_grid_frame, text="Visão Especial:").grid(row=0, column=4, padx=(15, 5), pady=8, sticky="e")
        self.combo_visao = ctk.CTkComboBox(self.top_grid_frame, values=["Padrão", "Horário", "Dias da Semana"], width=190)
        self.combo_visao.grid(row=0, column=5, padx=5, pady=8, sticky="ew")
        self.combo_visao.set("Padrão")

        # LINHA 1 - Botão Atualizar Gráficos alinhado abaixo de Ano
        self.btn_gerar = ctk.CTkButton(self.top_grid_frame, text="Atualizar Gráfico", command=self.gerar_grafico, fg_color="#1E88E5", hover_color="#1565C0", width=130)
        self.btn_gerar.grid(row=1, column=1, padx=5, pady=8, sticky="ew")

        ctk.CTkLabel(self.top_grid_frame, text="Métrica:").grid(row=1, column=2, padx=(15, 5), pady=8, sticky="e")
        self.combo_metrica = ctk.CTkComboBox(self.top_grid_frame, values=["Quantidade de Vendas", "Valor Total (R$)"], width=200)
        self.combo_metrica.grid(row=1, column=3, padx=5, pady=8, sticky="ew")
        self.combo_metrica.set("Valor Total (R$)")

        self.chart_frame = ctk.CTkFrame(self, fg_color="white")
        self.chart_frame.pack(side="top", fill="both", expand=True, padx=20, pady=(0, 10))
        
        self.fig = Figure(figsize=(8, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.chart_frame)
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=True)
        
        self.df_raw = None
        self.df_grouped = None
        self.after(100, self.carregar_dados)

    def acao_copiar_id(self):
        """Copia apenas o identificador da tela para o clipboard."""
        self.clipboard_clear()
        self.clipboard_append(self.id_str)
        self.update()
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL dinâmico completo seguindo o padrão premium visual."""
        query = self.get_current_query()
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", query)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def copiar_sql_debug(self, sql, window):
        self.clipboard_clear(); self.clipboard_append(sql); self.update()
        messagebox.showinfo("Clipboard", "SQL Copiado!", parent=window)
        window.destroy()

    def get_current_query(self):
        """Monta a query SQL refletindo os filtros reais de Ano e Mês da tela."""
        ano_sel = self.combo_ano.get()
        mes_sel = self.combo_mes.get()
        where_extra = ""
        
        if ano_sel != "Todos":
            where_extra += f" AND YEAR(CRMovDta) = {ano_sel}"
        if mes_sel != "Todos":
            mes_idx = self.combo_mes.get().split(" - ")[0]
            where_extra += f" AND MONTH(CRMovDta) = {mes_idx}"
            
        return f"SELECT CRMovDta, CRMov2CHor, CRMov2VlrO FROM crmov2 WHERE CRMovDta IS NOT NULL AND CRMov2Flag IN ('A', 'F'){where_extra} ORDER BY CRMovDta"

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: ANALISEVENDAS ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):
        try:
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config['servidor']};DATABASE={self.config['banco']};UID={self.config['usuario_bd']};PWD={self.config['senha_bd']}"
            conn = pyodbc.connect(conn_str)
            query = "SELECT CRMovDta, CRMov2CHor, CRMov2VlrO FROM crmov2 WHERE CRMovDta IS NOT NULL AND CRMov2Flag IN ('A', 'F')"
            self.ultima_query_bruta = query # Rastro Técnico SQL (?)
            self.df_raw = pd.read_sql(query, conn)
            conn.close()
            
            # Limpeza de dados básica
            self.df_raw['CRMovDta'] = pd.to_datetime(self.df_raw['CRMovDta'], errors='coerce')
            self.df_raw['CRMov2VlrO'] = pd.to_numeric(self.df_raw['CRMov2VlrO'], errors='coerce').fillna(0)
            
            # Extrair atributos principais e remover anos inválidos como 9999
            self.df_raw['Ano'] = self.df_raw['CRMovDta'].dt.year.astype('Int64')
            self.df_raw = self.df_raw[self.df_raw['Ano'] < 9999]
            
            self.df_raw['Mes_Int'] = self.df_raw['CRMovDta'].dt.month.astype('Int64')
            
            meses_dict = {1:"01 - Janeiro", 2:"02 - Fevereiro", 3:"03 - Março", 4:"04 - Abril", 5:"05 - Maio", 6:"06 - Junho", 7:"07 - Julho", 8:"08 - Agosto", 9:"09 - Setembro", 10:"10 - Outubro", 11:"11 - Novembro", 12:"12 - Dezembro"}
            self.df_raw['Mes_Extenso'] = self.df_raw['Mes_Int'].map(meses_dict)
            self.df_raw['Dia'] = self.df_raw['CRMovDta'].dt.day.astype('Int64')
            
            # Dias da Semana
            dias_dict = {0:"Segunda-feira", 1:"Terça-feira", 2:"Quarta-feira", 3:"Quinta-feira", 4:"Sexta-feira", 5:"Sábado", 6:"Domingo"}
            self.df_raw['Dia_Semana'] = self.df_raw['CRMovDta'].dt.dayofweek.map(dias_dict)
            
            dias_ordem = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]
            self.df_raw['Dia_Semana'] = pd.Categorical(self.df_raw['Dia_Semana'], categories=dias_ordem, ordered=True)
            
            # Hora (Tratando strings variadas)
            def extrair_hora(h):
                try: 
                    return int(str(h).replace(':', '')[:2])
                except: 
                    return None
            
            self.df_raw['Hora'] = self.df_raw['CRMov2CHor'].apply(extrair_hora)
            
            def extrair_faixa(h):
                try:
                    hora_int = int(str(h).replace(':', '')[:2])
                    if hora_int >= 24: return None
                    lim = (hora_int // 2) * 2
                    return f"{lim:02d}h às {lim+2:02d}h"
                except: return None
            
            self.df_raw['Faixa_Horario'] = self.df_raw['CRMov2CHor'].apply(extrair_faixa)
            
            # Preenche o ComboBox de Ano com os anos encontrados no banco (mais recente primeiro)
            anos_unicos = sorted(self.df_raw['Ano'].dropna().unique().tolist(), reverse=True)
            self.combo_ano.configure(values=["Todos"] + [str(a) for a in anos_unicos])
            
            self.gerar_grafico()
            
        except Exception as e:
            messagebox.showerror("Erro de Conexão", "Não foi possível conectar ao Banco de Dados com as credenciais fornecidas.\n\nVerifique os dados de acesso (Servidor, Banco, Usuário ou Senha) nas configurações e tente novamente.")
            # Abre a tela ConfigWindow (CFG_ACESSO) de forma autônoma
            ConfigWindow(self, self.config, lambda c: [setattr(self, 'config', c), self.carregar_dados()])

    def gerar_grafico(self):
        if self.df_raw is None or self.df_raw.empty:
            return
            
        ano_sel = self.combo_ano.get()
        mes_sel = self.combo_mes.get()
        visao = self.combo_visao.get()
        metrica = self.combo_metrica.get()
        
        df_valid = self.df_raw.copy()
        
        # Filtros Dinâmicos
        if ano_sel != "Todos":
            df_valid = df_valid[df_valid['Ano'] == int(ano_sel)]
        if mes_sel != "Todos":
            df_valid = df_valid[df_valid['Mes_Extenso'] == mes_sel]
            
        if df_valid.empty:
            self.ax.clear()
            self.ax.set_title("Nenhum dado encontrado para os filtros selecionados.")
            self.canvas.draw()
            return

        # Agrupamento base
        if visao == "Horário":
            grouper = 'Hora'
            agrupamento_str = "Faixa de Horário"
        elif visao == "Dias da Semana":
            grouper = 'Dia_Semana'
            agrupamento_str = "Dia da Semana"
        elif visao == "Dia da Semana + Horário":
            grouper = 'Dia_Horario'
            agrupamento_str = "Dia da Semana + Horário"
        else:
            if ano_sel != "Todos" and mes_sel == "Todos":
                grouper = 'Mes_Extenso'
                agrupamento_str = "Mês"
            elif ano_sel == "Todos" and mes_sel != "Todos":
                grouper = 'Ano'
                agrupamento_str = "Ano"
            elif ano_sel != "Todos" and mes_sel != "Todos":
                grouper = 'Dia'
                agrupamento_str = "Dia"
            else:
                grouper = 'Ano'
                agrupamento_str = "Ano"
            
        # Calcula dados
        if grouper not in df_valid.columns: return
        df_valid = df_valid.dropna(subset=[grouper])
        
        self.df_grouped = df_valid.groupby(grouper).agg(
            Qtd=('CRMovDta', 'size'),
            ValorTotal=('CRMov2VlrO', 'sum')
        ).reset_index()
        
        # Cria rótulo combinado para visão duo list
        if isinstance(grouper, list):
            self.df_grouped['Dia_Horario'] = self.df_grouped['Dia_Semana'].astype(str) + " - " + self.df_grouped['Hora'].map(lambda x: f"{int(x):02d}h")
            grouper_plot = 'Dia_Horario'
        else:
            grouper_plot = grouper
        
        if metrica == "Quantidade de Vendas":
            self.df_grouped['Valor'] = self.df_grouped['Qtd']
            ylabel = "Qtd. de Vendas"
        else:
            self.df_grouped['Valor'] = self.df_grouped['ValorTotal']
            ylabel = "Valor Total (R$)"
            
        # Pular se nulo
        if self.df_grouped.empty: return
        
        # Ordenar eixo X logicamente
        if grouper in ['Mes_Extenso', 'Dia_Semana', 'Dia_Horario']:
            self.df_grouped = self.df_grouped.sort_values(by=grouper)
        elif grouper == 'Ano':
            self.df_grouped = self.df_grouped.sort_values(by=grouper, ascending=False)
        else:
            self.df_grouped = self.df_grouped.sort_values(by=grouper)
        
        # Formata x para exibição (pandas period Mês não plota nativamente bem se for object string sem conversao)
        self.df_grouped[grouper_plot] = self.df_grouped[grouper_plot].astype(str)
        
        # Plot
        self.ax.clear()
        x_vals = self.df_grouped[grouper_plot]
        y_vals = self.df_grouped['Valor']
        
        # --- Lógica de Cores Dinâmicas (Degradê Vermelho/Verde) ---
        colors = []
        if not y_vals.empty:
            max_v = y_vals.max()
            min_v = y_vals.min()
            avg_v = y_vals.mean()
            
            for v in y_vals:
                if v > avg_v:
                    # ACIMA DA MÉDIA: Degradê Vermelho (Média -> Máximo)
                    # Da cor "Rosa Pastel" (#FFCDD2) para "Vermelho Intenso" (#B71C1C)
                    ratio = (v - avg_v) / (max_v - avg_v) if max_v != avg_v else 1.0
                    r = int(0xFF - (0xFF - 0xB7) * ratio)
                    g = int(0xCD - (0xCD - 0x1C) * ratio)
                    b = int(0xD2 - (0xD2 - 0x1C) * ratio)
                elif v < avg_v:
                    # ABAIXO DA MÉDIA: Degradê Verde (Mínimo -> Média)
                    # Da cor "Verde Escuro" (#1B5E20) para "Verde Pastel" (#C8E6C9)
                    ratio = (v - min_v) / (avg_v - min_v) if avg_v != min_v else 1.0
                    r = int(0x1B + (0xC8 - 0x1B) * ratio)
                    g = int(0x5E + (0xE6 - 0x5E) * ratio)
                    b = int(0x20 + (0xC9 - 0x20) * ratio)
                else:
                    # EXATAMENTE NA MÉDIA: Cinza Azulado Neutro
                    r, g, b = 0xCF, 0xD8, 0xDC
                
                colors.append(f'#{r:02x}{g:02x}{b:02x}')
        else:
            colors = '#1E88E5' # Fallback azul padrão
            
        bars = self.ax.bar(x_vals, y_vals, color=colors, edgecolor='#454545', linewidth=0.7)
        self.ax.set_title(f"{metrica} por {agrupamento_str}", fontdict={'fontsize': 14, 'fontweight': 'bold', 'color': '#1E3A8A'})
        self.ax.set_ylabel(ylabel, fontdict={'fontsize': 11, 'fontweight': 'bold'})
        self.ax.set_xlabel(agrupamento_str, fontdict={'fontsize': 11, 'fontweight': 'bold'})
        
        # Grid horizontal para facilitar leitura
        self.ax.yaxis.grid(True, linestyle='--', alpha=0.6, color='#E2E8F0')
        self.ax.set_axisbelow(True)
        
        # Remove bordas desnecessárias
        for spine in ['top', 'right']:
            self.ax.spines[spine].set_visible(False)
        
        # Criação do Tooltip Interativo ao passar o mouse
        cursor = mplcursors.cursor(bars, hover=True)
        
        @cursor.connect("add")
        def on_add(sel):
            idx = sel.index
            label_x = self.df_grouped.iloc[idx][grouper_plot]
            qtd = self.df_grouped.iloc[idx]['Qtd']
            valor = self.df_grouped.iloc[idx]['ValorTotal']
            
            # Formatação PT-BR
            texto = f"{label_x}\n----------------\n📦 Vendas: {int(qtd)}\n💰 Total: R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            sel.annotation.set_text(texto)
            sel.annotation.get_bbox_patch().set(boxstyle="round,pad=0.5", fc="white", alpha=0.95, ec="#1E88E5", lw=1.5)
            sel.annotation.set_fontsize(10)
        
        # Formatação
        if len(x_vals) > 10:
            self.ax.tick_params(axis='x', rotation=45, labelsize=9)
        else:
            self.ax.tick_params(axis='x', rotation=0, labelsize=10)
            
        if "Valor" in metrica:
            self.ax.get_yaxis().set_major_formatter(plt.FuncFormatter(lambda x, loc: f"R$ {x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')))
            
        self.fig.tight_layout()
        self.canvas.draw()
        
    def exportar_excel(self):
        if self.df_grouped is None or self.df_grouped.empty:
            messagebox.showwarning("Aviso", "Nenhum dado gerado para exportar.")
            return
        
        try:
            import tempfile, os, pandas as pd
            filepath = os.path.join(tempfile.gettempdir(), "Analise_Vendas.xlsx")
            
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            self.df_grouped.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(self.df_grouped.columns):
                 max_len = max(self.df_grouped[col].astype(str).map(len).max(), len(str(col))) + 2
                 from openpyxl.utils import get_column_letter
                 worksheet.column_dimensions[get_column_letter(i+1)].width = max_len
            writer.close()
            
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def exportar_pdf(self):
        if self.df_grouped is None or self.df_grouped.empty:
            messagebox.showwarning("Aviso", "Nenhum gráfico gerado para exportar.")
            return
        
        try:
            import tempfile, os
            filepath = os.path.join(tempfile.gettempdir(), "Grafico_Vendas.pdf")
            self.fig.savefig(filepath, format="pdf", bbox_inches="tight")
            os.startfile(filepath)  # Abre o PDF automaticamente
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def fechar_tela(self):
         if hasattr(self, 'hub'):
              self.hub.fechar_modulo_atual()
         else:
              self.destroy()

class MainHub(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SIL_IA_WIN")
        self.geometry("1000x650")
        self.after(100, lambda: self.state("zoomed"))
        ctk.set_appearance_mode("Light")
        
        self.config = {"nome_usuario": "Operador", "servidor": r"servidor\sqlexpress", "banco": "msinfor", "usuario_bd": "sa", "senha_bd": "Mabelu2011"}
        self.carregar_config()
        self.grid_rowconfigure(0, weight=1); self.grid_columnconfigure(1, weight=1)

        # Sidebar (Faixa Azul Escura conforme imagem)
        self.sidebar_frame = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color="#0D47A1") # Azul Marinho Rico
        self.sidebar_frame.grid(row=0, column=0, rowspan=2, sticky="nsew")
        
        # Frame centralizado no topo da barra
        self.sb_header = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        self.sb_header.pack(fill="x", padx=15, pady=(40, 20))
        
        logo_path = os.path.join(os.path.dirname(__file__), "logo_msinfor.jpg")
        if os.path.exists(logo_path):
            try:
                logo_img = Image.open(logo_path)
                # Logotipo maior
                self.logo_image = ctk.CTkImage(light_image=logo_img, dark_image=logo_img, size=(170, 170))
                self.logo_label = ctk.CTkLabel(self.sb_header, text="", image=self.logo_image)
                self.logo_label.pack(anchor="center") # Centralizado
            except Exception as e: pass

        ctk.CTkFrame(self.sidebar_frame, height=2, fg_color="#1565C0").pack(fill="x", padx=15, pady=(5, 15))

        # Botão Sair do Sistema na Sidebar (Embaixo de tudo)
        self.btn_sair = ctk.CTkButton(self.sidebar_frame, text="❌ Sair do Sistema", font=("Arial", 14, "bold"), command=self.confirmar_saida, fg_color="#C62828", hover_color="#B12020")
        self.btn_sair.pack(side="bottom", fill="x", padx=15, pady=(0, 20)) # pady=(top, bottom)

        # Botão Configurações na Sidebar (Acima do Sair)
        self.btn_config = ctk.CTkButton(self.sidebar_frame, text="⚙️ Configurações", font=("Arial", 14, "bold"), command=self.abrir_configuracoes, fg_color="#455A64", hover_color="#37474F")
        self.btn_config.pack(side="bottom", fill="x", padx=15, pady=(20, 10))

        # Main Area
        self.main_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=0)
        self.main_frame.grid(row=0, column=1, sticky="nsew")

        # --- CABEÇALHO GLOBAL DO HUB (NOVO) ---
        self.top_header_hub = ctk.CTkFrame(self.main_frame, fg_color="#F1F5F9", height=50, corner_radius=0)
        # Inicialmente oculto, aparece apenas no Módulo Posto
        self.top_header_hub.pack_forget()
        
        self.posto_frame_hub = ctk.CTkFrame(self.top_header_hub, fg_color="transparent")
        self.posto_frame_hub.pack(side="left", padx=20, pady=10)
        
        ctk.CTkLabel(self.posto_frame_hub, text="POSTO ATUAL:", font=("Arial", 15, "bold"), text_color="#1E293B").pack(side="left", padx=5)
        self.combo_posto_hub = ctk.CTkComboBox(self.posto_frame_hub, width=320, font=("Arial", 15, "bold"), state="readonly", command=self.on_posto_change_hub)
        self.combo_posto_hub.pack(side="left", padx=5)
        
        self.lbl_info_data_hub = ctk.CTkLabel(self.posto_frame_hub, text="", font=("Arial", 15, "bold"), text_color="#1565C0")
        self.lbl_info_data_hub.pack(side="left", padx=20)
        
        self.lbl_info_per_hub = ctk.CTkLabel(self.posto_frame_hub, text="", font=("Arial", 15, "bold"), text_color="#1565C0")
        self.lbl_info_per_hub.pack(side="left", padx=5)

        self.postos_map_hub = {}
        self.after(500, self.carregar_postos_hub)

        # Frame para os módulos (Canto Superior Esquerdo)
        self.modules_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.modules_frame.pack(fill="both", expand=True, padx=30, pady=30)
        
        # --- NOVO: Rodapé Global (Movido para cá) ---
        self.footer_frame = ctk.CTkFrame(self, height=35, fg_color="#1E1E1E")
        self.footer_frame.grid(row=1, column=1, sticky="ew")
        
        self.frame_id = ctk.CTkFrame(self.footer_frame, fg_color="transparent")
        self.frame_id.pack(side="right", padx=15)
        self.lbl_tela_id = ctk.CTkLabel(self.frame_id, text="Tela: PRINCIPAL", font=ctk.CTkFont(size=11, weight="bold"), text_color="white")
        self.lbl_tela_id.pack(side="left")
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=20, height=20, fg_color="transparent", hover_color="#333333", text_color="white", command=lambda: self.copiar_id("PRINCIPAL"))
        self.btn_copy_id.pack(side="left", padx=5)
        
        self.lbl_footer = ctk.CTkLabel(self.footer_frame, text="", text_color="white", font=("Arial", 11))
        self.lbl_footer.pack(side="left", padx=15, expand=True)
        self.atualizar_rodape()
        self.after(2000, self.testar_conexao_banco) # Aumentado delay inicial
        self.modules_frame.grid_columnconfigure((0, 1, 2, 3, 4), weight=1) # 5 Colunas
        self.modules_frame.grid_rowconfigure(0, weight=1)
        
        btn_width = 240 # Ligeiramente reduzido para acomodar 5 colunas
        
        # --- CONTAINER 1º COLUNA ---
        self.col1_frame = ctk.CTkFrame(self.modules_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col1_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        ctk.CTkLabel(self.col1_frame, text="1º Coluna", font=("Arial", 15, "bold"), text_color="#1E293B").pack(pady=(20, 15))

        # Botões Coluna 1 (Empilhados)
        self.btn_cep = ctk.CTkButton(self.col1_frame, text="📍 Atualizar CEP", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_cep, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_cep.pack(pady=8, padx=15)
        ToolTip(self.btn_cep, "Reparo automatizado de CEPs")

        self.btn_analise = ctk.CTkButton(self.col1_frame, text="📊 Vendas por Hora", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_analise, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_analise.pack(pady=8, padx=15)

        self.btn_analise_prod = ctk.CTkButton(self.col1_frame, text="🛒 Vendas por Produto", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_analise_produto, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_analise_prod.pack(pady=8, padx=15)

        self.btn_resumo_cli = ctk.CTkButton(self.col1_frame, text="🛍️ Análise Compras por Cliente", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_resumo_cliente, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_resumo_cli.pack(pady=8, padx=15)

        self.btn_pararam = ctk.CTkButton(self.col1_frame, text="🛑 Clientes Deixaram de Comprar", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_clientes_pararam, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_pararam.pack(pady=8, padx=15)

        self.btn_contas = ctk.CTkButton(self.col1_frame, text="💰 Resumo por Tipo Venda", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_posicao_contas, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_contas.pack(pady=8, padx=15)

        # --- 2º COLUNA ( OPERAÇÃO / ESTOQUE ) ---
        self.col2_frame = ctk.CTkFrame(self.modules_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col2_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col2_frame, text="2º Coluna", font=("Arial", 15, "bold"), text_color="#1E293B").pack(pady=(20, 15))

        self.btn_estoque_parado = ctk.CTkButton(self.col2_frame, text="📦 Estoque Parado", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_estoque_parado, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_estoque_parado.pack(pady=8, padx=15)

        self.btn_sugestao_compra = ctk.CTkButton(self.col2_frame, text="🛒 Sugestão de Compra", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_sugestao_compra, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_sugestao_compra.pack(pady=8, padx=15)

        self.btn_parcelas = ctk.CTkButton(self.col2_frame, text="💳 Análise das Parcelas", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_analise_parcelas, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_parcelas.pack(pady=8, padx=15)

        self.btn_cobranca = ctk.CTkButton(self.col2_frame, text="📞 Cobrança por Cliente", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_cobranca, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_cobranca.pack(pady=8, padx=15)

        self.btn_excluir_inativos = ctk.CTkButton(self.col2_frame, text="🗑️ Excluir Clientes Inativos", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_excluir_inativos, fg_color="#C62828", hover_color="#B12020")
        self.btn_excluir_inativos.pack(pady=8, padx=15)

        # --- 3º COLUNA ( BI / ANÁLISES ) ---
        self.col3_frame = ctk.CTkFrame(self.modules_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col3_frame.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col3_frame, text="3º Coluna", font=("Arial", 15, "bold"), text_color="#1E293B").pack(pady=(20, 15))
        
        self.btn_vendas_qtd = ctk.CTkButton(self.col3_frame, text="📊 Vendas Produtos x Qtde", width=btn_width, height=45, font=("Arial", 13, "bold"), 
                                           command=self.abrir_vendas_quantidade, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_vendas_qtd.pack(pady=8, padx=15)

        self.btn_volume_vendas_ticket = ctk.CTkButton(self.col3_frame, text="📈 Volume Vendas (Ticket Médio)", width=btn_width, height=45, font=("Arial", 13, "bold"), 
                                           command=self.abrir_volume_vendas_ticket, fg_color="#1E88E5", hover_color="#1565C0")
        self.btn_volume_vendas_ticket.pack(pady=8, padx=15)

        # --- OUTRA COLUNA VAZIA ( placeholder ) ---
        f4 = ctk.CTkFrame(self.modules_frame, fg_color="#F8FAFC", corner_radius=15, border_width=1, border_color="#E2E8F0")
        f4.grid(row=0, column=3, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(f4, text="4º Coluna", font=("Arial", 15, "bold"), text_color="#94A3B8").pack(pady=(20, 15))

        # --- 5º COLUNA ( POSTO DE COMBUSTÍVEL ) ---
        self.col5_frame = ctk.CTkFrame(self.modules_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col5_frame.grid(row=0, column=4, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col5_frame, text="5º Coluna", font=("Arial", 15, "bold"), text_color="#1E293B").pack(pady=(20, 15))
        
        self.btn_posto = ctk.CTkButton(self.col5_frame, text="⛽ POSTO DE COMBUSTÍVEL", width=btn_width, height=45, font=("Arial", 13, "bold"), command=self.abrir_posto, fg_color="#2E7D32", hover_color="#1B5E20")
        self.btn_posto.pack(pady=8, padx=15)
        ToolTip(self.btn_posto, "Módulo completo para Gestão de Postos")

    def on_posto_change_hub(self, choice):
        print(f"DEBUG: Mudou posto para -> {choice}")
        cod = self.postos_map_hub.get(choice)
        if cod is not None:
            print(f"DEBUG: ID encontrado -> {cod}")
            self.atualizar_info_posto_hub(cod)
            # Ao mudar o posto, se houver um módulo aberto, tentar recarregar seus dados
            if hasattr(self, 'modulo_atual') and hasattr(self.modulo_atual, 'carregar_dados'):
                try: 
                    print(f"DEBUG: Recarregando dados do módulo...")
                    self.modulo_atual.carregar_dados()
                except Exception as e: print(f"DEBUG Erro recarregar modulo: {e}")
        else:
            print(f"DEBUG: ID NÃO ENCONTRADO para {choice}. Mapa: {self.postos_map_hub}")

    def get_selected_posto_id(self):
        """Retorna o ID do posto atualmente selecionado no Hub."""
        try:
            choice = self.combo_posto_hub.get()
            return self.postos_map_hub.get(choice, 1)
        except: return 1

    def carregar_postos_hub(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config['servidor']};DATABASE={self.config['banco']};UID={self.config['usuario_bd']};PWD={self.config['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            cur.execute("SELECT POEmpCod, POEmpDes FROM POEmp ORDER BY POEmpCod")
            rows = cur.fetchall(); conn.close()
            nomes = []; default_val = None
            for cod, des in rows:
                nome_formatado = f"{int(cod)} - {str(des).strip().upper()}"
                nomes.append(nome_formatado)
                self.postos_map_hub[nome_formatado] = int(cod)
                if int(cod) == 1: default_val = nome_formatado
            self.combo_posto_hub.configure(values=nomes)
            if default_val:
                self.combo_posto_hub.set(default_val)
                self.on_posto_change_hub(default_val)
            elif nomes:
                self.combo_posto_hub.set(nomes[0])
                self.on_posto_change_hub(nomes[0])
        except Exception as e: print(f"Erro carregar postos hub: {e}")

    def atualizar_info_posto_hub(self, cod):
        """Atualiza a Data e Período no cabeçalho lendo da tabela POEmp."""
        try:
            # Feedback Visual Imediato
            self.lbl_info_data_hub.configure(text="📅 DATA: Carregando...")
            self.lbl_info_per_hub.configure(text="🕒 PER: ...")
            self.update_idletasks()

            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config['servidor']};DATABASE={self.config['banco']};UID={self.config['usuario_bd']};PWD={self.config['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            cur.execute("SELECT POEmpDtaUs, POEmpPer FROM POEmp WHERE POEmpCod = ?", cod)
            row = cur.fetchone(); conn.close()
            if row:
                dta, per = row
                dta_str = dta.strftime("%d/%m/%Y") if hasattr(dta, 'strftime') else str(dta)
                self.lbl_info_data_hub.configure(text=f"📅 DATA ATUAL DA CONTABILIZAÇÃO: {dta_str}")
                self.lbl_info_per_hub.configure(text=f"🕒 PERÍODO ATUAL: {per}")
                print(f"DEBUG: Hub Atualizado -> Data: {dta_str}, Per: {per}")
            else:
                print(f"DEBUG: Nenhum registro encontrado para POEmpCod={cod}")
            self.update_idletasks() 
        except Exception as e:
            print(f"CRITICAL ERROR carregar info hub: {e}")
            self.lbl_info_data_hub.configure(text="📅 Erro ao carregar")

    def confirmar_saida(self):
        """Exibe um diálogo premium para confirmação de saída do sistema."""
        class LogoutDialog(ctk.CTkToplevel):
            def __init__(self, master):
                super().__init__(master)
                self.title("Sair do Sistema")
                self.geometry("450x250")
                self.configure(fg_color="#F8FAFC")
                self.result = False
                self.grab_set()
                self.resizable(False, False)
                
                # Centralizar em relação ao master
                self.update_idletasks()
                x = master.winfo_x() + (master.winfo_width() // 2) - 225
                y = master.winfo_y() + (master.winfo_height() // 2) - 125
                self.geometry(f"+{x}+{y}")
                self.transient(master)

                # Icone de Saída Grande
                ctk.CTkLabel(self, text="🚪", font=("Arial", 60)).pack(pady=(25, 10))
                
                ctk.CTkLabel(self, text="ENCERRAR SESSÃO", font=("Arial", 18, "bold"), text_color="#1E293B").pack(pady=5)
                ctk.CTkLabel(self, text="Deseja realmente fechar o SIL3000?", font=("Arial", 14), text_color="#64748B").pack(pady=5)
                
                btn_frame = ctk.CTkFrame(self, fg_color="transparent")
                btn_frame.pack(pady=20)
                
                self.btn_sim = ctk.CTkButton(btn_frame, text="✅ SIM, SAIR", command=self.confirmar, fg_color="#D32F2F", hover_color="#B71C1C", width=140, height=40, font=("Arial", 13, "bold"))
                self.btn_sim.pack(side="left", padx=10)
                
                self.btn_nao = ctk.CTkButton(btn_frame, text="✖ CANCELAR", command=self.destroy, fg_color="#94A3B8", hover_color="#64748B", width=140, height=40, font=("Arial", 13, "bold"))
                self.btn_nao.pack(side="left", padx=10)
                
                self.bind("<Return>", lambda e: self.confirmar())
                self.bind("<Escape>", lambda e: self.destroy())

            def confirmar(self):
                self.result = True
                self.destroy()

        dialog = LogoutDialog(self)
        self.wait_window(dialog)
        if dialog.result:
            self.destroy()

    def testar_conexao_banco(self):
        import pyodbc
        from tkinter import messagebox
        
        server = self.config.get("servidor")
        database = self.config.get("banco")
        username = self.config.get("usuario_bd")
        password = self.config.get("senha_bd")
        
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
        
        try:
            conn = pyodbc.connect(conn_str)
            conn.close()
        except:
            try:
                # Fallback de Driver legado
                conn_str_alt = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
                conn = pyodbc.connect(conn_str_alt)
                conn.close()
            except Exception as e:
                # Alerta Amigável
                messagebox.showerror("Erro de Conexão", "Não foi possível conectar ao Banco de Dados com as credenciais fornecidas.\n\nPor favor, verifique os dados de acesso (Servidor, Banco, Usuário ou Senha) nas Configurações.")
                # Abre ConfigWindow (CFG_ACESSO) de forma automática
                self.abrir_configuracoes()

        self.protocol("WM_DELETE_WINDOW", lambda: None)
        self.after(100, self._desabilitar_close_windows)

    def _desabilitar_close_windows(self):
        try:
            import ctypes
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            if not hwnd: hwnd = self.winfo_id()
            hMenu = ctypes.windll.user32.GetSystemMenu(hwnd, False)
            if hMenu:
                 ctypes.windll.user32.EnableMenuItem(hMenu, 0xF060, 1 | 2)
        except: pass

    def copiar_id(self, texto):
         self.clipboard_clear()
         self.clipboard_append(texto)
         self.update()
         messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def atualizar_rodape(self):
        self.lbl_footer.configure(text=f"👤 {self.config['nome_usuario']}  |  🖥️ {self.config['servidor']}  |  🗄️ {self.config['banco']}")

    def carregar_config(self):
        """Carrega configurações do arquivo json se existir."""
        config_path = os.path.join(os.path.dirname(__file__), "config.json")
        if os.path.exists(config_path):
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    saved_config = json.load(f)
                    self.config.update(saved_config)
                    print(f"DEBUG: Configurações carregadas de {config_path}")
            except Exception as e:
                print(f"DEBUG: Erro ao carregar config.json: {e}")

    def salvar_config_persistente(self, nova_config):
        """Salva configurações no arquivo json e atualiza o estado atual."""
        self.config = nova_config
        self.atualizar_rodape()
        config_path = os.path.join(os.path.dirname(__file__), "config.json")
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
                print(f"DEBUG: Configurações salvas em {config_path}")
        except Exception as e:
            print(f"DEBUG: Erro ao salvar config.json: {e}")

    def abrir_configuracoes(self): 
        ConfigWindow(self, self.config, self.salvar_config_persistente)

    def fechar_modulo_atual(self):
        if hasattr(self, 'modulo_atual') and self.modulo_atual:
             self.modulo_atual.destroy()
             self.modulo_atual = None
        self.top_header_hub.pack_forget() # <--- Garante que o cabeçalho suma ao voltar pro Hub
        self.modules_frame.pack(fill="both", expand=True, padx=30, pady=30)
        
        # Bugfix CustomTkinter: Força recalculo de geometria
        self.update()
        w = self.winfo_width(); h = self.winfo_height()
        if w > 1 and h > 1:
             self.geometry(f"{w+1}x{h}")
             self.after(5, lambda: self.geometry(f"{w}x{h}"))

    def abrir_modulo(self, classe_modulo):
        if hasattr(self, "modulo_atual") and self.modulo_atual:
             self.modulo_atual.destroy()
        self.modules_frame.pack_forget()
        self.modulo_atual = classe_modulo(self.main_frame, self.config)
        self.modulo_atual.hub = self
        
        # --- Lógica de Exibição do Cabeçalho de Posto ---
        # Só aparece no ModuloPostoWindow ou em tabelas que marcaram 'is_posto_table'
        is_posto = isinstance(self.modulo_atual, ModuloPostoWindow) or getattr(self.modulo_atual, 'is_posto_table', False)
        if is_posto:
             # Pack do cabeçalho primeiro para ficar no topo
             self.top_header_hub.pack(side="top", fill="x")
        else:
             self.top_header_hub.pack_forget()

        self.modulo_atual.pack(fill="both", expand=True)
        
        # Bugfix CustomTkinter: Força recalculo ao abrir módulo
        self.update()
        w = self.winfo_width(); h = self.winfo_height()
        if w > 1 and h > 1:
             self.geometry(f"{w+1}x{h}")
             self.after(50, lambda: self.geometry(f"{w}x{h}")) # <--- Delay aumentado para 50ms

    def abrir_analise_produto(self): self.abrir_modulo(AnaliseProdutoWindow)
    def abrir_analise(self): self.abrir_modulo(AnaliseVendasWindow)
    def abrir_cep(self): self.abrir_modulo(App)
    def abrir_posto(self): self.abrir_modulo(ModuloPostoWindow)

    def abrir_resumo_cliente(self): self.abrir_modulo(ResumoClienteWindow)
    def abrir_clientes_pararam(self): self.abrir_modulo(ClientesPararamWindow)
    def abrir_posicao_contas(self): self.abrir_modulo(PosicaoContasReceberWindow)
    def abrir_estoque_parado(self): self.abrir_modulo(EstoqueParadoWindow)
    def abrir_sugestao_compra(self): self.abrir_modulo(SugestaoCompraWindow)
    def abrir_analise_parcelas(self): self.abrir_modulo(AnaliseParcelasWindow)
    def abrir_cobranca(self): self.abrir_modulo(CobrancaClienteWindow)
    def abrir_excluir_inativos(self): self.abrir_modulo(ExcluirClientesInativosWindow)
    def abrir_vendas_quantidade(self): self.abrir_modulo(AnaliseVendasQuantidadeWindow)
    def abrir_volume_vendas_ticket(self): self.abrir_modulo(VolumeVendasTicketWindow)

    def abrir_modulo_detalhes_inativos(self, cod, nome):
        if hasattr(self, "modulo_atual") and self.modulo_atual:
             self.modulo_atual.destroy()
        self.modules_frame.pack_forget()
        self.top_header_hub.pack_forget() # Oculta cabeçalho de Posto
        self.modulo_atual = DetalhesParcelasInativosWindow(self.main_frame, self.config, cod, nome)
        self.modulo_atual.hub = self
        self.modulo_atual.pack(fill="both", expand=True)

    def abrir_tabela_posto(self, key):
        config_tabela = POSTO_TABLES_CONFIG.get(key)
        if config_tabela:
            # Roteamento Especializado: POENF utiliza tela customizada com filtros avançados
            if key == "POENF":
                # Abrir com flag de retorno ao posto habilitada
                def instanciador_nf(parent, config):
                    win = EntradaNFPostoWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_nf)
                return

            if key == "POLMC1":
                def instanciador_lmc1(parent, config):
                    win = LMCVendaBombaWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_lmc1)
                return

            if key == "POCF1" or key == "MOV_POSTO":
                def instanciador_mov(parent, config):
                    win = MovimentoPostoWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_mov)
                return

            if key == "POLMC2":
                def instanciador_lmc2(parent, config):
                    win = LMCEstoqueWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_lmc2)
                return

            if key == "POLMC3":
                def instanciador_lmc3(parent, config):
                    win = LMCResumoWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_lmc3)
                return

            if key == "POCF2":
                def instanciador_cf2(parent, config):
                    win = CaixaFrentistaResumoWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_cf2)
                return

            if key == "POTCV":
                def instanciador_tcv(parent, config):
                    win = TotCompraVendaWindow(parent, config)
                    win.is_posto_table = True
                    return win
                self.abrir_modulo(instanciador_tcv)
                return

            # Custom load logic for the generic window
            def instanciador(parent, config):
                win = TabelaPostoWindow(parent, config, config_tabela)
                win.is_posto_table = True
                return win
            self.abrir_modulo(instanciador)

    def abrir_modulo_detalhes_cobranca(self, cod, nome):
        if hasattr(self, "modulo_atual") and self.modulo_atual:
             self.modulo_atual.destroy()
        self.modules_frame.pack_forget()
        self.top_header_hub.pack_forget() # Oculta cabeçalho de Posto
        self.modulo_atual = DetalhesParcelasCobrancaWindow(self.main_frame, self.config, cod, nome)
        self.modulo_atual.hub = self
        self.modulo_atual.pack(fill="both", expand=True)

class ConciliacaoConcentradorVendasWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Auditoria: CRMov2 x POCF1", "CONCILIACAO_CONCENTRADOR_VENDAS")
        self.config_db = config
        self.hub = parent
        self.is_posto_table = True

        # --- FRAME DE FILTROS (No Top Bar da BaseWindow) ---
        self.filter_container = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_container.pack(side="left", padx=20, pady=(0, 5), anchor="w")
        
        # Filtros de Data
        from datetime import datetime, date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")
        
        ctk.CTkLabel(self.filter_container, text="De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_container, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_ate = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))
        
        self.btn_filtrar = ctk.CTkButton(self.filter_container, text="🔎 Conciliar", width=120, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#0D47A1", hover_color="#1565C0", command=self.carregar_divergencias)
        self.btn_filtrar.pack(side="left", padx=10)
        
        # Configuração da Grid usando o Padrão Centralizado
        cols = ("data", "seq", "item", "cli_id", "cli_nome", "pro", "des", "qtd", "vlr", "status")
        heads = ("📅 Data", "🔢 Seq.", "📦 Item", "👤 ID", "🏢 Cliente", "🏷️ Cod.", "📝 Produto", "📊 Qtd.", "💰 Preço", "📢 Divergência")
        wids = (90, 70, 70, 70, 160, 70, 200, 80, 100, 250)
        alns = ("center", "center", "center", "center", "w", "center", "w", "center", "e", "w")
        self.configurar_grid(cols, heads, wids, alns)

        # Chamar carga automática com um pequeno delay para garantir populução do campos
        self.after(500, self.carregar_divergencias)

    def voltar(self):
        if hasattr(self, 'hub') and self.hub:
            self.hub.abrir_posto()

    def copiar_id_tela(self):
        self.clipboard_clear(); self.clipboard_append(self.id_str); self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")

    def get_current_query(self):
        """Monta o SQL real de auditoria fiscal cruzada (CRMov x POCF1) com filtros de data atuais."""
        posto_id = self.config_db.get("id_posto", "1")
        de = self.ent_de.get().strip()
        ate = self.ent_ate.get().strip()
        
        return f"""SELECT 
    c2.CRMovDta, c2.CRMovSeq, c4.CRMov4Ite, c2.CRMov2CodC, c2.CRMov2DesC, 
    c4.CEProCod, p.CEProDes, c4.CRMov4Qtd, c4.CRMov4VlrA,
    (SELECT COUNT(*) FROM POCF1 s 
     WHERE s.POEmpCod = c2.CMEmpCod 
     AND s.POCF1DtaMo = c2.CRMovDta 
     AND s.POCF1SeqMo = c2.CRMovSeq 
     AND s.POCF1IteMo = c4.CRMov4Ite) as QtdVendas
FROM CRMov2 c2
JOIN CRMov4 c4 ON c2.CMEmpCod = c4.CMEmpCod AND c2.CRMovDta = c4.CRMovDta AND c2.CRMovSeq = c4.CRMovSeq
LEFT JOIN CEPro p ON p.CEProCod = c4.CEProCod
WHERE c2.CMEmpCod = '{posto_id}' 
AND CAST(c2.CRMovDta AS DATE) BETWEEN '{de}' AND '{ate}'
AND c4.CEProCod < 10
AND c2.CRMov2Flag IN ('A', 'F')
AND (
    NOT EXISTS (SELECT 1 FROM POCF1 s 
                WHERE s.POEmpCod = c2.CMEmpCod 
                AND s.POCF1DtaMo = c2.CRMovDta 
                AND s.POCF1SeqMo = c2.CRMovSeq 
                AND s.POCF1IteMo = c4.CRMov4Ite)
    OR 
    (SELECT COUNT(*) FROM POCF1 s 
     WHERE s.POEmpCod = c2.CMEmpCod 
     AND s.POCF1DtaMo = c2.CRMovDta 
     AND s.POCF1SeqMo = c2.CRMovSeq 
     AND s.POCF1IteMo = c4.CRMov4Ite) > 1
)
ORDER BY c2.CRMovDta DESC"""

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL técnico detalhado seguindo o padrão premium."""
        query = self.get_current_query()
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", query)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def copiar_sql_debug(self, sql):
        self.clipboard_clear(); self.clipboard_append(sql); self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "COMANDO SQL COPIADO PARA O CLIPBOARD")

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            "--- ESTRUTURA SQL: CONCILIAÇÃO VENDAS X CONCENTRADOR ---\n"
            "TABELAS PRINCIPAIS: POCF1 (Cupons Concentrador) e CRMov2/4 (Vendas ERP)\n"
            "RELACIONAMENTOS:\n"
            "- Cruzamento de Cupons registrados fisicamente X notas processadas no ERP.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Data, Sequência, Cliente, Produto, Quantidade e Divergências de Batimento.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados da tela por Data/Período do caixa e Empresa.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordem Cronológica por Emissão (DESC)"
        )
    def carregar_divergencias(self):
        try:
            import pyodbc
            from datetime import datetime
            for i in self.tree.get_children(): self.tree.delete(i)
            
            # Formatar datas com strip para evitar erros de espaço
            try:
                de_raw = self.ent_de.get().strip()
                ate_raw = self.ent_ate.get().strip()
                if not de_raw or not ate_raw: return # Evitar erro se ainda vazio
                
                de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
                ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            except Exception as ex:
                print(f"DEBUG Erro Data: {ex} | DE: '{de_raw}' ATE: '{ate_raw}'")
                messagebox.showerror("Erro de Data", f"Formato inválido! Use DD/MM/AAAA\n\nErro: {str(ex)}")
                return
            
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            # Recuperar Posto através do Hub principal
            posto_id = 1
            if hasattr(self, "hub") and self.hub:
                posto_id = self.hub.get_selected_posto_id()
            
            # SQL DE BATIMENTO REFINADO - Com Filtro de Data Preciso
            sql = f"""
                SELECT 
                    c2.CRMovDta, 
                    c2.CRMovSeq, 
                    c4.CRMov4Ite, 
                    c2.CRMov2CodC as CliID,
                    c2.CRMov2DesC as CliNome,
                    c4.CEProCod as ProCod,
                    COALESCE(p.CEProDes, 'N/D') as ProDes,
                    COALESCE(c4.CRMov4Qtd, 0) as Qtd,
                    COALESCE(c4.CRMov4VlrA, 0) as Preco,
                    (SELECT COUNT(*) FROM POCF1 s 
                     WHERE s.POEmpCod = c2.CMEmpCod 
                     AND s.POCF1DtaMo = c2.CRMovDta 
                     AND s.POCF1SeqMo = c2.CRMovSeq 
                     AND s.POCF1IteMo = c4.CRMov4Ite) as QtdVendas
                FROM CRMov2 c2
                JOIN CRMov4 c4 ON c2.CMEmpCod = c4.CMEmpCod AND c2.CRMovDta = c4.CRMovDta AND c2.CRMovSeq = c4.CRMovSeq
                LEFT JOIN CEPro p ON p.CEProCod = c4.CEProCod
                WHERE c2.CMEmpCod = ? 
                AND CAST(c2.CRMovDta AS DATE) BETWEEN ? AND ?
                AND c4.CEProCod < 10
                AND (
                    NOT EXISTS (SELECT 1 FROM POCF1 s 
                                WHERE s.POEmpCod = c2.CMEmpCod 
                                AND s.POCF1DtaMo = c2.CRMovDta 
                                AND s.POCF1SeqMo = c2.CRMovSeq 
                                AND s.POCF1IteMo = c4.CRMov4Ite)
                    OR 
                    (SELECT COUNT(*) FROM POCF1 s 
                     WHERE s.POEmpCod = c2.CMEmpCod 
                     AND s.POCF1DtaMo = c2.CRMovDta 
                     AND s.POCF1SeqMo = c2.CRMovSeq 
                     AND s.POCF1IteMo = c4.CRMov4Ite) > 1
                )
                ORDER BY c2.CRMovDta DESC
            """
            # Executar passando os parâmetros formatados
            # Overlay de Processamento Centralizado (Diretamente na tela, não PopUp)
            self.overlay_prog = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#1E88E5", corner_radius=15)
            self.overlay_prog.place(relx=0.5, rely=0.5, anchor="center")
            
            ctk.CTkLabel(self.overlay_prog, text="🚀 AUDITORIA CONCENTRADOR X VENDAS", font=("Arial", 14, "bold"), text_color="#1E293B").pack(padx=30, pady=(20, 5))
            
            self.p_bar = ctk.CTkProgressBar(self.overlay_prog, width=300, height=12, corner_radius=10, fg_color="#E2E8F0", progress_color="#1E88E5")
            self.p_bar.pack(padx=30, pady=10)
            self.p_bar.set(0)
            
            self.lbl_prog_val = ctk.CTkLabel(self.overlay_prog, text="⏳ Iniciando conexão...", font=("Arial", 12), text_color="#1E88E5")
            self.lbl_prog_val.pack(pady=(0, 20))
            
            self.update() # Força a pintura do overlay antes da query pesada
            
            self.ultima_query_bruta = sql.replace("?", f"'{posto_id}'", 1).replace("?", f"'{de}'", 1).replace("?", f"'{ate}'", 1)
            cur.execute(sql, (posto_id, de, ate))
            rows = cur.fetchall()
            
            total_regs = len(rows)
            for i, row in enumerate(rows):
                # row[0]:Dta, row[1]:Seq, row[2]:Ite, row[3]:CliID, row[4]:CliNom, row[5]:CodPro, row[6]:DesPro, row[7]:Qtd, row[8]:Preco, row[9]:QtdVendas
                status = "❌ NÃO LANÇADO (POCF1)" if row[9] == 0 else f"⚠️ REPLICADO {row[9]}X"
                v_data = row[0].strftime("%d/%m/%Y") if hasattr(row[0], "strftime") else str(row[0])
                
                self.tree.insert("", "end", values=(
                    v_data, row[1], row[2], row[3], str(row[4]).strip().upper(), 
                    row[5], row[6], f"{row[7]:.2f}", f"R$ {row[8]:.2f}", status
                ))
                
                # Atualizar progresso centralizado e Barra
                if (i + 1) % 50 == 0 or (i + 1) == total_regs:
                    self.lbl_prog_val.configure(text=f"⚙️ Processando: {i+1} de {total_regs}...")
                    self.p_bar.set((i + 1) / total_regs)
                    self.update()

            conn.close()
            self.overlay_prog.destroy()
            
            # Finalização no Rodapé
            if hasattr(self, "lbl_total_regs"):
                self.lbl_total_regs.configure(text=f"📊 Total de Divergências: {total_regs} registros", text_color="#1E293B")
            
            if total_regs == 0:
                messagebox.showinfo("Sucesso Conciliação", "✅ Nenhuma divergência encontrada!\n\nTodos os registros de vendas do Concentrador (CRMov2) foram localizados corretamente no SIL (POCF1).")

            self.re_zebrar()
            
        except Exception as e:
            messagebox.showerror("Erro Conciliação", f"Falha ao cruzar dados:\n\n{str(e)}")

class ConciliacaoPOCF4Window(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Auditoria: POCF4 x POCF1", "CONCILIACAO_POCF4_POCF1")
        self.config_db = config
        self.hub = parent
        self.is_posto_table = True

        # --- FRAME DE FILTROS (No Top Bar da BaseWindow) ---
        self.filter_container = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_container.pack(side="left", padx=20, pady=(0, 5), anchor="w")
        
        # Filtros de Data
        from datetime import datetime, date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")
        
        ctk.CTkLabel(self.filter_container, text="De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_container, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_ate = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))
        
        self.btn_filtrar = ctk.CTkButton(self.filter_container, text="🔎 Conciliar", width=120, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#0D47A1", hover_color="#1565C0", command=self.carregar_divergencias)
        self.btn_filtrar.pack(side="left", padx=10)

        # Configuração da Grid usando o Padrão BaseWindow
        cols = ("leitura", "regis", "qtd", "vlr_bo", "status")
        heads = ("📅 Data\Hora da Leitura Concentrador", "📋 Registro", "📊 Quantidade de Litros", "💰 Valor Abastecimento", "📢 Divergência")
        wids = (250, 110, 150, 150, 300)
        alns = ("center", "center", "center", "e", "w")
        self.configurar_grid(cols, heads, wids, alns)

        self.after(500, self.carregar_divergencias)

    def voltar(self): 
        if hasattr(self, 'hub') and self.hub: self.hub.abrir_posto()

    def copiar_id_tela(self):
        self.clipboard_clear()
        self.clipboard_append(self.id_str)
        # Apenas copia, seguindo o padrão simplificado solicitado


    def get_current_query(self):
        de_raw = self.ent_de.get().strip()
        ate_raw = self.ent_ate.get().strip()
        posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
        try:
            from datetime import datetime
            di = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            df = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
        except: di, df = '1900-01-01', '1900-01-01'

        return f"""
            SELECT 
                f4.POCF4Tst, f4.POCF4Regis, f4.POCF4Qtd, f4.POCF4VlrBo
            FROM POCF4 f4
            WHERE f4.POEmpCod = {posto_id}
              AND CAST(f4.POCF4Tst AS DATE) >= '{di}'
              AND CAST(f4.POCF4Tst AS DATE) <= '{df}'
              AND NOT EXISTS (
                  SELECT 1
                  FROM POCF1 f1
                  WHERE f1.POEmpCod = f4.POEmpCod
                    AND f1.POCF1Tst = f4.POCF4Tst
              )
            ORDER BY f4.POCF4Tst DESC
        """.strip()

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL Pronto/Executado em tempo real, seguindo o padrão premium."""
        query = self.get_current_query()
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", query)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def abrir_calendario(self, entry_target):
        pop = ctk.CTkToplevel(self)
        pop.title("📅 Selecionar Data")
        pop.geometry("320x400")
        pop.attributes("-topmost", True)
        pop.grab_set()
        from datetime import date
        import calendar
        hoje = date.today()
        cal_state = {"mes": hoje.month, "ano": hoje.year}
        main_f = ctk.CTkFrame(pop, fg_color="white")
        main_f.pack(fill="both", expand=True, padx=10, pady=10)
        header_f = ctk.CTkFrame(main_f, fg_color="transparent")
        header_f.pack(fill="x", pady=5)
        lbl_mes_ano = ctk.CTkLabel(header_f, text="", font=("Arial", 14, "bold"), text_color="#1E88E5")
        def update_view():
            mes, ano = cal_state["mes"], cal_state["ano"]
            nome_mes = ["?", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][mes]
            lbl_mes_ano.configure(text=f"{nome_mes} {ano}")
            for widget in grid_f.winfo_children():
                if int(widget.grid_info()["row"]) > 0: widget.destroy()
            cal = calendar.monthcalendar(ano, mes)
            for r, week in enumerate(cal):
                for c, day in enumerate(week):
                    if day != 0:
                        btn = ctk.CTkButton(grid_f, text=str(day), width=35, height=35, fg_color="#F1F5F9", text_color="black", hover_color="#CBD5E1", command=lambda d=day: set_date(d))
                        btn.grid(row=r+1, column=c, padx=2, pady=2)
        def mudar_mes(inc):
            cal_state["mes"] += inc
            if cal_state["mes"] > 12: cal_state["mes"] = 1; cal_state["ano"] += 1
            elif cal_state["mes"] < 1: cal_state["mes"] = 12; cal_state["ano"] -= 1
            update_view()
        ctk.CTkButton(header_f, text="◀", width=30, height=30, fg_color="#94A3B8", command=lambda: mudar_mes(-1)).pack(side="left", padx=10)
        lbl_mes_ano.pack(side="left", expand=True)
        ctk.CTkButton(header_f, text="▶", width=30, height=30, fg_color="#94A3B8", command=lambda: mudar_mes(1)).pack(side="right", padx=10)
        grid_f = ctk.CTkFrame(main_f, fg_color="transparent")
        grid_f.pack(fill="both", expand=True)
        dias_sem = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
        for i, ds in enumerate(dias_sem): ctk.CTkLabel(grid_f, text=ds, font=("Arial", 10, "bold"), text_color="gray").grid(row=0, column=i, pady=5)
        def set_date(dia):
            entry_target.delete(0, "end"); entry_target.insert(0, f"{dia:02d}/{cal_state['mes']:02d}/{cal_state['ano']}")
            pop.destroy()
        update_view()

    def carregar_divergencias(self):
        try:
            import pyodbc
            from datetime import datetime
            for i in self.tree.get_children(): self.tree.delete(i)
            de_raw = self.ent_de.get().strip(); ate_raw = self.ent_ate.get().strip()
            if not de_raw or not ate_raw: return
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, "hub") and self.hub else 1
            
            sql = f"""
                SELECT 
                    f4.POCF4Tst, 
                    f4.POCF4Regis, 
                    f4.POCF4Qtd,
                    f4.POCF4VlrBo
                FROM POCF4 f4
                WHERE f4.POEmpCod = ?
                  AND CAST(f4.POCF4Tst AS DATE) >= ?
                  AND CAST(f4.POCF4Tst AS DATE) <= ?
                  AND NOT EXISTS (
                      SELECT 1
                      FROM POCF1 f1
                      WHERE f1.POEmpCod = f4.POEmpCod
                        AND f1.POCF1Tst = f4.POCF4Tst
                  )
                ORDER BY f4.POCF4Tst DESC
            """
            self.overlay_prog = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#1E88E5", corner_radius=15)
            self.overlay_prog.place(relx=0.5, rely=0.5, anchor="center")
            ctk.CTkLabel(self.overlay_prog, text="🚀 AUDITORIA POCF4 X POCF1", font=("Arial", 14, "bold"), text_color="#1E293B").pack(padx=30, pady=(20, 5))
            self.p_bar = ctk.CTkProgressBar(self.overlay_prog, width=300, height=12, corner_radius=10, fg_color="#E2E8F0", progress_color="#1E88E5")
            self.p_bar.pack(padx=30, pady=10); self.p_bar.set(0)
            self.lbl_prog_val = ctk.CTkLabel(self.overlay_prog, text="⏳ Cruzando registros...", font=("Arial", 12), text_color="#1E88E5")
            self.lbl_prog_val.pack(pady=(0, 20))
            self.update()

            cur.execute(sql, (posto_id, de, ate))
            rows = cur.fetchall()
            total_regs = len(rows)
            total_valor = 0
            for i, row in enumerate(rows):
                # row[0]:Tst, row[1]:Regis, row[2]:Qtd, row[3]:VlrBo
                total_valor += float(row[3] or 0)
                status = "❌ NÃO LANÇADO (POCF1)" 
                self.tree.insert("", "end", values=(row[0], row[1], f"{row[2]:.2f}", f"R$ {row[3]:.2f}", status))
                if (i + 1) % 50 == 0 or (i + 1) == total_regs:
                    msg = f"⚙️ Processando: {i+1} de {total_regs}..."
                    self.lbl_prog_val.configure(text=msg)
                    self.p_bar.set((i + 1) / total_regs); self.update()
            
            conn.close(); self.overlay_prog.destroy()
            
            # Atualizar Totais no Rodapé (Padrão BaseWindow: tree_totais)
            if hasattr(self, "tree_totais"):
                for i in self.tree_totais.get_children(): self.tree_totais.delete(i)
                # cols = ("leitura", "regis", "qtd", "vlr_bo", "status")
                self.tree_totais.insert("", "end", values=("TOTAL GERAL", "", "", f"R$ {total_valor:,.2f}", ""))
            
            if hasattr(self, "lbl_total_regs"):
                self.lbl_total_regs.configure(text=f"📊 Total: {total_regs} registros | 💰 Soma: R$ {total_valor:,.2f}")
            
            if total_regs == 0:
                # Formatação bonita para o período
                p_inicio = datetime.strptime(de, "%Y-%m-%d").strftime("%d/%m/%Y")
                p_fim = datetime.strptime(ate, "%Y-%m-%d").strftime("%d/%m/%Y")
                messagebox.showinfo("✅ Conciliação Concluída", f"TUDO OK! Nenhuma divergência encontrada no período de {p_inicio} até {p_fim}.")
            
            self.re_zebrar()
        except Exception as e: messagebox.showerror("Erro Conciliação", str(e))

class ConciliacaoPOCF1CRMOV4Window(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Auditoria: POCF1 x CRMov4", "CONCILIACAO_POCF1_CRMOV4")
        self.config_db = config
        self.hub = parent
        self.is_posto_table = True

        # --- FRAME DE FILTROS (No Top Bar da BaseWindow) ---
        self.filter_container = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_container.pack(side="left", padx=20, pady=(0, 5), anchor="w")
        
        # Filtros de Data
        from datetime import datetime, date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")
        
        ctk.CTkLabel(self.filter_container, text="De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_container, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_ate = ctk.CTkEntry(self.filter_container, width=90, height=32); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_container, text="📅", width=30, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))
        
        self.btn_filtrar = ctk.CTkButton(self.filter_container, text="🔎 Conciliar", width=120, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#0D47A1", hover_color="#1565C0", command=self.carregar_divergencias)
        self.btn_filtrar.pack(side="left", padx=10)

        # Configuração da Grid usando o Padrão BaseWindow (Personalizado)
        cols = ("leitura", "regis", "usu", "vlr_mo", "obs", "tipmo")
        heads = ("📅 Data\Hora Movimento", "📋 Registro", "👤 Usuário", "💰 Valor Movimento", "📝 Observação", "⚙️ Tipo Mov.")
        wids = (220, 110, 130, 150, 200, 100)
        alns = ("center", "center", "center", "e", "w", "center")
        self.configurar_grid(cols, heads, wids, alns)

        self.after(500, self.carregar_divergencias)

    def voltar(self): 
        if hasattr(self, 'hub') and self.hub: self.hub.abrir_posto()

    def copiar_id_tela(self):
        self.clipboard_clear(); self.clipboard_append(self.id_str)
        messagebox.showinfo("Informação", f"Copiado ID da Tela: {self.id_str}")

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: CONCILIACAOPOCF1CRMOV4 ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def abrir_calendario(self, entry_target):
        pop = ctk.CTkToplevel(self); pop.title("📅 Data"); pop.geometry("320x400"); pop.attributes("-topmost", True); pop.grab_set()
        import calendar; hoje = date.today(); cal_state = {"mes": hoje.month, "ano": hoje.year}
        main_f = ctk.CTkFrame(pop, fg_color="white"); main_f.pack(fill="both", expand=True, padx=10, pady=10)
        header_f = ctk.CTkFrame(main_f, fg_color="transparent"); header_f.pack(fill="x", pady=5)
        lbl_mes_ano = ctk.CTkLabel(header_f, text="", font=("Arial", 14, "bold"), text_color="#1E88E5")
        def update_view():
            mes, ano = cal_state["mes"], cal_state["ano"]
            nome_mes = ["?", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][mes]
            lbl_mes_ano.configure(text=f"{nome_mes} {ano}")
            for widget in grid_f.winfo_children():
                if int(widget.grid_info()["row"]) > 0: widget.destroy()
            cal = calendar.monthcalendar(ano, mes)
            for r, week in enumerate(cal):
                for c, day in enumerate(week):
                    if day != 0:
                        btn = ctk.CTkButton(grid_f, text=str(day), width=35, height=35, fg_color="#F1F5F9", text_color="black", hover_color="#CBD5E1", command=lambda d=day: set_date(d))
                        btn.grid(row=r+1, column=c, padx=2, pady=2)
        def mudar_mes(inc):
            cal_state["mes"] += inc
            if cal_state["mes"] > 12: cal_state["mes"] = 1; cal_state["ano"] += 1
            elif cal_state["mes"] < 1: cal_state["mes"] = 12; cal_state["ano"] -= 1
            update_view()
        ctk.CTkButton(header_f, text="◀", width=30, height=30, command=lambda: mudar_mes(-1)).pack(side="left", padx=10); lbl_mes_ano.pack(side="left", expand=True); ctk.CTkButton(header_f, text="▶", width=30, height=30, command=lambda: mudar_mes(1)).pack(side="right", padx=10)
        grid_f = ctk.CTkFrame(main_f, fg_color="transparent"); grid_f.pack(fill="both", expand=True)
        def set_date(dia):
            entry_target.delete(0, "end"); entry_target.insert(0, f"{dia:02d}/{cal_state['mes']:02d}/{cal_state['ano']}"); pop.destroy()
        update_view()

    def get_current_query(self):
        """Retorna a query de auditoria baseada no SQL fornecido pelo usuário."""
        try:
            de_raw = self.ent_de.get().strip(); ate_raw = self.ent_ate.get().strip()
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, "hub") and self.hub else 1
            
            return f"""
                SELECT
                    f1.POCF1Tst,
                    f1.POCF1Regis,
                    f1.POCF1Usu,
                    f1.POCF1VlrMo,
                    f1.POCF1Obs,
                    f1.POCF1TipMo,
                    f1.POEmpCod,
                    f1.POCF1DtaMo,
                    f1.POCF1SeqMo,
                    f1.POCF1IteMo
                FROM POCF1 f1
                WHERE f1.POEmpCod = {posto_id}
                  AND f1.POCF1PuBom > 0
                  AND CAST(f1.POCF1Tst AS DATE) >= '{de}'
                  AND CAST(f1.POCF1Tst AS DATE) <= '{ate}'
                  AND NOT EXISTS (
                      SELECT 1
                      FROM CRMov4 c
                      WHERE c.CMEmpCod = f1.POEmpCod
                        AND c.CRMovDta = f1.POCF1DtaMo
                        AND c.CRMovSeq = f1.POCF1SeqMo
                        AND c.CRMov4Ite = f1.POCF1IteMo
                  )
                ORDER BY f1.POCF1Tst DESC
            """.strip()
        except Exception as e: 
            return f"Erro ao gerar SQL: {str(e)}"

    def mostrar_sql_bruto(self):
        """Exibe o comando SQL seguindo o padrão premium visual."""
        query = self.get_current_query()
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar
        w, h = 800, 600
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=(25, 15))
        
        f = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        f.pack(fill="both", expand=True, padx=40, pady=10)
        
        t = ctk.CTkTextbox(f, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        t.pack(fill="both", expand=True, padx=15, pady=15)
        t.insert("0.0", query)
        t.configure(state="disabled")
        
        btn_f = ctk.CTkFrame(win, fg_color="transparent")
        btn_f.pack(pady=25)
        
        ctk.CTkButton(btn_f, text="📋 COPIAR SQL", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#2E7D32", hover_color="#1B5E20", corner_radius=8,
                     command=lambda: [self.clipboard_clear(), self.clipboard_append(query), self.update(), messagebox.showinfo("OK", "Copiado!")]).pack(side="left", padx=10)
        
        ctk.CTkButton(btn_f, text="FECHAR", width=170, height=40, font=("Arial", 13, "bold"), 
                     fg_color="#475569", hover_color="#334155", corner_radius=8,
                     command=win.destroy).pack(side="left", padx=10)

    def carregar_divergencias(self):
        try:
            import pyodbc; from datetime import datetime
            for i in self.tree.get_children(): self.tree.delete(i)
            de_raw = self.ent_de.get().strip(); ate_raw = self.ent_ate.get().strip()
            if not de_raw or not ate_raw: return
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, "hub") and self.hub else 1
            
            sql = self.get_current_query()
            
            # Overlay de Processamento Premium
            self.overlay_prog = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#1E88E5", corner_radius=15)
            self.overlay_prog.place(relx=0.5, rely=0.5, anchor="center")
            ctk.CTkLabel(self.overlay_prog, text="🚀 AUDITORIA POCF1 X CRMOV4", font=("Arial", 14, "bold"), text_color="#1E293B").pack(padx=30, pady=(20, 5))
            self.p_bar = ctk.CTkProgressBar(self.overlay_prog, width=300, height=12, corner_radius=10, fg_color="#E2E8F0", progress_color="#1E88E5")
            self.p_bar.pack(padx=30, pady=10); self.p_bar.set(0)
            self.lbl_prog_val = ctk.CTkLabel(self.overlay_prog, text="⏳ Capturando registros não integrados...", font=("Arial", 12), text_color="#1E88E5")
            self.lbl_prog_val.pack(pady=(0, 20))
            self.update()

            cur.execute(sql)
            rows = cur.fetchall()
            total_regs = len(rows); total_valor = 0
            
            for i, row in enumerate(rows):
                # row[0]:Tst, row[1]:Regis, row[2]:Usu, row[3]:VlrMo, row[4]:Obs, row[5]:TipMo
                vlr = float(row[3]) if row[3] is not None else 0.0
                total_valor += vlr
                
                self.tree.insert("", "end", values=(
                    row[0], row[1], str(row[2]).strip(), f"R$ {vlr:,.2f}", 
                    str(row[4]).strip() if row[4] else "", str(row[5]).strip() if row[5] else ""
                ))
                
                if (i + 1) % 50 == 0 or (i + 1) == total_regs:
                    self.lbl_prog_val.configure(text=f"⚙️ Processando: {i+1} de {total_regs}...")
                    self.p_bar.set((i + 1) / total_regs); self.update()
            
            conn.close(); self.overlay_prog.destroy()
            
            # Atualizar Totais no Rodapé (Padrão Definido)
            if hasattr(self, "tree_totais"):
                for i in self.tree_totais.get_children(): self.tree_totais.delete(i)
                # cols = ("leitura", "regis", "usu", "vlr_mo", "obs", "tipmo")
                self.tree_totais.insert("", "end", values=("TOTAL GERAL", "", "", f"R$ {total_valor:,.2f}", "", ""))
            
            if hasattr(self, "lbl_total_regs"):
                self.lbl_total_regs.configure(text=f"📊 Total: {total_regs} registros | 💰 Soma: R$ {total_valor:,.2f}")
            
            if total_regs == 0:
                p_inicio = datetime.strptime(de, "%Y-%m-%d").strftime("%d/%m/%Y")
                p_fim = datetime.strptime(ate, "%Y-%m-%d").strftime("%d/%m/%Y")
                messagebox.showinfo("✅ Auditado com Sucesso", f"TUDO OK! Nenhuma divergência no período de {p_inicio} até {p_fim}.")
                
            self.re_zebrar()
        except Exception as e:
            if hasattr(self, "overlay_prog") and self.overlay_prog: self.overlay_prog.destroy()
            messagebox.showerror("Erro Conciliação", str(e))

class ModuloPostoWindow(ctk.CTkFrame):
    def __init__(self, parent, config):
        super().__init__(parent, fg_color="#FFFFFF", corner_radius=0)
        self.config = config
        self.id_str = "MODULO_POSTO"
        
        # O Módulo Posto agora utiliza o Cabeçalho Global do Hub Principal (MainHub)
        # para seleção de Posto, Data e Período.

        # para seleção de Posto, Data e Período.

        # Grid de botões (Multi-Colunas)
        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.pack(fill="both", expand=True, padx=20, pady=20)
        self.content_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        btn_width = 240
        btn_height = 40
        
        # --- COLUNA 1: EQUIPAMENTOS ---
        self.col1 = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col1.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col1, text="🏗️ Equipamentos", font=("Arial", 14, "bold"), text_color="#1E293B").pack(pady=(15, 10))
        
        btns_col1 = [
            ("🏢 Cadastro Posto (POEmp)", "POEmp"),
            ("📻 Bombas (POBom/POMBo)", "POBom"),
            ("🔧 Intervenção Técnica (POBIT)", "POBIT"),
            ("🛢️ Cadastro de Tanques (POTnq)", "POTnq"),
            ("🔐 Cadastro de Lacres (POLac)", "POLac"),
            ("🔗 Vínculo Bico/Bomba (POBxB)", "POBxB")
        ]
        for txt, key in btns_col1:
            cmd = lambda k=key: self.hub.abrir_tabela_posto(k)
            ctk.CTkButton(self.col1, text=txt, width=btn_width, height=btn_height, font=("Arial", 12, "bold"), command=cmd, fg_color="#1E88E5", hover_color="#1565C0").pack(pady=5, padx=15)

        # --- NOVA SEÇÃO: CONSISTÊNCIAS ---
        # Posicionada abaixo de Equipamentos (Linha 1, Coluna 0)
        self.col_consistencias = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col_consistencias.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col_consistencias, text="🔍 Consistências", font=("Arial", 14, "bold"), text_color="#1E293B").pack(pady=(15, 10))
        
        # Botão: CRMov2 x POCF1
        ctk.CTkButton(self.col_consistencias, text="🎯 CRMov2 x POCF1", width=btn_width, height=btn_height, font=("Arial", 12, "bold"), 
                     command=lambda: self.hub.abrir_modulo(ConciliacaoConcentradorVendasWindow), 
                     fg_color="#F59E0B", hover_color="#D97706").pack(pady=5, padx=15)

        # Botão: POCF4 x POCF1
        ctk.CTkButton(self.col_consistencias, text="🔍 POCF4 x POCF1", width=btn_width, height=btn_height, font=("Arial", 12, "bold"), 
                     command=lambda: self.hub.abrir_modulo(ConciliacaoPOCF4Window), 
                     fg_color="#F59E0B", hover_color="#D97706").pack(pady=5, padx=15)

        # Botão: POCF1 x CRMov4 (Vinculado)
        ctk.CTkButton(self.col_consistencias, text="🎯 POCF1 x CRMov4", width=btn_width, height=btn_height, font=("Arial", 12, "bold"), 
                     command=lambda: self.hub.abrir_modulo(ConciliacaoPOCF1CRMOV4Window), 
                     fg_color="#F59E0B", hover_color="#D97706").pack(pady=5, padx=15)

        # --- COLUNA 2: MOVIMENTAÇÃO / LMC ---
        self.col2 = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col2, text="📖 Movimento / LMC", font=("Arial", 14, "bold"), text_color="#1E293B").pack(pady=(15, 10))
        
        btns_col2 = [
            ("⏱️ Velocímetros (POVel1)", "POVel1"),
            ("📜 Hist. Velocímetros (POVel2)", "POVel2"),
            ("📊 LMC 1: Venda p/ Bomba", "POLMC1"),
            ("📈 LMC 2: Estoque/Entrada", "POLMC2"),
            ("📉 LMC 3: Resumo p/ Prod", "POLMC3"),
            ("📚 Livros do LMC (POLvr)", "POLvr")
        ]
        for txt, key in btns_col2:
            cmd = lambda k=key: self.hub.abrir_tabela_posto(k)
            ctk.CTkButton(self.col2, text=txt, width=btn_width, height=btn_height, font=("Arial", 12, "bold"), command=cmd, fg_color="#1E88E5", hover_color="#1565C0").pack(pady=5, padx=15)

        # --- COLUNA 3: FINANCEIRO / CAIXA ---
        self.col3 = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col3.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col3, text="💰 Financeiro / Caixa", font=("Arial", 14, "bold"), text_color="#1E293B").pack(pady=(15, 10))
        
        btns_col3 = [
            ("⭐ Movimentos (POCF1)", "POCF1"),
            ("👤 Caixa Frentista (POCF2)", "POCF2"),
            ("📊 Totais Frentista (POCF3)", "POCF3"),
            ("📑 Itens Frentista (POCF4)", "POCF4"),
            ("🏦 Caixa Geral (POCxa1)", "POCxa1"),
            ("📒 Detalhe Caixa (POCxa2)", "POCxa2"),
            ("🎫 Cadastro de Vales (POVal)", "POVal")
        ]
        for txt, key in btns_col3:
            cmd = lambda k=key: self.hub.abrir_tabela_posto(k)
            ctk.CTkButton(self.col3, text=txt, width=btn_width, height=btn_height, font=("Arial", 12, "bold"), command=cmd, fg_color="#1E88E5", hover_color="#1565C0").pack(pady=5, padx=15)

        # --- COLUNA 4: SUPRIMENTO / DIVERSOS ---
        self.col4 = ctk.CTkFrame(self.content_frame, fg_color="#F8FAFC", corner_radius=15, border_width=2, border_color="#E2E8F0")
        self.col4.grid(row=0, column=3, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(self.col4, text="🏁 Suprimento / Env", font=("Arial", 14, "bold"), text_color="#1E293B").pack(pady=(15, 10))
        
        btns_col4 = [
            ("🚚 Entrada NF (POENF)", "POENF"),
            ("🤝 Fornecedores (POFor)", "POFor"),
            ("🕒 Períodos Caixa (POPer)", "POPer"),
            ("🔥 Tipos Combustível (POTCo)", "POTCo"),
            ("📝 Total Compra/Venda (POTCV)", "POTCV")
        ]
        for txt, key in btns_col4:
            cmd = lambda k=key: self.hub.abrir_tabela_posto(k)
            ctk.CTkButton(self.col4, text=txt, width=btn_width, height=btn_height, font=("Arial", 12, "bold"), command=cmd, fg_color="#1E88E5", hover_color="#1565C0").pack(pady=5, padx=15)




        # Rodapé com Fechar e Copiar
        self.bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_bar.pack(side="bottom", fill="x", padx=20, pady=10)
        
        self.btn_voltar = ctk.CTkButton(self.bottom_bar, text="✖  Fechar Módulo", width=140, height=35, command=self.fechar_tela, fg_color="transparent", border_width=2, border_color="black", text_color="black", hover_color="#E0E0E0", font=("Arial", 12, "bold"))
        self.btn_voltar.pack(side="left")
        
        self.frame_id = ctk.CTkFrame(self.bottom_bar, fg_color="transparent")
        self.frame_id.pack(side="right")
        
        ctk.CTkLabel(self.frame_id, text=f"Tela: {self.id_str}", font=("Arial", 11, "bold"), text_color="gray").pack(side="left")
        
        self.btn_copy_id = ctk.CTkButton(self.frame_id, text="📋", width=25, height=25, fg_color="transparent", hover_color="#E0E0E0", text_color="black", font=("Arial", 11, "bold"), command=self.copiar_id)
        self.btn_copy_id.pack(side="left", padx=5)
        ToolTip(self.btn_copy_id, "COPIAR NOME TELA")

    def copiar_id(self):
        self.clipboard_clear()
        self.clipboard_append(self.id_str)
        self.update()
        from tkinter import messagebox
        messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")
        
    def fechar_tela(self):
        if hasattr(self, 'hub'):
            self.hub.fechar_modulo_atual()
        else:
            self.destroy()



def centralizar_tela(tela, w, h):
    x = (tela.winfo_screenwidth() // 2) - (w // 2)
    y = (tela.winfo_screenheight() // 2) - (h // 2)
    tela.geometry(f"{w}x{h}+{x}+{y}")


# =====================================================================
# TELA 4: ANÁLISE DE VENDAS POR PRODUTO (CFG_PRODS)
# =====================================================================
# NOVAS TELAS: BASE E ESPECIALIZADAS
# =====================================================================



POSTO_TABLES_CONFIG = {
    "POEmp": {"titulo": "Cadastro de Postos", "id_tela": "CAD_POSTO", "tabela": "POEmp", "colunas": ["POEmpCod", "POEmpDes", "POEmpRazSo", "POEmpCGC", "PoEmpEnd", "POEmpI_E"], "cabecalhos": ["Empresa", "Descrição", "Razão Social", "CNPJ", "Endereço", "Ins. Estadual"], "widths": [80, 180, 220, 130, 200, 130], "aligns": ["center", "w", "w", "center", "w", "center"]},
    "POBom": {"titulo": "Cadastro de Bombas", "id_tela": "CAD_BOMBAS", "tabela": "POBom", "colunas": ["POBomCod", "POBomDes", "POBomVlrUn", "POBomTnqCo"], "cabecalhos": ["Código", "Descrição", "Valor Unitário Atual", "N° do Tanque"], "widths": [80, 200, 150, 100], "aligns": ["center", "w", "e", "center"]},
    "POBIT": {"titulo": "Intervenção Técnica", "id_tela": "CAD_BICOS", "tabela": "POBIT", "colunas": ["POBomCod", "POMBoCod", "POBomDes", "POBITNroIn", "POBITDta", "POBITMot", "POBITNom", "POBITCNPJ", "POBITCPF"], "cabecalhos": ["Código Bico", "Código Bomba", "Descrição Bomba", "N° da Intervenção Técnica", "Data da Intervenção Técnica", "Motivo da Intervenção Técnica", "Nome do Interventor", "CNPJ da Empresa de Intervenção", "CPF do Interventor"], "widths": [100, 100, 180, 180, 150, 250, 180, 160, 140], "aligns": ["center", "center", "w", "center", "center", "w", "w", "center", "center"]},
    "POTnq": {"titulo": "Cadastro de Tanques", "id_tela": "CAD_TANQUES", "tabela": "POTnq", "colunas": ["POTnqCod", "POTnqDes", "POTnqDesLM", "POLvrCod", "POTCoDes", "POTnqCusUn", "POTnqQtd"], "cabecalhos": ["Tanque", "Descrição", "Desc. LMC", "Livro", "Tipo Combustível", "Custo Úl. Ent.", "Capacidade (Lts)"], "widths": [80, 180, 150, 80, 180, 150, 150], "aligns": ["center", "w", "w", "center", "w", "e", "e"]},
    "POLac": {"titulo": "Cadastro de Lacres", "id_tela": "CAD_LACRES", "tabela": "POLac", "colunas": ["POEmpCod", "POMBoCod", "POLacNro", "POLacDta"], "cabecalhos": ["Posto", "Bomba", "Nro Lacre", "Data"], "widths": [80, 80, 150, 120], "aligns": ["center", "center", "center", "center"]},
    "POBxB": {"titulo": "Vínculo Bico/Bomba", "id_tela": "VIN_BICO_BOMBA", "tabela": "POBxB", "colunas": ["POBomCod", "BicoDes", "POMBoCod", "BombaDes", "POBxBSts"], "cabecalhos": ["Cód. Bico", "Descrição Bico", "Cód. Bomba", "Descrição Bomba", "Status"], "widths": [100, 200, 100, 200, 80], "aligns": ["center", "w", "center", "w", "center"]},
    "POVel1": {"titulo": "Velocímetros", "id_tela": "VELOCIMETROS", "tabela": "POVel1", "colunas": ["POEmpCod", "POBomCod", "POVel1Dta", "POVel1VlrI", "POVel1VlrF"], "cabecalhos": ["Posto", "Bomba", "Data", "Inicial", "Final"], "widths": [60, 60, 120, 120, 120], "aligns": ["center", "center", "center", "e", "e"]},
    "POVel2": {"titulo": "Hist. Velocímetros", "id_tela": "HIST_VELOCIMETROS", "tabela": "POVel2", "colunas": ["POEmpCod", "POBomCod", "POVel2Dta", "POVel2Hor", "POVel2Vlr"], "cabecalhos": ["Posto", "Bomba", "Data", "Hora", "Leitura"], "widths": [60, 60, 100, 80, 120], "aligns": ["center", "center", "center", "center", "e"]},
    "POLMC1": {"titulo": "LMC 1: Venda p/ Bomba", "id_tela": "LMC_VENDA_BOMBA", "tabela": "POLMC1", "colunas": ["POEmpCod", "POLMC1Dta", "POBomCod", "POLMC1Qtd"], "cabecalhos": ["Posto", "Data", "Bomba", "Quantidade"], "widths": [80, 120, 80, 120], "aligns": ["center", "center", "center", "e"]},
    "POLMC2": {"titulo": "LMC 2: Estoque/Entrada", "id_tela": "LMC_ESTOQUE", "tabela": "POLMC2", "colunas": ["POTnqCod", "PODtaMov", "POLMC2EIn", "POLMC2E_S"], "cabecalhos": ["Tanque", "Data Mov", "Est. Inicial", "Entr/Saída"], "widths": [80, 120, 120, 120], "aligns": ["center", "center", "e", "e"]},
    "POLMC3": {"titulo": "LMC 3: Resumo p/ Prod", "id_tela": "LMC_RESUMO", "tabela": "POLMC3", "colunas": ["POEmpCod", "POLMC3Dta", "POTCoCod", "POLMC3Vnd"], "cabecalhos": ["Posto", "Data", "Tipo", "Total Venda"], "widths": [80, 120, 80, 150], "aligns": ["center", "center", "center", "e"]},
    "POLvr": {"titulo": "Livros do LMC", "id_tela": "LMC_LIVROS", "tabela": "POLvr", "colunas": ["POLvrCod", "POLvrDes"], "cabecalhos": ["Cód", "Descrição do Livro"], "widths": [80, 300], "aligns": ["center", "w"]},
    "POCF1": {"titulo": "⭐ Movimentos do Posto", "id_tela": "MOV_POSTO", "tabela": "POCF1", "colunas": ["POCF1Tst", "POCF1Usu", "POCF1BomCo", "POCF1Qtd", "POCF1VlrMo", "POCF1Sts"], "cabecalhos": ["Data/Hora", "Frentista", "Bomba", "Qtde", "Valor", "Sts"], "widths": [160, 100, 70, 90, 100, 50], "aligns": ["center", "w", "center", "e", "e", "center"]},
    "POCF2": {"titulo": "Caixa Frentista (Resumo)", "id_tela": "CX_FRENTISTA_RES", "tabela": "POCF2", "colunas": ["POCF2Usu", "POCF2Tst", "POCF2Sts", "POCF2VlrMo", "POCF2TSTFe"], "cabecalhos": ["Frentista", "Abertura", "Sts", "Total", "Fechamento"], "widths": [120, 160, 60, 120, 160], "aligns": ["w", "center", "center", "e", "center"]},
    "POCF3": {"titulo": "Totais Frentista", "id_tela": "CX_FRENTISTA_TOT", "tabela": "POCF3", "colunas": ["POCF2Usu", "POCF3TipMo", "POCF3VlrFi"], "cabecalhos": ["Frentista", "Tipo Mov.", "Valor Total"], "widths": [120, 100, 120], "aligns": ["w", "center", "e"]},
    "POCF4": {"titulo": "Itens do Caixa", "id_tela": "CX_FRENTISTA_ITENS", "tabela": "POCF4", "colunas": ["POCF4Tst", "POCF4Usu", "POCF4ProDe", "POCF4Qtd", "POCF4VlrMo"], "cabecalhos": ["Data/Hora", "Frentista", "Produto", "Qtde", "Valor"], "widths": [160, 100, 200, 90, 100], "aligns": ["center", "w", "w", "e", "e"]},
    "POCxa1": {"titulo": "Caixa Geral do Posto", "id_tela": "CX_GERAL_POSTO", "tabela": "POCxa1", "colunas": ["POCxa1DCx", "POCxa1Sts", "POCxa1Vlr"], "cabecalhos": ["Data Caixa", "Status", "Valor Total"], "widths": [120, 80, 150], "aligns": ["center", "center", "e"]},
    "POCxa2": {"titulo": "Detalhes Caixa Geral", "id_tela": "CX_GERAL_DET", "tabela": "POCxa2", "colunas": ["POCxa1DCx", "POCxa2Tip", "POCxa2Vlr"], "cabecalhos": ["Data Caixa", "Tipo", "Valor"], "widths": [120, 100, 150], "aligns": ["center", "center", "e"]},
    "POVal": {"titulo": "Cadastro de Vales", "id_tela": "CAD_VALES", "tabela": "POVal", "colunas": ["POValCod", "POValDta", "POValUsu", "POValVlr"], "cabecalhos": ["Cód", "Data", "Frentista", "Valor Vale"], "widths": [80, 100, 120, 120], "aligns": ["center", "center", "w", "e"]},
    "POENF": {"titulo": "Entrada de Notas (Posto)", "id_tela": "ENTRADA_NF_POSTO", "tabela": "POENF", "colunas": ["PODtaMov", "POForCod", "FOR_DES", "POForNNF", "POForDNF", "POTnqCod", "TNQ_DES", "POENFQtd", "POENFVlrTo", "POENFSts"], "cabecalhos": ["📅 Movimento", "🔑 Cód. Forn", "🤝 Fornecedor", "📄 Nro Nota", "📅 Data Nota", "🛢️ Tanque", "📝 Tanque Det.", "📦 Qtd", "💰 Vlr. Total", "🔘 Status"], "widths": [110, 80, 200, 110, 110, 70, 150, 90, 120, 90], "aligns": ["center", "center", "w", "center", "center", "center", "w", "center", "e", "center"]},
    "POFor": {"titulo": "Fornecedores Posto", "id_tela": "FORNECEDORES_POSTO", "tabela": "POFor", "colunas": ["POForCod", "POForDes"], "cabecalhos": ["Cód", "Razão Social"], "widths": [80, 300], "aligns": ["center", "w"]},
    "POPer": {"titulo": "Períodos de Caixa", "id_tela": "PERIODOS_CAIXA", "tabela": "POPer", "colunas": ["POEmpCod", "POPerCod", "POPerDes"], "cabecalhos": ["Posto", "Período", "Descrição"], "widths": [80, 80, 200], "aligns": ["center", "center", "w"]},
    "POTCo": {"titulo": "Tipos Combustível", "id_tela": "TIPOS_COMBUSTIVEL", "tabela": "POTCo", "colunas": ["POTCoCod", "POTCoDes"], "cabecalhos": ["Cód", "Tipo de Combustível"], "widths": [80, 250], "aligns": ["center", "w"]},
    "POTCV": {"titulo": "Total Compra/Venda", "id_tela": "TOT_COMPRA_VENDA", "tabela": "POTCV", "colunas": ["POEmpCod", "POTCVDta", "POTCVVnd", "POTCVCom"], "cabecalhos": ["Posto", "Data", "Vendas", "Compras"], "widths": [80, 120, 120, 120], "aligns": ["center", "center", "e", "e"]}
}

POSTO_FIELD_LABELS = {
    "POEmpCod": "Empresa",
    "POEmpDes": "Descrição Empresa",
    "POEmpDtaUs": "Data Atual do Movimento",
    "POEmpPer": "Período Atual",
    "POEmpRazSo": "Razão Social",
    "PoEmpEnd": "Endereço",
    "POEmpCGC": "C.N.P.J.",
    "POEmpI_E": "Inscrição Estadual",
    "POEmpQtdPe": "Quantidade de Períodos",
    "POMBoCod": "Bomba",
    "POBomCod": "Bico",
    "POBITNroIn": "N° da Intervenção Técnica",
    "POBITDta": "Data da Intervenção Técnica",
    "POBITMot": "Motivo da Intervenção Técnica",
    "POBITNom": "Nome do Interventor",
    "POBITCNPJ": "CNPJ da Empresa de Intervenção",
    "POBITCPF": "CPF do Interventor",
    "POTnqCod": "Tanque",
    "POTnqDes": "Descrição",
    "POTnqDesLM": "Descrição Tanque no LMC",
    "POLvrCod": "Livro",
    "POTnqTipCo": "Tipo de Combustível",
    "POTnqCusUn": "Custo Unitário da Última Entrada da Nota",
    "POTnqConSp": "Código do Tanque para Consolidar no Speed",
    "POTnqQtd": "Quantidade de Litros que Cabe no Tanque",
    "PODtaMov": "Data do Movimento",
    "POForCod": "Código do Fornecedor",
    "POForNNF": "Número da Nota Fiscal",
    "POForDNF": "Data da Nota Fiscal",
    "POENFQtd": "Quantidade de Entrada",
    "POENFVlrTo": "Valor Total da Nota",
    "POENFVlrUn": "Valor Unitário do Produto",
    "POENFObs": "Observações da Nota",
    "POENFSts": "Situação (A-Aberto, F-Fechado, C-Cancelado)",
    "POENFDtaPg": "Data do Pagamento",
    "POENFOrg": "Origem (0-Comb, 1-Lub, 9-Div)",
    "POENFNFE": "Chave de Acesso NFE (60 dígitos)",
    "POENFBCO": "Código do Banco",
    "POENFCHQ": "Número do Cheque"
}

class FormularioPostoWindow(ctk.CTkToplevel):
    def __init__(self, parent, config_db, table_info, pk_value=None):
        super().__init__(parent)
        self.config_db = config_db
        self.table_info = table_info
        self.pk_value = pk_value # Valor da PK se for ALTERAR
        
        self.title(f"{'📝 Alterar' if pk_value else '✚ Incluir'} - {table_info['titulo']} (Todos os Campos)")
        self.geometry("650x800")
        self.grab_set() 
        centralizar_tela(self, 650, 800)

        self.main_frame = ctk.CTkFrame(self, fg_color="#F1F5F9", corner_radius=15)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Área de rolagem para suportar todos os campos da tabela
        self.scroll_frame = ctk.CTkScrollableFrame(self.main_frame, fg_color="transparent")
        self.scroll_frame.pack(fill="both", expand=True, padx=10, pady=(10, 80))
        
        ctk.CTkLabel(self.scroll_frame, text=f"{'📝 EDITAR REGISTRO COMPLETO' if pk_value else '✚ NOVO REGISTRO COMPLETO'}", font=("Arial", 18, "bold"), text_color="#1E293B").pack(pady=20)
        
        self.fields = {}
        self.all_columns = []
        self.carregar_layout_dinamico()
        
        # Nome Exclusivo da Tela e Botão Copiar (Canto Inferior Direito)
        self.id_manut = f"MANUT_{table_info['id_tela']}"
        self.footer_form = ctk.CTkFrame(self.main_frame, fg_color="transparent", height=30)
        self.footer_form.place(relx=1.0, rely=1.0, x=-20, y=-10, anchor="se")
        
        ctk.CTkLabel(self.footer_form, text=f"ID: {self.id_manut}", font=("Arial", 10, "bold"), text_color="gray").pack(side="left", padx=5)
        btn_copy = ctk.CTkButton(self.footer_form, text="📋", width=25, height=25, fg_color="transparent", hover_color="#E2E8F0", text_color="#64748B", font=("Arial", 11, "bold"), command=self.copiar_id_form)
        btn_copy.pack(side="left")

        # Barra de Ações (Fixo no rodapé)
        self.actions_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.actions_frame.place(relx=0.5, rely=0.90, anchor="center", relwidth=0.8)
        
        self.btn_cancel = ctk.CTkButton(self.actions_frame, text="✖ Fechar sem Salvar", font=("Arial", 14, "bold"), height=48, fg_color="#E53935", hover_color="#B71C1C", command=self.destroy)
        self.btn_cancel.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.btn_save = ctk.CTkButton(self.actions_frame, text="💾 SALVAR", font=("Arial", 14, "bold"), height=48, fg_color="#2E7D32", hover_color="#1B5E20", command=self.salvar)
        self.btn_save.pack(side="left", fill="x", expand=True)

    def copiar_id_form(self):
        self.clipboard_clear()
        self.clipboard_append(self.id_manut)
        from tkinter import messagebox
        messagebox.showinfo("Copiado", f"ID da tela copiado: {self.id_manut}")

    def carregar_layout_dinamico(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            tabela = self.table_info['tabela']
            posto_id = 1
            if hasattr(self.master, 'hub'): posto_id = self.master.hub.get_selected_posto_id()
            
            # 1. Buscar todas as colunas da tabela no banco
            cur.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{tabela}' ORDER BY ORDINAL_POSITION")
            self.all_columns = [row[0] for row in cur.fetchall()]
            
            # 2. Se for alteração, buscar os dados atuais de TODAS as colunas
            full_data = {}
            if self.pk_value:
                # CASO ESPECIAL: POBIT (Chave Composta)
                if tabela.upper() == "POBIT" and isinstance(self.pk_value, (list, tuple)):
                    sql = f"SELECT * FROM {tabela} WHERE POEmpCod = ? AND POBomCod = ? AND POBITNroIn = ?"
                    cur.execute(sql, self.pk_value)
                # CASO ESPECIAL: POBXB (Chave Composta)
                elif tabela.upper() == "POBXB" and isinstance(self.pk_value, (list, tuple)):
                    sql = f"SELECT * FROM {tabela} WHERE POEmpCod = ? AND POMBoCod = ? AND POBomCod = ?"
                    cur.execute(sql, self.pk_value)
                elif isinstance(self.pk_value, (list, tuple)):
                    # CASO GERAL CHAVE COMPOSTA (Dinâmico: Suporta POENF com 4 campos)
                    key_cols = self.all_columns[:len(self.pk_value)]
                    where_clauses = [f"{col} = ?" for col in key_cols]
                    sql = f"SELECT * FROM {tabela} WHERE {' AND '.join(where_clauses)}"
                    cur.execute(sql, self.pk_value)
                else:
                    # Caso Padrão (Chave Simples)
                    sql = f"SELECT * FROM {tabela} WHERE {self.all_columns[0]} = ?"
                    cur.execute(sql, self.pk_value)
                
                row = cur.fetchone()
                if row:
                    full_data = {col: val for col, val in zip(self.all_columns, row)}
                else:
                    print(f"DEBUG: Registro não encontrado na tabela {tabela} para PK {self.pk_value}")
            
            # 3. Gerar campos na UI
            for i, col in enumerate(self.all_columns):
                frame = ctk.CTkFrame(self.scroll_frame, fg_color="transparent")
                frame.pack(fill="x", padx=20, pady=4)
                
                lbl_color = "#1565C0"; lbl_font = ("Arial", 12, "bold")
                exibir_nome = POSTO_FIELD_LABELS.get(col, col)
                ctk.CTkLabel(frame, text=f"{exibir_nome}:", font=lbl_font, text_color=lbl_color, width=180, anchor="w").pack(side="left")
                
                # --- NOVO: Lógica de ComboBox para Campos Específicos (POBXB e POENF) ---
                use_combo = False
                if tabela.upper() == "POBXB" and col.upper() in ["POMBOCOD", "POBOMCOD"]: use_combo = True
                if tabela.upper() == "POENF" and col.upper() in ["POTNQCOD", "POFORCOD"]: use_combo = True
                
                if use_combo:
                    options = []
                    try:
                        if col.upper() == "POMBOCOD" or (tabela.upper() == "POBXB" and col.upper() == "POBOMCOD"):
                            cur.execute("SELECT POBomCod, POBomDes FROM POBom WHERE POEmpCod = ?", (posto_id,))
                            options = [f"{r[0]} - {str(r[1]).strip()}" for r in cur.fetchall()]
                        elif col.upper() == "POTNQCOD":
                            cur.execute("SELECT POTnqCod, POTnqDes FROM POTnq WHERE POEmpCod = ?", (posto_id,))
                            options = [f"{r[0]} - {str(r[1]).strip()}" for r in cur.fetchall()]
                        elif col.upper() == "POFORCOD":
                            cur.execute("SELECT POForCod, POForDes FROM POFor ORDER BY POForDes")
                            options = [f"{r[0]} - {str(r[1]).strip()}" for r in cur.fetchall()]
                    except Exception as e: print(f"DEBUG Erro carregar combos form: {e}")
                    
                    if not options: options = ["1", "2"]
                    
                    widget = ctk.CTkComboBox(frame, values=options, font=("Arial", 13), height=32, dropdown_font=("Arial", 13))
                    widget.pack(side="left", fill="x", expand=True)
                    
                    val = full_data.get(col, "")
                    if val is not None:
                        # Se já temos o valor, tentamos setar o item que começa com o ID
                        found = False
                        for opt in options:
                            if opt.startswith(f"{val} -"):
                                widget.set(opt)
                                found = True
                                break
                        if not found: widget.set(str(val))
                    
                    if self.pk_value and i in [1, 2]: 
                         widget.configure(state="disabled")
                
                else:
                    widget = ctk.CTkEntry(frame, font=("Arial", 13), height=32, border_color="#CBD5E1")
                    widget.pack(side="left", fill="x", expand=True)
                    widget.bind("<KeyRelease>", self.force_uppercase)
                    
                    if self.pk_value:
                        val = full_data.get(col, "")
                        if val is None: val = ""
                        elif hasattr(val, 'strftime'): val = val.strftime("%d/%m/%Y")
                        widget.insert(0, str(val).upper())
                        
                        # Travar Empresa ou chaves iniciais se for alteração
                        if i == 0: 
                            widget.configure(state="disabled", fg_color="#E2E8F0", text_color="#94A3B8")
                
                self.fields[col] = widget

            conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro de Schema", f"Não foi possível mapear os campos da tabela {tabela}:\n\n{str(e)}")

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro de Schema", f"Não foi possível mapear os campos da tabela {tabela}:\n\n{str(e)}")

    def force_uppercase(self, event):
        # Diferenciando se o widget é um CTkEntry (o bind passa o widget interno as vezes)
        widget = event.widget
        try:
            pos = widget.index("insert")
            val = widget.get().upper()
            widget.delete(0, "end")
            widget.insert(0, val)
            widget.icursor(pos)
        except: pass

    def salvar(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            tabela = self.table_info['tabela']
            # Extrair valores tratando o caso de "ID - Descrição" nos ComboBoxes
            vals = []
            for col in self.all_columns:
                raw = self.fields[col].get().upper()
                if " - " in raw:
                    raw = raw.split(" - ")[0].strip()
                vals.append(raw)
            
            if self.pk_value: # UPDATE
                # Caso especial para chaves compostas no WHERE
                if tabela.upper() == "POBIT" and isinstance(self.pk_value, (list, tuple)):
                    pk_where = "POEmpCod = ? AND POBomCod = ? AND POBITNroIn = ?"
                elif tabela.upper() == "POBXB" and isinstance(self.pk_value, (list, tuple)):
                    pk_where = "POEmpCod = ? AND POMBoCod = ? AND POBomCod = ?"
                elif isinstance(self.pk_value, (list, tuple)):
                    pk_where = f"POEmpCod = ? AND {self.all_columns[1]} = ?"
                else:
                    pk_where = f"{self.all_columns[0]} = ?"
                
                # Campos para SET (todos exceto os que compõem a PK, mas por simplicidade aqui mandamos todos)
                # O ideal no SQL Server é não atualizar colunas de PK, mas aqui vamos pular a primeira se for Empresa
                update_cols = self.all_columns
                set_parts = [f"{c} = ?" for c in update_cols]
                sql = f"UPDATE {tabela} SET {', '.join(set_parts)} WHERE {pk_where}"
                
                params = vals + (list(self.pk_value) if isinstance(self.pk_value, (list, tuple)) else [self.pk_value])
            else: # INSERT
                placeholders = ", ".join(["?" for _ in self.all_columns])
                sql = f"INSERT INTO {tabela} ({', '.join(self.all_columns)}) VALUES ({placeholders})"
                params = vals
            
            cur.execute(sql, params)
            conn.commit()
            conn.close()
            
            from tkinter import messagebox
            messagebox.showinfo("Sucesso", "Registro atualizado com sucesso!")
            if hasattr(self.master, "carregar_dados"):
                self.master.carregar_dados()
            
            # Sincronização Global: Se alterou dados de posto (POEmp), atualizar Hub principal
            if self.table_info['tabela'] == "POEmp" and hasattr(self.master, 'hub'):
                try: self.master.hub.carregar_postos_hub()
                except: pass

            self.destroy()
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao gravar dados:\n\n{str(e)}")

class TotCompraVendaWindow(BaseWindow):
    """Tela especializada para Total Compra/Venda (POTCV) com filtro de período e totais."""
    def __init__(self, parent, config):
        super().__init__(parent, "Total Compra/Venda", "TOT_COMPRA_VENDA")
        self.config_db = config

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        # --- FILTROS DE PERÍODO ---
        flt = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        flt.pack(side="left", padx=20)

        ctk.CTkLabel(flt, text="De:", font=("Arial", 12, "bold")).pack(side="left", padx=(0, 5))
        self.entry_dt_ini = ctk.CTkEntry(flt, width=100, font=("Arial", 12), placeholder_text="dd/mm/aaaa")
        self.entry_dt_ini.insert(0, ini_mes)
        self.entry_dt_ini.pack(side="left", padx=(0, 10))

        ctk.CTkLabel(flt, text="Até:", font=("Arial", 12, "bold")).pack(side="left", padx=(0, 5))
        self.entry_dt_fim = ctk.CTkEntry(flt, width=100, font=("Arial", 12), placeholder_text="dd/mm/aaaa")
        self.entry_dt_fim.insert(0, fim_mes)
        self.entry_dt_fim.pack(side="left", padx=(0, 10))

        ctk.CTkButton(flt, text="🔍 Filtrar", width=100, height=30, fg_color="#1565C0", hover_color="#0D47A1", font=("Arial", 12, "bold"), command=self.carregar_dados).pack(side="left", padx=5)

        # --- GRID ---
        cols = ["POTCVDta", "POTCVVnd", "POTCVCom"]
        hdrs = ["📅 Data", "💰 Vendas", "🛒 Compras"]
        wids = [150, 180, 180]
        alns = ["center", "e", "e"]
        self.configurar_grid(columns=cols, headings=hdrs, widths=wids, aligns=alns)

        # --- RODAPÉ DE TOTAIS (Padrão Ouro - row=2) ---
        totais_frame = ctk.CTkFrame(self, height=35, fg_color="transparent", corner_radius=0)
        totais_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 0))
        self.tree_totais = ttk.Treeview(totais_frame, columns=cols, show="", height=1, style="TotaisTCV.Treeview")
        self.tree_totais.column("#0", width=0, stretch=False)
        for c, w, a in zip(cols, wids, alns):
            self.tree_totais.column(c, width=w, minwidth=w, anchor=a)
        st = ttk.Style()
        st.configure("TotaisTCV.Treeview", background="#1565C0", foreground="white", font=("Arial", 12, "bold"), rowheight=28)
        self.tree_totais.pack(fill="x", expand=True)

        self.after(200, self.carregar_dados)

    def get_current_query(self):
        from datetime import datetime
        posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
        dt_i = datetime.strptime(self.entry_dt_ini.get().strip(), "%d/%m/%Y").strftime("%Y%m%d")
        dt_f = datetime.strptime(self.entry_dt_fim.get().strip(), "%d/%m/%Y").strftime("%Y%m%d")
        sql = (
            f"SELECT POTCVDta, POTCVVnd, POTCVCom "
            f"FROM POTCV "
            f"WHERE POEmpCod = {posto_id} "
            f"AND POTCVDta BETWEEN '{dt_i}' AND '{dt_f}' "
            f"ORDER BY POTCVDta DESC"
        )
        return sql

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: TOTCOMPRAVENDA ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import datetime
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()

            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            dt_i = datetime.strptime(self.entry_dt_ini.get().strip(), "%d/%m/%Y").strftime("%Y%m%d")
            dt_f = datetime.strptime(self.entry_dt_fim.get().strip(), "%d/%m/%Y").strftime("%Y%m%d")

            sql = "SELECT POTCVDta, POTCVVnd, POTCVCom FROM POTCV WHERE POEmpCod = ? AND POTCVDta BETWEEN ? AND ? ORDER BY POTCVDta DESC"
            self.ultima_query_bruta = sql.replace("?", f"'{posto_id}'", 1).replace("?", f"'{dt_i}'", 1).replace("?", f"'{dt_f}'", 1)
            cur.execute(sql, (posto_id, dt_i, dt_f))
            rows = cur.fetchall()

            for item in self.tree.get_children(): self.tree.delete(item)

            def fmt_moeda(v):
                return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            total_vnd, total_com = 0.0, 0.0
            for r in rows:
                dt_val = r[0].strftime("%d/%m/%Y") if hasattr(r[0], 'strftime') else str(r[0])
                vnd = float(r[1] or 0)
                com = float(r[2] or 0)
                total_vnd += vnd
                total_com += com
                self.tree.insert("", "end", values=(dt_val, fmt_moeda(vnd), fmt_moeda(com)))

            # Atualizar tree_totais
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            self.tree_totais.insert("", "end", values=(
                f"REGISTROS: {len(rows)}", fmt_moeda(total_vnd), fmt_moeda(total_com)))

            self.re_zebrar(); conn.close()
        except Exception as e:
            print(f"Erro ao carregar POTCV: {e}")
            from tkinter import messagebox
            messagebox.showerror("Erro", f"Erro ao carregar Total Compra/Venda:\n\n{str(e)}")

class TabelaPostoWindow(BaseWindow):
    def __init__(self, parent, config, info):
        super().__init__(parent, info['titulo'], info['id_tela'])
        self.config_db = config
        self.table_info = info
        
        # Adicionar Botão Incluir na Barra Superior (Top Frame)
        self.btn_incluir = ctk.CTkButton(self.top_frame, text="➕ Incluir Novo", width=140, height=35, font=("Arial", 12, "bold"), fg_color="#2E7D32", hover_color="#1B5E20", command=self.abrir_formulario_incluir)
        self.btn_incluir.pack(side="right", padx=20, pady=10)

        self.configurar_grid(
            columns=info['colunas'],
            headings=info['cabecalhos'],
            widths=info['widths'],
            aligns=info['aligns']
        )
        
        # --- BARRA AZUL DE REGISTROS (Padrão Ouro - row=2) ---
        # Remove espaço entre grid e barra e tira padding horizontal para ocupar tela inteira
        self.grid_frame.grid_configure(padx=0, pady=(0, 0))
        self.bar_regs = ctk.CTkFrame(self, height=30, fg_color="#1565C0", corner_radius=0)
        self.bar_regs.grid(row=2, column=0, sticky="ew", padx=0, pady=(0, 5))
        self.lbl_bar_regs = ctk.CTkLabel(self.bar_regs, text="REGISTROS: 0", font=("Arial", 12, "bold"), text_color="white")
        self.lbl_bar_regs.pack(side="left", padx=10, pady=2)
        
        # Ocultar o contador da bottom_bar (substituído pela barra azul)
        self.lbl_total_regs.pack_forget()
        
        # Bind do Duplo Clique para Alterar
        self.tree.bind("<Double-1>", self.abrir_formulario_alterar)
        
        self.after(100, self.carregar_dados)

    def abrir_formulario_incluir(self):
        FormularioPostoWindow(self, self.config_db, self.table_info)

    def abrir_formulario_alterar(self, event):
        try:
            item = self.tree.identify_row(event.y)
            if not item: return
            
            row_values = self.tree.item(item, "values")
            if not row_values: return
            
            tab = self.table_info['tabela'].upper()
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1

            if tab == "POBIT":
                pk_val = [posto_id, row_values[0], row_values[3]]
            elif tab == "POTNQ":
                pk_val = [posto_id, row_values[0]]
            elif tab == "POBXB":
                pk_val = [posto_id, row_values[2], row_values[0]]
            elif tab == "POEMP":
                pk_val = row_values[0]
            else:
                pk_val = [posto_id, row_values[0]]
                
            FormularioPostoWindow(self, self.config_db, self.table_info, pk_value=pk_val)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro ao Abrir Manutenção", f"Não foi possível abrir a tela de edição:\n\n{str(e)}")
            print(f"DEBUG Erro abrir manutenção {self.table_info.get('tabela')}: {e}")

    def get_sql_summary(self):
        """Retorna o resumo técnico da lógica de negócio desta tela no Padrão Ouro."""
        titulo = self.table_info.get('titulo', self.table_info.get('tabela', '')).upper()
        tabela = self.table_info.get('tabela', 'N/A')
        colunas = ", ".join(self.table_info.get('colunas', []))
        return (
            f"--- ESTRUTURA SQL: {titulo} ---\n"
            f"TABELA PRINCIPAL: {tabela} (Cadastro Dinâmico)\n"
            "RELACIONAMENTOS:\n"
            "- Operações diretas na tabela base na maioria dos casos.\n"
            "- Casos especiais (como Tanques/Bicos) recebem JOINs específicos no método carregar_dados.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            f"- {colunas}\n\n"
            "FILTROS APLICADOS:\n"
            "- Posto/Empresa: (ID Selecionado no Hub globalmente)\n\n"
            "OPERAÇÕES SUPORTADAS:\n"
            "- MODO CRUD Ativo: Permite Inclusão (Insert), Edição (Update) e Exclusão."
        )

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            cols = ", ".join(self.table_info['colunas'])
            tab = self.table_info['tabela']
            
            # Pega o Posto selecionado globalmente no Dashboard
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            
            # --- CASO ESPECIAL: CAD_BICOS (POBIT) com JOIN para Bomba ---
            if tab.upper() == "POBIT":
                sql = """
                    SELECT 
                        t.POBomCod, 
                        v.POMBoCod, 
                        b.POBomDes, 
                        t.POBITNroIn, 
                        t.POBITDta, 
                        t.POBITMot, 
                        t.POBITNom, 
                        t.POBITCNPJ, 
                        t.POBITCPF
                    FROM POBIT t
                    LEFT JOIN POBxB v ON t.POEmpCod = v.POEmpCod AND t.POBomCod = v.POBomCod
                    LEFT JOIN POBom b ON v.POEmpCod = b.POEmpCod AND v.POMBoCod = b.POBomCod
                    WHERE t.POEmpCod = ?
                    ORDER BY t.POBomCod DESC
                """
                params = (posto_id,)
                # Debug SQL Bruto
                self.ultima_query_bruta = sql.replace("?", f"'{params[0]}'")
                cur.execute(sql, params)
            
            # --- CASO ESPECIAL: CAD_TANQUES (POTnq) com JOIN para Tipo Combustível ---
            elif tab.upper() == "POTNQ":
                sql = """
                    SELECT 
                        t.POTnqCod, 
                        t.POTnqDes, 
                        t.POTnqDesLM, 
                        t.POLvrCod, 
                        c.POTCoDes, 
                        t.POTnqCusUn, 
                        t.POTnqQtd
                    FROM POTnq t
                    LEFT JOIN POTCo c ON t.POTnqTipCo = c.POTCoCod
                    WHERE t.POEmpCod = ?
                    ORDER BY t.POTnqCod DESC
                """
                params = (posto_id,)
                self.ultima_query_bruta = sql.replace("?", f"'{params[0]}'")
                cur.execute(sql, params)
            
            # --- CASO ESPECIAL: VIN_BICO_BOMBA (POBxB) com JOIN Duplo ---
            elif tab.upper() == "POBXB":
                sql = """
                    SELECT 
                        v.POBomCod, 
                        b1.POBomDes, 
                        v.POMBoCod, 
                        b2.POBomDes, 
                        v.POBxBSts
                    FROM POBxB v
                    LEFT JOIN POBom b1 ON v.POEmpCod = b1.POEmpCod AND v.POBomCod = b1.POBomCod
                    LEFT JOIN POBom b2 ON v.POEmpCod = b2.POEmpCod AND v.POMBoCod = b2.POBomCod
                    WHERE v.POEmpCod = ?
                    ORDER BY v.POMBoCod, v.POBomCod
                """
                self.ultima_query_bruta = sql.replace("?", str(posto_id))
                cur.execute(sql, (posto_id,))

            # --- CASO ESPECIAL: ENTRADA NF (POENF) com JOIN p/ Nome Fornecedor e Tanque ---
            elif tab.upper() == "POENF":
                sql = """
                    SELECT 
                        t.PODtaMov, t.POForCod, f.POForDes as FOR_DES, t.POForNNF, 
                        t.POForDNF, t.POTnqCod, nq.POTnqDes as TNQ_DES, t.POENFQtd, 
                        t.POENFVlrTo, t.POENFSts
                    FROM POENF t
                    LEFT JOIN POFor f ON t.POForCod = f.POForCod
                    LEFT JOIN POTnq nq ON t.POTnqCod = nq.POTnqCod AND t.POEmpCod = nq.POEmpCod
                    WHERE t.POEmpCod = ?
                    ORDER BY t.PODtaMov DESC
                """
                self.ultima_query_bruta = sql.replace("?", str(posto_id))
                cur.execute(sql, (posto_id,))

            # --- CASO GERAL: Sem JOINs (Tabelas Simples) ---
            elif tab.upper() != "POEMP":
                try:
                    # Verifica se a tabela possui o campo POEmpCod
                    cur.execute(f"SELECT TOP 0 POEmpCod FROM {tab}")
                    
                    # Se tiver POEmpCod, filtra pelo posto selecionado
                    sql = f"SELECT TOP 1000 {cols} FROM {tab} WHERE POEmpCod = ? ORDER BY {self.table_info['colunas'][0]} DESC"
                    self.ultima_query_bruta = sql.replace("?", str(posto_id))
                    cur.execute(sql, (posto_id,))
                except:
                    # Caso não tenha POEmpCod (EX: POFor, POLvr, POTCo...), carrega sem filtro
                    sql = f"SELECT TOP 1000 {cols} FROM {tab} ORDER BY {self.table_info['colunas'][0]} DESC"
                    self.ultima_query_bruta = sql
                    cur.execute(sql)
            else:
                # Caso seja POEmp, mostra todos os postos cadastrados
                sql = f"SELECT TOP 1000 {cols} FROM {tab} ORDER BY {self.table_info['colunas'][0]}"
                self.ultima_query_bruta = sql
                cur.execute(sql)

            rows = cur.fetchall()
            
            for item in self.tree.get_children(): self.tree.delete(item)
            
            for r in rows:
                fmt_vals = []
                for idx, v in enumerate(r):
                    if hasattr(v, 'strftime'): 
                         fmt_vals.append(v.strftime("%d/%m/%Y"))
                    elif isinstance(v, (int, float)):
                         col_name = str(cur.description[idx][0]).upper()
                         # Formatação para Moeda / Preços (2 ou 3 casas)
                         if "VLR" in col_name or "CUS" in col_name or "TOT" in col_name or "SAL" in col_name:
                             # Se for valor unitário, usar 3 casas decimais (Padrão Posto p/ Bicos e Notas)
                             dec = 3 if "UN" in col_name or "BOM" in col_name else 2
                             fmt = f"R$ {v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
                             fmt_vals.append(fmt)
                         elif "QTD" in col_name or "LMC" in col_name:
                             fmt_vals.append(f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                         else:
                             fmt_vals.append(v)
                    else: 
                         fmt_vals.append(str(v or ""))
                self.tree.insert("", "end", values=tuple(fmt_vals))
            
            self.re_zebrar()
            self.lbl_bar_regs.configure(text=f"REGISTROS: {len(rows)}")
            conn.close()
        except Exception as e:
            print(f"Erro ao carregar {self.table_info['tabela']}: {e}")

class EntradaNFPostoWindow(BaseWindow):
    def __init__(self, parent, config):
        self.table_info = POSTO_TABLES_CONFIG["POENF"]
        super().__init__(parent, self.table_info['titulo'], self.table_info['id_tela'])
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent # Fallback p/ MainHub
        
        # --- BARRA DE FILTROS (Premium) ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", fill="x", expand=True, padx=20, pady=(0, 5))

        # Filtro de Período
        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_frame, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(5, 2))
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        # Seletor de Tipo de Data
        ctk.CTkLabel(self.filter_frame, text="Filtrar por Data:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 2))
        self.combo_tipo_data = ctk.CTkComboBox(self.filter_frame, values=["Movimento", "Data Nota"], width=130, height=32, font=("Arial", 12))
        self.combo_tipo_data.pack(side="left", padx=2); self.combo_tipo_data.set("Movimento")

        # Filtro de Fornecedor (ComboBox Dinâmico conforme Grid)
        ctk.CTkLabel(self.filter_frame, text="🤝 Fornecedor:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 2))
        self.combo_fornecedor = ctk.CTkComboBox(self.filter_frame, values=["TODOS"], width=200, height=32, font=("Arial", 12))
        self.combo_fornecedor.pack(side="left", padx=2); self.combo_fornecedor.set("TODOS")

        # Busca Textual
        ctk.CTkLabel(self.filter_frame, text="🔍 Busca:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 2))
        self.ent_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Forn. ou Nro Nota...", width=180, height=32, font=("Arial", 13))
        self.ent_busca.pack(side="left", padx=2)
        self.ent_busca.bind("<Return>", lambda e: self.carregar_dados())

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=90, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        # Botão Incluir
        self.btn_incluir = ctk.CTkButton(self.top_frame, text="➕ Lançar Nota", width=140, height=35, font=("Arial", 12, "bold"), 
                                        fg_color="#2E7D32", hover_color="#1B5E20", command=self.abrir_formulario_incluir)
        self.btn_incluir.pack(side="right", padx=20, pady=10)

        self.configurar_grid(
            columns=self.table_info['colunas'],
            headings=self.table_info['cabecalhos'],
            widths=self.table_info['widths'],
            aligns=self.table_info['aligns']
        )
        
        self.tree.bind("<Double-1>", self.abrir_formulario_alterar)
        
        # --- RODAPÉ DE TOTAIS (Excel Style) ---
        # Utiliza o tree_totais de BaseWindow para alinhamento automático com as colunas
        self.after(200, self.carregar_dados)

    def abrir_formulario_incluir(self):
        FormularioPostoWindow(self, self.config_db, self.table_info)

    def abrir_formulario_alterar(self, event):
        item = self.tree.identify_row(event.y)
        if not item: return
        vals = self.tree.item(item, "values")
        posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
        
        # CHAVE COMPOSTA POENF: Empresa, Movimento(0), Tanque(5), Fornecedor(1), Nota(3)
        pk_val = [posto_id, vals[0], vals[5], vals[1], vals[3]]
        FormularioPostoWindow(self, self.config_db, self.table_info, pk_value=pk_val)

    def get_current_query(self):
        """Helper para auditoria SQL."""
        posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
        de_raw = self.ent_de.get().strip()
        ate_raw = self.ent_ate.get().strip()
        busca = self.ent_busca.get().strip().upper()
        
        where_extra = ""
        params = [posto_id]
        
        if de_raw and ate_raw:
            from datetime import datetime
            try:
                d_ini = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
                d_fim = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
                field_map = {"Movimento": "t.PODtaMov", "Data Nota": "t.POForDNF"}
                selected_field = field_map.get(self.combo_tipo_data.get(), "t.PODtaMov")
                where_extra += f" AND CAST({selected_field} AS DATE) BETWEEN ? AND ? "
                params.extend([d_ini, d_fim])
            except: pass

        forn_sel = self.combo_fornecedor.get()
        if forn_sel and forn_sel != "TODOS":
            try:
                cod_forn = int(forn_sel.split(" - ")[0])
                where_extra += " AND t.POForCod = ? "
                params.append(cod_forn)
            except: pass

        if busca:
            where_extra += " AND (f.POForDes LIKE ? OR CAST(t.POForNNF AS VARCHAR) LIKE ?) "
            params.extend([f"%{busca}%", f"%{busca}%"])

        sql = f"""
            SELECT 
                t.PODtaMov, t.POForCod, f.POForDes as FOR_DES, t.POForNNF, 
                t.POForDNF, t.POTnqCod, nq.POTnqDes as TNQ_DES, t.POENFQtd, 
                t.POENFVlrTo, t.POENFSts
            FROM POENF t
            LEFT JOIN POFor f ON t.POForCod = f.POForCod
            LEFT JOIN POTnq nq ON t.POTnqCod = nq.POTnqCod AND t.POEmpCod = nq.POEmpCod
            WHERE t.POEmpCod = ? {where_extra}
            ORDER BY t.PODtaMov DESC
        """.strip()
        return sql, params

    def carregar_combo_fornecedores(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            cur.execute("SELECT POForCod, POForDes FROM POFor ORDER BY POForDes")
            rows = cur.fetchall()
            conn.close()
            
            opcoes = ["TODOS"]
            for r in rows:
                if r[0] and r[1]:
                    opcoes.append(f"{int(r[0])} - {str(r[1]).strip().upper()}")
            self.combo_fornecedor.configure(values=opcoes)
        except Exception as e:
            print(f"Erro carregar combo fornecedores POENF: {e}")


    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            
            # Filtros
            de = self.ent_de.get().strip()
            ate = self.ent_ate.get().strip()
            busca = self.ent_busca.get().strip().upper()
            
            where_extra = ""
            params = [posto_id]
            
            if de and ate:
                from datetime import datetime
                d_ini = datetime.strptime(de, "%d/%m/%Y").strftime("%Y-%m-%d")
                d_fim = datetime.strptime(ate, "%d/%m/%Y").strftime("%Y-%m-%d")
                field_map = {"Movimento": "t.PODtaMov", "Data Nota": "t.POForDNF"}
                selected_field = field_map.get(self.combo_tipo_data.get(), "t.PODtaMov")
                where_extra += f" AND CAST({selected_field} AS DATE) BETWEEN ? AND ? "
                params.extend([d_ini, d_fim])
            
            # Filtro Fornecedor (Combo)
            forn_sel = self.combo_fornecedor.get()
            if forn_sel and forn_sel != "TODOS":
                try: 
                    cod_forn = int(forn_sel.split(" - ")[0])
                    where_extra += " AND t.POForCod = ? "
                    params.append(cod_forn)
                except: pass
            
            if busca:
                where_extra += " AND (f.POForDes LIKE ? OR CAST(t.POForNNF AS VARCHAR) LIKE ?) "
                params.extend([f"%{busca}%", f"%{busca}%"])

            sql = f"""
                SELECT 
                    t.PODtaMov, t.POForCod, f.POForDes as FOR_DES, t.POForNNF, 
                    t.POForDNF, t.POTnqCod, nq.POTnqDes as TNQ_DES, t.POENFQtd, 
                    t.POENFVlrTo, t.POENFSts
                FROM POENF t
                LEFT JOIN POFor f ON t.POForCod = f.POForCod
                LEFT JOIN POTnq nq ON t.POTnqCod = nq.POTnqCod AND t.POEmpCod = nq.POEmpCod
                WHERE t.POEmpCod = ? {where_extra}
                ORDER BY t.PODtaMov DESC
            """
            
            # --- DEBUG INFO ---
            self.ultima_query_bruta = sql
            for p in params:
                val = f"'{p}'" if isinstance(p, str) else str(p)
                self.ultima_query_bruta = self.ultima_query_bruta.replace("?", val, 1)

            cur.execute(sql, params)
            rows = cur.fetchall()

            # --- NOVO: Atualizar Combo Fornecedores apenas com os dados carregados (se estiver em TODOS) ---
            if forn_sel == "TODOS":
                fornecedores_grid = set()
                for r in rows:
                    if r[1] and r[2]: # POForCod and POForDes
                        fornecedores_grid.add(f"{int(r[1])} - {str(r[2]).strip().upper()}")
                opcoes = ["TODOS"] + sorted(list(fornecedores_grid))
                self.combo_fornecedor.configure(values=opcoes)

            for item in self.tree.get_children(): self.tree.delete(item)
            fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            total_qtd = 0
            total_vlr = 0
            count = len(rows)
            
            for r in rows:
                total_qtd += float(r[7] or 0)
                total_vlr += float(r[8] or 0)
                
                fmt_vals = []
                for i, v in enumerate(r):
                    if hasattr(v, 'strftime'): fmt_vals.append(v.strftime("%d/%m/%Y"))
                    elif i == 8: fmt_vals.append(fmt(float(v or 0)))
                    elif i == 7: fmt_vals.append(f"{int(float(v or 0))}") # Qtd sem decimais conforme pedido
                    else: fmt_vals.append(str(v or ""))
                self.tree.insert("", "end", values=tuple(fmt_vals))
            
            # Atualizar Totais no Rodapé (Alinhado com colunas)
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            
            row_totais = [""] * len(self.table_info['colunas'])
            row_totais[0] = f"REGISTROS: {count}"
            row_totais[7] = f"{int(total_qtd)}" # Qtd sem decimais
            row_totais[8] = fmt(total_vlr)
            self.tree_totais.insert("", "end", values=tuple(row_totais))

            self.re_zebrar()
            conn.close()
        except Exception as e:
            print(f"Erro ao carregar especializado POENF: {e}")


class LMCVendaBombaWindow(BaseWindow):
    def __init__(self, parent, config):
        self.table_info = POSTO_TABLES_CONFIG["POLMC1"]
        super().__init__(parent, "📊 LMC 1: Venda p/ Bomba (Especializado)", "LMC_VENDA_BOMBA")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent 
        self.is_posto_table = True

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", fill="x", expand=True, padx=20, pady=(0, 5))

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_frame, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(5, 2))
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=100, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        # Configuração da Grid
        cols = ("Data", "BomCod", "BomDes", "TnqCod", "TnqDes", "VelIni", "VelFin", "Dif", "Qtd")
        heads = ("📅 Data", "🔢 Cód. Bomba", "⛽ Bomba", "🛢️ Tanque", "📝 Tanque Desc.", "⏲️ Vel. Inicial", "⏲️ Vel. Final", "📈 Dif", "📊 Qtd (Lts)")
        wids = (110, 80, 180, 80, 150, 110, 110, 110, 110)
        alns = ("center", "center", "w", "center", "w", "e", "e", "e", "e")
        self.configurar_grid(cols, heads, wids, alns)
        
        self.after(200, self.carregar_dados)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            "--- ESTRUTURA SQL: LMC VENDA BOMBA ---\n"
            "TABELAS PRINCIPAIS: POLMC1 (Vendas), POBom (Bombas), POTnq (Tanques)\n"
            "RELACIONAMENTOS:\n"
            "- LMC1 JOIN Bombas (POBomCod) JOIN Tanques (POTnqCod)\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Volumem Vendido, Totalizadores Iniciais/Finais por Bomba.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtro de Bomba Opcional, Período de Data e Posto Logado.\n\n"
            "ORDENAÇÃO:\n"
            "- Data (Decrescente) e Tanque (Crescente)"
        )

    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import datetime
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_raw = self.ent_de.get().strip()
            ate_raw = self.ent_ate.get().strip()
            
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")

            sql = """
                SELECT 
                    l.PODtaMov, 
                    l.POBomCod, 
                    MAX(b.POBomDes) as BomDes,
                    MAX(b.POBomTnqCo) as TnqCod,
                    MAX(t.POTnqDes) as TnqDes,
                    SUM(l.POLMC1Qtd) as Qtd,
                    MAX(v1.POVel2Ini) as VelIni,
                    MAX(v2.POVel2Fin) as VelFin
                FROM POLMC1 l
                LEFT JOIN POBom b ON l.POBomCod = b.POBomCod AND l.POEmpCod = b.POEmpCod
                LEFT JOIN POTnq t ON b.POBomTnqCo = t.POTnqCod AND b.POEmpCod = t.POEmpCod
                LEFT JOIN POVel2 v1 ON l.POEmpCod = v1.POEmpCod 
                                   AND l.PODtaMov = v1.PODtaMov 
                                   AND l.POBomCod = v1.POBomCod
                                   AND v1.POVel1Per = 1
                LEFT JOIN POVel2 v2 ON l.POEmpCod = v2.POEmpCod 
                                   AND l.PODtaMov = v2.PODtaMov 
                                   AND l.POBomCod = v2.POBomCod
                                   AND v2.POVel1Per = (SELECT TOP 1 POEmpQtdPe FROM POEmp WHERE POEmpCod = l.POEmpCod)
                WHERE l.POEmpCod = ?
                  AND CAST(l.PODtaMov AS DATE) BETWEEN ? AND ?
                GROUP BY l.PODtaMov, l.POBomCod
                HAVING SUM(l.POLMC1Qtd) > 0
                ORDER BY l.PODtaMov DESC
            """
            
            self.ultima_query_bruta = sql.replace("?", f"'{posto_id}'", 1).replace("?", f"'{de}'", 1).replace("?", f"'{ate}'", 1)
            cur.execute(sql, (posto_id, de, ate))
            rows = cur.fetchall()

            for item in self.tree.get_children(): self.tree.delete(item)
            
            total_qtd = 0
            total_dif = 0
            for r in rows:
                v_qtd = float(r[5] or 0)
                v_ini = float(r[6] or 0)
                v_fin = float(r[7] or 0)
                v_dif = v_fin - v_ini
                
                total_qtd += v_qtd
                total_dif += v_dif
                
                fmt_vals = [
                    r[0].strftime("%d/%m/%Y") if hasattr(r[0], 'strftime') else str(r[0]),
                    str(r[1]),
                    str(r[2] or "N/A"),
                    str(r[3] or "---"),
                    str(r[4] or "---"),
                    f"{v_ini:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"{v_fin:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"{v_dif:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                    f"{v_qtd:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                ]
                self.tree.insert("", "end", values=tuple(fmt_vals))

            # Atualizar Rodapé
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            row_totais = [""] * 9
            row_totais[0] = f"REGISTROS: {len(rows)}"
            row_totais[7] = f"{total_dif:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_totais[8] = f"{total_qtd:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.tree_totais.insert("", "end", values=tuple(row_totais))

            self.re_zebrar()
            conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao carregar POLMC1:\n{e}")
            print(f"Erro ao carregar POLMC1 Especializado: {e}")



class LMCEstoqueWindow(BaseWindow):
    def __init__(self, parent, config):
        self.table_info = POSTO_TABLES_CONFIG["POLMC2"]
        super().__init__(parent, "📈 LMC 2: Estoque/Entrada (Especializado)", "LMC_ESTOQUE")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent 
        self.is_posto_table = True

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", fill="x", expand=True, padx=20, pady=(0, 5))

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 De:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=(0, 10))
        
        ctk.CTkLabel(self.filter_frame, text="Até:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(5, 2))
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=95, height=32, font=("Arial", 13)); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        ctk.CTkLabel(self.filter_frame, text="🛢️ Tanque:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 2))
        self.combo_tanque = ctk.CTkComboBox(self.filter_frame, values=["✨ Todos"], width=130, font=("Arial", 12))
        self.combo_tanque.pack(side="left", padx=2)
        self.combo_tanque.set("✨ Todos")

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=100, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        self.after(500, self.carregar_tanques)

        # Configuração da Grid
        cols = ("Data", "TnqCod", "TnqDes", "EstIni", "Compras", "Vendas", "Ajustes", "Transf", "EstFin")
        heads = ("📅 Data", "🛢️ Tanq", "📝 Descrição", "📦 Est. Inicial", "🔋 Compras", "⛽ Vendas", "⚙️ Ajustes", "🔄 Transf.", "🏁 Est. Final")
        wids = (100, 60, 180, 100, 100, 100, 100, 100, 100)
        alns = ("center", "center", "w", "e", "e", "e", "e", "e", "e")
        self.configurar_grid(cols, heads, wids, alns)
        
        self.after(200, self.carregar_dados)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            "--- ESTRUTURA SQL: LMC ESTOQUE / ENTRADA ---\n"
            "TABELAS PRINCIPAIS: POLMC2 (Estoques), POTnq (Tanques)\n"
            "RELACIONAMENTOS:\n"
            "- LMC2 LEFT JOIN Tanques. Subqueries em POENF (Compras) e POLMC1 (Vendas).\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Volumes Iniciais, Finais, Compras, Vendas, Ajustes por Tanque.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtro de Tanque, Período e ID do Posto Global.\n\n"
            "ORDENAÇÃO:\n"
            "- Data (Decrescente) e Tanque (Crescente)"
        )

    def carregar_tanques(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor()
                posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
                cur.execute("SELECT POTnqCod, POTnqDes FROM POTnq WHERE POEmpCod = ? ORDER BY POTnqCod", (posto_id,))
                rows = cur.fetchall()
                if not rows:
                    print(f"DEBUG: Nenhum tanque encontrado para posto {posto_id}")
                vals = ["✨ Todos"] + [f"{r[0]} - {r[1]}" for r in rows]
                self.combo_tanque.configure(values=vals)
                self.combo_tanque.set("✨ Todos")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showwarning("Aviso", f"Não foi possível carregar a lista de tanques:\n{e}")
            print(f"Erro tanques: {e}")

    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import datetime
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_raw = self.ent_de.get().strip()
            ate_raw = self.ent_ate.get().strip()
            
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")

            filtro_tanque = ""
            params = [posto_id, de, ate]
            sel_tanque = self.combo_tanque.get()
            if sel_tanque != "✨ Todos":
                tnq_id = sel_tanque.split(" - ")[0]
                filtro_tanque = " AND l.POTnqCod = ? "
                params.append(tnq_id)

            sql = f"""
                SELECT 
                    l.PODtaMov,
                    l.POTnqCod,
                    MAX(t.POTnqDes) as TnqDes,
                    MAX(l.POLMC2EIn) as EstIni,
                    COALESCE((SELECT SUM(e.POENFQtd) FROM POENF e WHERE e.POEmpCod = l.POEmpCod AND e.PODtaMov = l.PODtaMov AND e.POTnqCod = l.POTnqCod), 0) as Compras,
                    COALESCE((SELECT SUM(l1.POLMC1Qtd) FROM POLMC1 l1 JOIN POBom b ON l1.POBomCod = b.POBomCod AND l1.POEmpCod = b.POEmpCod 
                              WHERE l1.POEmpCod = l.POEmpCod AND l1.PODtaMov = l.PODtaMov AND b.POBomTnqCo = l.POTnqCod), 0) as Vendas,
                    MAX(l.POLMC2E_S) as Ajustes,
                    MAX(l.POLMC2TrsQ) as Transf,
                    (MAX(l.POLMC2EIn) + 
                     COALESCE((SELECT SUM(e.POENFQtd) FROM POENF e WHERE e.POEmpCod = l.POEmpCod AND e.PODtaMov = l.PODtaMov AND e.POTnqCod = l.POTnqCod), 0) - 
                     COALESCE((SELECT SUM(l1.POLMC1Qtd) FROM POLMC1 l1 JOIN POBom b ON l1.POBomCod = b.POBomCod AND l1.POEmpCod = b.POEmpCod 
                               WHERE l1.POEmpCod = l.POEmpCod AND l1.PODtaMov = l.PODtaMov AND b.POBomTnqCo = l.POTnqCod), 0) + 
                     MAX(l.POLMC2E_S) + MAX(l.POLMC2TrsQ)) as EstFin
                FROM POLMC2 l
                LEFT JOIN POTnq t ON l.POTnqCod = t.POTnqCod AND l.POEmpCod = t.POEmpCod
                WHERE l.POEmpCod = ?
                  AND CAST(l.PODtaMov AS DATE) BETWEEN ? AND ?
                  {filtro_tanque}
                GROUP BY l.POEmpCod, l.PODtaMov, l.POTnqCod
                ORDER BY l.PODtaMov DESC, l.POTnqCod ASC
            """
            
            self.ultima_query_bruta = sql.replace("?", f"'{posto_id}'", 1).replace("?", f"'{de}'", 1).replace("?", f"'{ate}'", 1)
            cur.execute(sql, params)
            rows = cur.fetchall()

            for item in self.tree.get_children(): self.tree.delete(item)
            
            totais = [0.0] * 6 # EstIni, Compras, Vendas, Ajustes, Transf, EstFin
            primeiro_lanca_tanque = {} # tnq_cod -> (data, val_ini)
            ultimo_lanca_tanque = {}   # tnq_cod -> (data, val_fin)
            
            for r in rows:
                dt_mov = r[0]
                tnq_cod = str(r[1])
                vals = [dt_mov.strftime("%d/%m/%Y") if hasattr(dt_mov, 'strftime') else str(dt_mov), tnq_cod, str(r[2] or "N/A")]
                
                v_ini = float(r[3] or 0)
                v_fin = float(r[8] or 0)

                # Estoque Inicial (Data mais antiga do período para este tanque)
                if tnq_cod not in primeiro_lanca_tanque or dt_mov < primeiro_lanca_tanque[tnq_cod][0]:
                    primeiro_lanca_tanque[tnq_cod] = (dt_mov, v_ini)
                
                # Estoque Final (Data mais recente do período para este tanque)
                if tnq_cod not in ultimo_lanca_tanque or dt_mov > ultimo_lanca_tanque[tnq_cod][0]:
                    ultimo_lanca_tanque[tnq_cod] = (dt_mov, v_fin)

                for i in range(3, 9):
                    val = float(r[i] or 0)
                    if i in (4, 5, 6, 7): # Apenas Compras, Vendas, Ajustes e Transf são somados acumulados
                        totais[i-3] += val
                    vals.append(f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                
                self.tree.insert("", "end", values=tuple(vals))

            # Rodapé (Lógica especializada para Estoques Inicial e Final)
            total_real_est_ini = sum(item[1] for item in primeiro_lanca_tanque.values())
            total_real_est_fin = sum(item[1] for item in ultimo_lanca_tanque.values())
            
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            row_tot = [""] * 9
            row_tot[0] = f"SOMA:"
            # i=3: Est. Inicial
            row_tot[3] = f"{total_real_est_ini:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            # i=4..7: Compras, Vendas, Ajustes, Transf
            for i in range(4, 8):
                row_tot[i] = f"{totais[i-3]:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            # i=8: Est. Final (Último lançamento somado)
            row_tot[8] = f"{total_real_est_fin:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            self.tree_totais.insert("", "end", values=tuple(row_tot))

            self.re_zebrar()
            conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao carregar POLMC2:\n{e}")
            print(f"Erro ao carregar POLMC2 Especializado: {e}")


class LMCResumoWindow(BaseWindow):
    def __init__(self, parent, config):
        self.table_info = POSTO_TABLES_CONFIG["POLMC3"]
        super().__init__(parent, "📊 LMC 3: Resumo Período (Especializado)", "LMC_RESUMO")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent 
        self.is_posto_table = True

        # --- BARRA DE FILTROS (Alinhada à Esquerda) ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 Período:", font=("Arial", 12, "bold")).pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=90, height=32, font=("Arial", 12))
        self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=2)

        ctk.CTkLabel(self.filter_frame, text="até", font=("Arial", 12)).pack(side="left", padx=5)
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=90, height=32, font=("Arial", 12))
        self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        ctk.CTkLabel(self.filter_frame, text="⛽ Prod:", font=("Arial", 12, "bold")).pack(side="left", padx=(10, 2))
        self.combo_produto = ctk.CTkComboBox(self.filter_frame, values=["✨ Todos"], width=160, font=("Arial", 12))
        self.combo_produto.pack(side="left", padx=2)
        self.combo_produto.set("✨ Todos")

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=100, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        # Configuração da Grid (POLMC3)
        cols = ("Data", "CombCod", "CombDes", "EstIni", "Entradas", "Vendas", "Ajustes", "EstFin")
        heads = ("📅 Data", "🔢 Cód.", "⛽ Combustível", "📦 Est. Inicial", "📥 Entradas", "📤 Vendas", "⚙️ Ajustes", "🏁 Est. Final")
        wids = (110, 60, 200, 110, 110, 110, 110, 110)
        alns = ("center", "center", "w", "e", "e", "e", "e", "e")
        self.configurar_grid(cols, heads, wids, alns)
        
        self.after(500, self.carregar_produtos)
        self.after(800, self.carregar_dados)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            "--- ESTRUTURA SQL: LMC RESUMO PERÍODO ---\n"
            "TABELAS PRINCIPAIS: POLMC3 (Resumo Combustíveis)\n"
            "RELACIONAMENTOS:\n"
            "- LEFT JOIN POTCo c ON l.POLMC3TCoC = c.POTCoCod\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Entradas, Vendas, Ajustes e Saldo de Estoque Dinâmico.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtro de Combustível (Tela) e Posto logado.\n\n"
            "ORDENAÇÃO:\n"
            "- Data (Decrescente) e Combustível (Crescente)"
        )

    def carregar_produtos(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor()
                cur.execute("SELECT POTCoCod, POTCoDes FROM POTCo ORDER BY POTCoCod")
                rows = cur.fetchall()
                vals = ["✨ Todos"] + [f"{int(r[0])} - {str(r[1]).strip().upper()}" for r in rows]
                self.combo_produto.configure(values=vals)
        except Exception as e:
            print(f"Erro ao carregar combustíveis no resumo: {e}")

    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import datetime
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_raw = self.ent_de.get().strip(); ate_raw = self.ent_ate.get().strip()
            de = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            ate = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")

            filtro_prod = ""
            params = [posto_id, de, ate]
            sel_prod = self.combo_produto.get()
            if sel_prod != "✨ Todos":
                prod_cod = sel_prod.split(" - ")[0]
                filtro_prod = " AND l.POLMC3TCoC = ? "
                params.append(prod_cod)

            sql = f"""
                SELECT 
                    l.PODtaMov,
                    l.POLMC3TCoC,
                    MAX(c.POTCoDes) as CombDes,
                    SUM(l.POLMC3Ein) as EstIni,
                    SUM(l.POLMC3EntN) as Entradas,
                    SUM(l.POLMC3Vda) as Vendas,
                    SUM(l.POLMC3E_S) as Ajustes,
                    (SUM(l.POLMC3Ein) + SUM(l.POLMC3EntN) - SUM(l.POLMC3Vda) + SUM(l.POLMC3E_S)) as EstFin
                FROM POLMC3 l
                LEFT JOIN POTCo c ON l.POLMC3TCoC = c.POTCoCod
                WHERE l.POEmpCod = ?
                  AND CAST(l.PODtaMov AS DATE) BETWEEN ? AND ?
                  {filtro_prod}
                GROUP BY l.PODtaMov, l.POLMC3TCoC
                ORDER BY l.PODtaMov DESC, l.POLMC3TCoC ASC
            """
            self.ultima_query_bruta = sql.replace("?", f"'{posto_id}'", 1).replace("?", f"'{de}'", 1).replace("?", f"'{ate}'", 1)
            if filtro_prod: self.ultima_query_bruta = self.ultima_query_bruta.replace("?", f"'{prod_cod}'", 1)
            
            cur.execute(sql, params)
            rows = cur.fetchall()

            for item in self.tree.get_children(): self.tree.delete(item)
            
            totais_mov = [0.0] * 3 # Ent, Vnd, Aju
            pri_est = {} # cat -> (data, val)
            ult_est = {} # cat -> (data, val)
            
            for r in rows:
                dt = r[0]; cat = str(r[1])
                v_ini = float(r[3] or 0); v_ent = float(r[4] or 0)
                v_vnd = float(r[5] or 0); v_aju = float(r[6] or 0)
                v_fin = float(r[7] or 0)

                if cat not in pri_est or dt < pri_est[cat][0]: pri_est[cat] = (dt, v_ini)
                if cat not in ult_est or dt > ult_est[cat][0]: ult_est[cat] = (dt, v_fin)
                
                totais_mov[0] += v_ent; totais_mov[1] += v_vnd; totais_mov[2] += v_aju

                vals = [dt.strftime("%d/%m/%Y") if hasattr(dt, 'strftime') else str(dt), cat, str(r[2] or "N/A")]
                for val in [v_ini, v_ent, v_vnd, v_aju, v_fin]:
                    vals.append(f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                
                self.tree.insert("", "end", values=tuple(vals))

            # Rodapé Inteligente
            t_ini = sum(v[1] for v in pri_est.values())
            t_fin = sum(v[1] for v in ult_est.values())
            
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            row_tot = [""] * 8
            row_tot[0] = "SOMA:"
            row_tot[3] = f"{t_ini:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_tot[4] = f"{totais_mov[0]:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_tot[5] = f"{totais_mov[1]:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_tot[6] = f"{totais_mov[2]:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_tot[7] = f"{t_fin:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            self.tree_totais.insert("", "end", values=tuple(row_tot))
            self.re_zebrar(); conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao carregar POLMC3:\n{e}")
            print(f"Erro ao carregar POLMC3 Especializado: {e}")


class MovimentoPostoWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "⭐ Movimentos do Posto (Auditoria Completa)", "MOV_POSTO")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent 
        self.is_posto_table = True

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 Período:", font=("Arial", 12, "bold")).pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=90, height=32); self.ent_de.pack(side="left", padx=2)
        self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=2)

        ctk.CTkLabel(self.filter_frame, text="até", font=("Arial", 12)).pack(side="left", padx=5)
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=90, height=32); self.ent_ate.pack(side="left", padx=2)
        self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=100, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        ctk.CTkLabel(self.filter_frame, text="👤 Frentista:", font=("Arial", 12, "bold")).pack(side="left", padx=(10, 2))
        self.combo_frentista = ctk.CTkComboBox(self.filter_frame, values=["Todos"], width=130, height=32)
        self.combo_frentista.pack(side="left", padx=2)
        self.combo_frentista.set("Todos")

        # --- GRID CONFIG (45 COLUNAS - Auditoria Total) ---
        cols = (
            "Tst", "Regis", "VlrCx", "Usu", "VlrMo", "VlrDs", "VlrAc", "VlrSa", "TipMo", "DesTi", "Sts", 
            "TstFe", "Obs", "Bic", "BomCo", "ProCo", "ProDe", "Qtd", "VelIn", "VelFi", "VlrUn", 
            "VlrBo", "TstCo", "Can", "DataA", "HoraA", "PUBom", "VlrLu", "DCx", "VUF", "CusUn", "CusTo", "VTF", "VDFFa",
            "Mrg", "VlrCC", "TipCC", "CarCn", "PECDs", "VDF", "NroNf", "LibGe", "Canal", "TipOr"
        )
        heads = (
            "🕒 Lançamento", "🔢 Regis.", "💰 Vlr Cx", "👤 Frentista", "📈 Movim.", "📉 Desc.", "➕ Acrésc.", "⚖️ Saldo", "📂 Tipo", "📝 Desc. Tipo", "🔘 Sts", 
            "🕒 Fech.", "📝 Obs.", "📦 Bico", "⛽ Bomba", "🏷️ Cód. Pro.", "📦 Produto", "📊 Qtd.", "🏁 Vel. Ini.", "🏁 Vel. Fim.", "💰 Unit.", 
            "⛽ Vlr Bomba", "🕒 Confir.", "🚫 Can.", "📅 Data Ab.", "🕒 Hora Ab.", "⛽ P. Bomba", "📈 Lucro", "🕒 Data Cx", "💎 Unit. Fis.", "💰 Custo Un.", "💰 Custo Tot.", "💎 Tot. Fis.", "📉 Desc. Fis. Fake",
            "📊 Margem", "💳 Vlr Adic. CC", "💳 Tipo CC", "🛒 Car. Conv.", "📉 PEC Desc.", "📉 Desc. Fis.", "📄 NF", "🔑 Lib. Ger.", "📡 Canal", "📂 Tip. Orig."
        )
        wids = (140, 80, 100, 100, 100, 90, 90, 100, 60, 120, 50, 140, 150, 60, 60, 90, 180, 80, 100, 100, 90, 100, 140, 50, 90, 80, 90, 90, 140, 90, 90, 100, 100, 110, 80, 100, 80, 80, 90, 90, 80, 70, 60, 60)
        alns = ("center", "center", "e", "w", "e", "e", "e", "e", "center", "w", "center", "center", "w", "center", "center", "center", "w", "e", "e", "e", "e", "e", "center", "center", "center", "center", "e", "e", "center", "e", "e", "e", "e", "e", "e", "e", "center", "center", "e", "e", "center", "center", "center", "center")
        
        self.configurar_grid(cols, heads, wids, alns)

        # Adicionar Scrollbar Horizontal (Necessário para 33 colunas)
        self.h_scroll = ttk.Scrollbar(self.grid_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(xscrollcommand=self.h_scroll.set)
        self.h_scroll.pack(side="bottom", fill="x")

        self.after(200, self.carregar_dados)

    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import datetime
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_dt = datetime.strptime(self.ent_de.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
            ate_dt = datetime.strptime(self.ent_ate.get(), "%d/%m/%Y").strftime("%Y-%m-%d")

            frentista_sel = self.combo_frentista.get()
            where_ext = " AND POCF1Usu = ? " if frentista_sel != "Todos" else ""
            params = [posto_id, de_dt, ate_dt]
            if frentista_sel != "Todos": params.append(frentista_sel)

            sql = f"""
                SELECT 
                    POCF1Tst, POCF1Regis, POCF1VlrCx, POCF1Usu, POCF1VlrMo, POCF1VlrDs, POCF1VlrAc, POCF1VlrSa, 
                    POCF1TipMo, 
                    CASE POCF1TipMo 
                        WHEN 'V' THEN 'VENDA' WHEN 'R' THEN 'RECEBIMENTO' WHEN 'S' THEN 'SAIDA' WHEN 'E' THEN 'ENTRADA' 
                        ELSE POCF1TipMo END as DesTi,
                    POCF1Sts, POCF1TstFe, POCF1Obs, POCF1Bic, POCF1BomCo, POCF1ProCo, POCF1ProDe, 
                    POCF1Qtd, POCF1VelIn, POCF1VelFi, POCF1VlrUn, POCF1VlrBo, POCF1TstCo, POCF1Can, POCF1DataA, 
                    POCF1HoraA, POCF1PUBom, POCF1VlrLu, POCF1DCx, POCF1VUF, 
                    POCF1CusUn,
                    (POCF1Qtd * ISNULL(POCF1CusUn, 0)) as CusTo, 
                    (POCF1VUF * POCF1Qtd) as VTF,
                    CASE 
                        WHEN POCF1VlrCx > (POCF1VUF * POCF1Qtd) AND (POCF1VUF * POCF1Qtd) <> 0 
                        THEN POCF1VlrCx - (POCF1VUF * POCF1Qtd) 
                        ELSE 0 END as VDFFa,
                    POCF1Mrg, POCF1VlrCC, POCF1TipCC, POCF1CarCn, POCF1PECDs, POCF1VDF, 
                    POCF1NroNf, POCF1LibGe, POCF1Canal, POCF1TipOr
                FROM POCF1
                WHERE POEmpCod = ? AND CAST(POCF1Tst AS DATE) BETWEEN ? AND ?
                {where_ext}
                ORDER BY POCF1Tst DESC
            """
            self.ultima_query_bruta = sql
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()

            # Atualizar Combobox
            if frentista_sel == "Todos":
                frentistas = sorted(list(set(str(r[3]).strip() for r in rows if r[3])))
                self.combo_frentista.configure(values=["Todos"] + frentistas)

            for item in self.tree.get_children(): self.tree.delete(item)
            fmt_moeda = lambda v: f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            fmt_4 = lambda v: f"{v:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            for r in rows:
                v = list(r)
                # Datas: 0(Tst), 11(TstFe), 22(TstCo), 24(DataA), 28(DCx)
                for idx in [0, 11, 22, 28]:
                    if v[idx] and hasattr(v[idx], "strftime"): v[idx] = v[idx].strftime("%d/%m/%Y %H:%M")
                    else: v[idx] = "" if not v[idx] else v[idx]
                if v[24] and hasattr(v[24], "strftime"): v[24] = v[24].strftime("%d/%m/%Y")

                # Numéricos: VlrCx(2), VlrMo(4), VlrDs(5), VlrAc(6), VlrSa(7), Qtd(17), Vel(18,19), Un(20), VlrBo(21), PUB(26), Luc(27), VUF(29), CusUn(30), CusTo(31), VTF(32), VDFFa(33), Mrg(34), VlrCC(35), PECDs(38), VDF(39)
                for idx in [2, 4, 5, 6, 7, 17, 18, 19, 20, 21, 26, 27, 29, 30, 31, 32, 33, 34, 35, 38, 39]:
                    if v[idx] is not None: 
                        if idx in [29, 30, 38]: v[idx] = fmt_4(float(v[idx]))
                        else: v[idx] = fmt_moeda(float(v[idx]))
                
                self.tree.insert("", "end", values=tuple(v))

            # Totais no rodapé (Quantidade, Valor Movimento e Valor Caixa)
            total_qtd = sum(float(r[17] or 0) for r in rows)
            total_mov = sum(float(r[4] or 0) for r in rows)
            total_cx = sum(float(r[2] or 0) for r in rows)
            
            info_tot = f"Total: {len(rows)} | 📊 Qtd: {total_qtd:,.3f} | 📈 Movim: {fmt_moeda(total_mov)} | 💰 Caixa: {fmt_moeda(total_cx)}"
            self.lbl_total_regs.configure(text=info_tot)
            self.re_zebrar(); conn.close()
        except Exception as e:
            print(f"Erro POCF1: {e}")
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", str(e))

class CaixaFrentistaResumoWindow(BaseWindow):
    def __init__(self, parent, config, initial_filters=None):
        self.table_info = POSTO_TABLES_CONFIG["POCF2"]
        super().__init__(parent, "👥 Caixa Frentista: Resumo (Especializado)", "CX_FRENTISTA_RES")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent 
        self.is_posto_table = True
        self.initial_filters = initial_filters

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")

        from datetime import date
        hoje = date.today()
        ini_mes = date(hoje.year, hoje.month, 1).strftime("%d/%m/%Y")
        fim_mes = hoje.strftime("%d/%m/%Y")

        ctk.CTkLabel(self.filter_frame, text="📅 Período:", font=("Arial", 12, "bold")).pack(side="left", padx=2)
        self.ent_de = ctk.CTkEntry(self.filter_frame, width=90, height=32, font=("Arial", 12))
        self.ent_de.pack(side="left", padx=2)
        if self.initial_filters:
            self.ent_de.insert(0, self.initial_filters.get('de', ini_mes))
        else:
            self.ent_de.insert(0, ini_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_de)).pack(side="left", padx=2)

        ctk.CTkLabel(self.filter_frame, text="até", font=("Arial", 12)).pack(side="left", padx=5)
        self.ent_ate = ctk.CTkEntry(self.filter_frame, width=90, height=32, font=("Arial", 12))
        self.ent_ate.pack(side="left", padx=2)
        if self.initial_filters:
            self.ent_ate.insert(0, self.initial_filters.get('ate', fim_mes))
        else:
            self.ent_ate.insert(0, fim_mes)
        ctk.CTkButton(self.filter_frame, text="📅", width=25, height=32, fg_color="transparent", text_color="black", hover_color="#E0E0E0", 
                     command=lambda: self.abrir_calendario(self.ent_ate)).pack(side="left", padx=(0, 10))

        self.btn_filtrar = ctk.CTkButton(self.filter_frame, text="⚙️ Filtrar", width=100, height=32, font=("Arial", 12, "bold"), 
                                        fg_color="#1E88E5", hover_color="#1565C0", command=self.carregar_dados)
        self.btn_filtrar.pack(side="left", padx=10)

        ctk.CTkLabel(self.filter_frame, text="👤 Frentista:", font=("Arial", 12, "bold")).pack(side="left", padx=(10, 2))
        self.combo_frentista = ctk.CTkComboBox(self.filter_frame, values=["Todos"], width=130, height=32, font=("Arial", 12), command=lambda _: self.carregar_dados())
        self.combo_frentista.pack(side="left", padx=2)
        
        if self.initial_filters and self.initial_filters.get('frentista'):
            self.combo_frentista.set(self.initial_filters['frentista'])
        else:
            self.combo_frentista.set("Todos")

        # Configuração da Grid (POCF2 - Simplificada p/ auditoria)
        cols = ("Usu", "Abertura", "VlrFi", "TSTFe", "UsuFe", "VlrBo", "CofVl", "VlrDe", "VlrRe", "RecCa", "LucCo", "DCx")
        heads = ("👤 Frentista", "🕒 Abertura", "🏁 Valor Final", "🕒 Fech.", "👤 Conf.", "⛽ Bombas", "📥 Cofre", "📉 Desc.", "💰 Rec.", "💳 Cartão", "📈 Lucro", "📅 Data Cx")
        wids = (120, 150, 110, 150, 100, 110, 110, 90, 110, 110, 90, 100)
        alns = ("w", "center", "e", "center", "w", "e", "e", "e", "e", "e", "e", "center")
        self.configurar_grid(cols, heads, wids, alns)
        
        self.after(200, self.carregar_dados)
        self.tree.bind("<Button-1>", self.on_double_click)
        self.tree.bind("<Double-1>", self.on_double_click)
        self.tree.bind("<Return>", self.on_double_click)

    def on_double_click(self, event):
        """Captura o clique (simples ou duplo) e abre a tela de detalhes (POCF1)."""
        # Padrão Ouro: Identificar exatamente se clicou em uma linha
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            # Fallback para seleção atual se for via teclado (Enter) ou clique manual
            selection = self.tree.selection()
            if not selection: return
            item_id = selection[0]
        
        if not item_id or "|" not in str(item_id): return

        try:
            usu_raw, tst_raw = str(item_id).split("|")
            
            # Salvar filtros atuais para o retorno
            filtros_salvos = {
                'de': self.ent_de.get(),
                'ate': self.ent_ate.get(),
                'frentista': self.combo_frentista.get()
            }
            
            # Navegação premium: Substituímos o frame atual
            self.pack_forget()
            detalhe = CaixaFrentistaItensWindow(self.hub.main_frame, self.config_db, usu_raw, tst_raw, filtros_salvos)
            detalhe.hub = self.hub
            detalhe.pack(fill="both", expand=True)
            self.hub.modulo_atual = detalhe
            self.destroy()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro de Navegação", f"Não foi possível abrir o detalhamento:\n{e}")
            print(f"Erro ao abrir detalhes: {e}")

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: CAIXAFRENTISTARESUMO ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Gera o comando SQL técnico com base nos filtros atuais da UI."""
        try:
            from datetime import datetime
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_raw = self.ent_de.get().strip(); ate_raw = self.ent_ate.get().strip()
            
            # Tenta converter datas para o formato SQL, mas não trava se falhar (ex: string vazia)
            try:
                de_dt = datetime.strptime(de_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
                ate_dt = datetime.strptime(ate_raw, "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                de_dt = "YYYY-MM-DD"; ate_dt = "YYYY-MM-DD"

            frentista_sel = self.combo_frentista.get()
            where_frentista = f" AND p.POCF2Usu = '{frentista_sel}' " if frentista_sel != "Todos" else ""

            sql = f"""SELECT 
    p.POCF2Usu, p.POCF2Tst, p.POCF2Sts, p.POCF2VlrMa, 
    (SELECT SUM(f.POCF3VlrFi) FROM POCF3 f 
     WHERE f.POEmpCod = p.POEmpCod AND f.POCF2Usu = p.POCF2Usu AND f.POCF2Tst = p.POCF2Tst) as VlrFi,
    (p.POCF2VlrMa - ISNULL((SELECT SUM(f.POCF3VlrFi) FROM POCF3 f 
                            WHERE f.POEmpCod = p.POEmpCod AND f.POCF2Usu = p.POCF2Usu AND f.POCF2Tst = p.POCF2Tst), 0)) as VlrDi,
    p.POCF2TSTFe, p.POCF2UsuFe, p.POCF2VlrBo, p.POCF2CofVl, p.POCF2VlrDe, p.POCF2VlrRe, p.POCF2RecCa, p.POCF2LucCo,
    p.POCF2VlrCh, p.POCF2VlrMo, p.POCF2DCx
FROM POCF2 p
WHERE p.POEmpCod = {posto_id}
  AND CAST(p.POCF2Tst AS DATE) BETWEEN '{de_dt}' AND '{ate_dt}'
  {where_frentista}
ORDER BY p.POCF2Tst DESC"""
            return sql
        except Exception as e:
            return f"-- Erro ao gerar SQL dinâmico: {e}"

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            frentista_sel = self.combo_frentista.get()
            sql = self.get_current_query()
            # Ajustar sql para usar placeholders (parametragem real para execução)
            sql_exec = sql.replace(f"'{frentista_sel}'", "?") if frentista_sel != "Todos" else sql
            sql_exec = sql_exec.replace("'YYYY-MM-DD'", "?") # Fallback safety
            
            # Reprocessar parâmetros reais para o cursor
            from datetime import datetime
            posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
            de_dt = datetime.strptime(self.ent_de.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
            ate_dt = datetime.strptime(self.ent_ate.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
            
            params = [posto_id, de_dt, ate_dt]
            if frentista_sel != "Todos": params.append(frentista_sel)

            # Re-montar a query limpa com placeholders para o pyodbc
            where_frentista = " AND p.POCF2Usu = ? " if frentista_sel != "Todos" else ""
            sql_final = f"""
                SELECT 
                    p.POCF2Usu, p.POCF2Tst, p.POCF2Sts, p.POCF2VlrMa, 
                    (SELECT SUM(f.POCF3VlrFi) FROM POCF3 f 
                     WHERE f.POEmpCod = p.POEmpCod AND f.POCF2Usu = p.POCF2Usu AND f.POCF2Tst = p.POCF2Tst) as VlrFi,
                    (p.POCF2VlrMa - ISNULL((SELECT SUM(f.POCF3VlrFi) FROM POCF3 f 
                                            WHERE f.POEmpCod = p.POEmpCod AND f.POCF2Usu = p.POCF2Usu AND f.POCF2Tst = p.POCF2Tst), 0)) as VlrDi,
                    p.POCF2TSTFe, p.POCF2UsuFe, p.POCF2VlrBo, p.POCF2CofVl, p.POCF2VlrDe, p.POCF2VlrRe, p.POCF2RecCa, p.POCF2LucCo,
                    p.POCF2VlrCh, p.POCF2VlrMo, p.POCF2DCx
                FROM POCF2 p
                WHERE p.POEmpCod = ?
                  AND CAST(p.POCF2Tst AS DATE) BETWEEN ? AND ?
                  {where_frentista}
                ORDER BY p.POCF2Tst DESC
            """
            
            self.ultima_query_bruta = sql # Para Auditoria fiel ao que o usuário vê (com valores)
            cur.execute(sql_final, tuple(params))
            rows = cur.fetchall()

            # Atualizar Combobox de Frentistas se estiver em 'Todos'
            if frentista_sel == "Todos":
                frentistas_encontrados = sorted(list(set(str(r[0]).strip() for r in rows)))
                self.combo_frentista.configure(values=["Todos"] + frentistas_encontrados)
            
            # Se voltamos de uma tela com filtro, tentar selecionar o frentista original
            if self.initial_filters and self.initial_filters.get('frentista'):
                self.combo_frentista.set(self.initial_filters['frentista'])
                # Limpamos para não forçar em recargas manuais
                self.initial_filters['frentista'] = None

            for item in self.tree.get_children(): self.tree.delete(item)
            total_calculado = 0.0
            total_desconto = 0.0
            
            for r in rows:
                usu = str(r[0]).strip()
                tst_ab = r[1]
                v_fi = float(r[4] or 0)
                tst_fe = r[6]; usu_fe = str(r[7] or "").strip()
                v_bo = float(r[8] or 0); v_cof = float(r[9] or 0); v_dec = float(r[10] or 0)
                v_rec = float(r[11] or 0); v_crt = float(r[12] or 0); v_luc = float(r[13] or 0)
                dcx = r[16]
                
                total_calculado += v_fi
                total_desconto += v_dec
                
                ab_fmt = tst_ab.strftime("%d/%m/%Y %H:%M") if hasattr(tst_ab, "strftime") and tst_ab.year < 9999 else "N/A"
                fe_fmt = tst_fe.strftime("%d/%m/%Y %H:%M") if hasattr(tst_fe, "strftime") and tst_fe.year < 9999 else "ABER"
                dcx_fmt = dcx.strftime("%d/%m/%Y") if hasattr(dcx, "strftime") and dcx.year < 9999 else "-"
                
                # Formatar Numericos (Apenas colunas visiveis)
                vals = [usu, ab_fmt]
                vals.append(f"{v_fi:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                vals.append(fe_fmt)
                vals.append(usu_fe)
                for v in [v_bo, v_cof, v_dec, v_rec, v_crt, v_luc]:
                    vals.append(f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                vals.append(dcx_fmt)

                # Usamos IID para guardar as chaves de busca (Frentista | Timestamp)
                iid_key = f"{usu}|{tst_ab.strftime('%Y-%m-%d %H:%M:%S.%f') if hasattr(tst_ab, 'strftime') else str(tst_ab)}"
                self.tree.insert("", "end", iid=iid_key, values=tuple(vals))

            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            row_tot = [""] * 12
            row_tot[0] = f"REGISTROS: {len(rows)}"
            row_tot[2] = f"{total_calculado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            row_tot[7] = f"{total_desconto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.tree_totais.insert("", "end", values=tuple(row_tot))
            
            self.re_zebrar(); conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao carregar POCF2:\n{e}")
            print(f"Erro ao carregar POCF2 Especializado: {e}")

    def mostrar_sql_bruto(self):
        """Sobrescreve a auditoria global para aplicar o padrão VIP 2025 nesta tela específica."""
        resumo = self.get_sql_summary()
        query = self.get_current_query()
        
        win = ctk.CTkToplevel(self)
        win.title(f"AUDITORIA SQL ({self.id_str})")
        win.geometry("800x600")
        win.configure(fg_color="#F1F5F9")
        win.transient(self); win.grab_set()
        
        # Centralizar 800x600
        sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
        win.geometry(f"800x600+{(sw-800)//2}+{(sh-600)//2}")
        
        ctk.CTkLabel(win, text="🔍 COMANDO SQL PRONTO PARA EXECUÇÃO", font=("Arial", 18, "bold"), text_color="#1565C0").pack(pady=(20, 10))
        
        # Container Branco Arredondado
        container = ctk.CTkFrame(win, fg_color="white", corner_radius=12)
        container.pack(fill="both", expand=True, padx=40, pady=(0, 20))
        
        text_area = ctk.CTkTextbox(container, font=("Consolas", 12), fg_color="white", text_color="#1E293B", border_width=0)
        text_area.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Inserir Conteúdo (Resumo + Query)
        conteudo_completo = f"{resumo}\n\n{'='*50}\n\n{query}"
        text_area.insert("0.0", conteudo_completo)
        text_area.configure(state="disabled")
        
        # Botões de Rodapé VIP
        footer = ctk.CTkFrame(win, fg_color="transparent")
        footer.pack(fill="x", side="bottom", pady=20)
        
        def copiar():
            self.clipboard_clear()
            self.clipboard_append(query)
            self.update()
            from tkinter import messagebox
            messagebox.showinfo("Copiado", "COMANDO SQL TÉCNICO COPIADO!")

        ctk.CTkButton(footer, text="📋 COPIAR SQL", font=("Arial", 12, "bold"), fg_color="#2E7D32", hover_color="#1B5E20", 
                     corner_radius=8, height=35, command=copiar).pack(side="left", padx=(40, 10))
        
        ctk.CTkButton(footer, text="FECHAR", font=("Arial", 12, "bold"), fg_color="#475569", hover_color="#334155", 
                     corner_radius=8, height=35, command=win.destroy).pack(side="left", padx=10)


class CaixaFrentistaItensWindow(BaseWindow):
    """Tela de Detalhamento de Movimentos (POCF1) vinculada ao fechamento do Resumo (POCF2)."""
    def __init__(self, parent, config, usu, tst_fe, parent_filters=None):
        self.usu_filter = usu
        self.tst_filter = tst_fe
        self.parent_filters = parent_filters
        title = f"📋 Itens do Caixa - Frentista: {usu}"
        super().__init__(parent, title, "CX_FRENTISTA_ITENS")
        self.config_db = config
        self.hub = parent.hub if hasattr(parent, 'hub') else parent
        self.is_posto_table = True

        # Botão Voltar Customizado (Prioridade Alta)
        self.btn_voltar = ctk.CTkButton(self.top_frame, text="⬅  VOLTAR AO RESUMO", width=180, height=32, 
                                       fg_color="#475569", hover_color="#334155", font=("Arial", 12, "bold"),
                                       command=self.voltar)
        self.btn_voltar.pack(side="right", padx=20, pady=5)

        # --- FILTRO POR TIPO DE MOVIMENTO ---
        self.tipmo_legend = {
            "Todos": "Todos",
            "V": "V - VENDA A VISTA",
            "P": "P - VENDA PRAZO",
            "C": "C - VENDA CARTÃO",
            "S": "S - VALE CLIENTE",
            "E": "E - VALE POSTO",
            "R": "R - RECEBIMENTOS",
            "$": "$ - VENDA PRAZO (M)",
            "A": "A - BAIXA DE VALE",
            "O": "O - OUTROS",
            "B": "B - BOMBA AGUARD.",
            "M": "M - ENTRADA/SAIDA",
            "1": "1 - SAIU $ ENTROU CARTÃO",
            "2": "2 - SAIU CARTÃO ENTROU $",
            "D": "D - DESPESAS",
            "3": "3 - CARTÃO DÉBITO",
            "4": "4 - CARTÃO CRÉDITO",
            "5": "5 - PIX",
            "6": "6 - VENDA CARTÃO DÉBITO",
            "7": "7 - CARTÃO ALELO",
            "8": "8 - TRANSF. $ LOJA"
        }
        
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=5)
        
        ctk.CTkLabel(self.filter_frame, text="🔍 Filtrar Tipo:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_tipmo = ctk.CTkComboBox(self.filter_frame, values=list(self.tipmo_legend.values()), 
                                         width=220, height=32, command=lambda _: self.carregar_dados())
        self.combo_tipmo.pack(side="left", padx=5)
        self.combo_tipmo.set("Todos")

        # Checkbox: Somente com Desconto
        self.chk_desc_var = ctk.BooleanVar(value=False)
        self.chk_desc = ctk.CTkCheckBox(self.filter_frame, text="Só c/ Desconto", variable=self.chk_desc_var,
                                        font=("Arial", 12, "bold"), command=self.carregar_dados,
                                        checkbox_width=20, checkbox_height=20)
        self.chk_desc.pack(side="left", padx=(15, 5))

        # Configuração da Grid (Colunas Essenciais da POCF1)
        cols = ("Tst", "Regis", "VlrCx", "VlrDs", "TipMo", "DesTi", "Obs", "ProDe", "Qtd", "VlrUn", "DCx")
        heads = ("🕒 Lançamento", "Regis.", "💰 Vlr Cx", "📉 Desconto", "Tipo", "Descrição", "Obs.", "📦 Produto", "📊 Qtd", "💰 Unit.", "📅 Data Cx")
        wids = (140, 80, 100, 100, 60, 120, 150, 200, 80, 90, 100)
        alns = ("center", "center", "e", "e", "center", "w", "w", "w", "e", "e", "center")
        
        self.configurar_grid(cols, heads, wids, alns)
        
        # Scrollbar Horizontal se necessário
        self.h_scroll = ttk.Scrollbar(self.grid_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(xscrollcommand=self.h_scroll.set)
        self.h_scroll.pack(side="bottom", fill="x")

        # --- RODAPÉ DE TOTAIS (tree_totais espelhada - Padrão Ouro) ---
        totais_frame = ctk.CTkFrame(self, height=35, fg_color="transparent", corner_radius=0)
        totais_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0,0))
        
        self.tree_totais = ttk.Treeview(totais_frame, columns=cols, show="", height=1, style="TotaisItens.Treeview")
        self.tree_totais.column("#0", width=0, stretch=False)
        for c, w, a in zip(cols, wids, alns):
            self.tree_totais.column(c, width=w, minwidth=w, anchor=a)
        
        style_t = ttk.Style()
        style_t.configure("TotaisItens.Treeview", background="#1565C0", foreground="white", font=("Arial", 12, "bold"), rowheight=28)
        self.tree_totais.pack(fill="x", expand=True)



        self.after(200, self.carregar_dados)

    def fechar_tela(self):
        """Sobrescreve o fechar da BaseWindow para retornar ao resumo, não ao Posto."""
        self.voltar()

    def voltar(self):
        """Retorna para a tela de resumo preservando os filtros originais."""
        self.pack_forget()
        resumo = CaixaFrentistaResumoWindow(self.hub.main_frame, self.config_db, self.parent_filters)
        resumo.hub = self.hub
        resumo.pack(fill="both", expand=True)
        self.hub.modulo_atual = resumo
        self.destroy()

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: CAIXAFRENTISTAITENS ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        posto_id = self.hub.get_selected_posto_id() if hasattr(self, 'hub') else 1
        tst_seg = self.tst_filter[:19]
        
        # Obter o código do tipo selecionado no combo
        tip_sel_label = self.combo_tipmo.get()
        tip_code = "Todos"
        for code, label in self.tipmo_legend.items():
            if label == tip_sel_label:
                tip_code = code
                break
        
        where_tip = f" AND POCF1TipMo = '{tip_code}' " if tip_code != "Todos" else ""
        where_desc = " AND POCF1VlrDs <> 0 " if self.chk_desc_var.get() else ""

        return f"""SELECT 
    POCF1Tst, POCF1Regis, POCF1VlrCx, POCF1VlrDs, POCF1TipMo, 
    CASE RTRIM(POCF1TipMo) 
        WHEN 'V' THEN 'VENDA A VISTA' WHEN 'P' THEN 'VENDA PRAZO' WHEN 'C' THEN 'VENDA CARTÃO'
        WHEN 'S' THEN 'VALE CLIENTE' WHEN 'E' THEN 'VALE POSTO' WHEN 'R' THEN 'RECEBIMENTOS'
        WHEN '$' THEN 'VENDA PRAZO (M)' WHEN 'A' THEN 'BAIXA DE VALE' WHEN 'O' THEN 'OUTROS'
        WHEN 'B' THEN 'BOMBA AGUARD.' WHEN 'M' THEN 'ENTRADA/SAIDA' WHEN 'D' THEN 'DESPESAS'
        WHEN '1' THEN 'SAIU $ ENTROU CARTÃO' WHEN '2' THEN 'SAIU CARTÃO ENTROU $'
        WHEN '3' THEN 'CARTÃO DÉBITO' WHEN '4' THEN 'C.CRÉDITO LJ.CONV.' WHEN '5' THEN 'PIX'
        WHEN '6' THEN 'VENDA CARTÃO DÉBITO' WHEN '7' THEN 'CARTÃO ALELO' WHEN '8' THEN 'TRANSF. $ LOJA'
        ELSE ISNULL(RTRIM(POCF1TipMo), 'ERRO') END as DesTi,
    POCF1Obs, POCF1ProDe, POCF1Qtd, POCF1VlrUn, POCF1DCx
FROM POCF1
WHERE POEmpCod = {posto_id}
  AND (CONVERT(VARCHAR, POCF1TstFe, 120) LIKE '{tst_seg}%' OR CONVERT(VARCHAR, POCF1TstFe, 121) LIKE '{tst_seg}%')
  AND RTRIM(POCF1Usu) = '{self.usu_filter.strip()}'
  {where_tip}
  {where_desc}
ORDER BY POCF1Tst ASC"""

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            sql = self.get_current_query()
            print(f"\n[DEBUG] SQL POCF1: {sql}")
            self.ultima_query_bruta = sql
            cur.execute(sql)
            rows = cur.fetchall()
            from tkinter import messagebox
            if len(rows) == 0:
                tst_msg = self.tst_filter[:19] # Local p/ evitar erro de escopo
                messagebox.showwarning("Diagnóstico POCF1", f"Atenção: Nenhum registro encontrado para este frentista/horário.\n\nFiltro: {tst_msg}\nFrentista: {self.usu_filter}")
            print(f"[DEBUG] POCF1 retornou {len(rows)} registros.")

            for item in self.tree.get_children(): self.tree.delete(item)
            fmt_moeda = lambda v: f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            total_vlr_cx = 0.0
            total_vlr_ds = 0.0

            for r in rows:
                v = list(r)
                # Mapeamento Novo: 0:Tst, 1:Regis, 2:VlrCx, 3:VlrDs, 4:TipMo, 5:DesTi, 6:Obs, 7:ProDe, 8:Qtd, 9:VlrUn, 10:DCx
                
                # Somar totais
                total_vlr_cx += float(v[2] or 0)
                total_vlr_ds += float(v[3] or 0)

                # Formatar Datas: 0(Tst), 10(DCx)
                if v[0] and hasattr(v[0], "strftime"): v[0] = v[0].strftime("%d/%m/%Y %H:%M")
                if v[10] and hasattr(v[10], "strftime"): v[10] = v[10].strftime("%d/%m/%Y")
                # Formatar Moedas: 2(VlrCx), 3(VlrDs), 9(VlrUn)
                for idx in [2, 3, 9]:
                    v[idx] = fmt_moeda(float(v[idx] or 0))
                # Quantidade: 8
                v[8] = f"{float(v[8] or 0):,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
                
                self.tree.insert("", "end", values=tuple(v))

            # Atualizar tree_totais (Padrão Ouro - barra azul alinhada)
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            self.tree_totais.insert("", "end", values=(
                f"REGISTROS: {len(rows)}", "", fmt_moeda(total_vlr_cx), fmt_moeda(total_vlr_ds),
                "", "", "", "", "", "", ""))

            self.re_zebrar(); conn.close()
        except Exception as e:
            print(f"Erro ao carregar detalhes POCF1: {e}")


class AnaliseProdutoWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Análise de Vendas por Produto", "CFG_PRODS")
        self.config_db = config

        # Barra de Filtros (Alinhada à esquerda e transparente)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")


        # LINHA 1
        ctk.CTkLabel(self.filter_frame, text="Ano:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_ano = ctk.CTkComboBox(self.filter_frame, values=["Todos"], width=90, font=("Arial", 15))
        self.combo_ano.pack(side="left", padx=5); self.combo_ano.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Mês:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.combo_mes = ctk.CTkComboBox(self.filter_frame, values=self.meses_ext, width=120, font=("Arial", 15))
        self.combo_mes.pack(side="left", padx=5); self.combo_mes.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Métrica:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.combo_metrica = ctk.CTkComboBox(self.filter_frame, values=["Quantidade", "Valor Venda", "Lucro"], width=130, font=("Arial", 15))
        self.combo_metrica.pack(side="left", padx=5); self.combo_metrica.set("Quantidade")

        ctk.CTkLabel(self.filter_frame, text="Exibir:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.combo_limite = ctk.CTkComboBox(self.filter_frame, values=["Top 10", "Top 20", "Top 50", "Top 100"], width=110, font=("Arial", 15))
        self.combo_limite.pack(side="left", padx=5); self.combo_limite.set("Top 20")

        ctk.CTkLabel(self.filter_frame, text="Produto:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(25, 5))
        self.txt_pesquisa = ctk.CTkEntry(self.filter_frame, placeholder_text="Filtrar por nome...", width=200, font=("Arial", 15))
        self.txt_pesquisa.pack(side="left", padx=5)

        btn_processar = ctk.CTkButton(self.filter_frame, text="🔍 Atualizar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=110, font=("Arial", 13, "bold"))
        btn_processar.pack(side="left", padx=15)

        # Ajuste Colunas
        self.tree.configure(columns=("Rank", "Cod", "Desc", "Qtd", "Venda", "Lucro"))
        self.tree.heading("Rank", text="🏆 #"); self.tree.heading("Cod", text="🔑 Código"); self.tree.heading("Desc", text="📦 Produto"); self.tree.heading("Qtd", text="🛒 Qtd"); self.tree.heading("Venda", text="💰 Total Venda"); self.tree.heading("Lucro", text="📈 Lucro")
        self.tree.column("Rank", width=50, anchor="center"); self.tree.column("Cod", width=90, anchor="center"); self.tree.column("Desc", width=350); self.tree.column("Qtd", width=90, anchor="center"); self.tree.column("Venda", width=160, anchor="e"); self.tree.column("Lucro", width=150, anchor="e")

        # Bootstrap
        self.after(200, self.carregar_anos)


    
    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: ANALISEPRODUTO ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Monta o SQL real que será executado na análise de produtos, com filtros dinâmicos."""
        ano = self.combo_ano.get()
        mes = self.combo_mes.get()
        limite = self.combo_limite.get().replace("Top ", "")
        metrica = self.combo_metrica.get()
        pesquisa = self.txt_pesquisa.get().strip()
        
        where_clauses = ["c4.CRMovDta IS NOT NULL", "YEAR(c4.CRMovDta) != 9999", "c2.CRMov2Flag IN ('A', 'F')"]
        if ano != "Todos": where_clauses.append(f"YEAR(c4.CRMovDta) = {int(ano)}")
        if mes != "Todos": 
            try: mes_idx = self.meses_ext.index(mes); where_clauses.append(f"MONTH(c4.CRMovDta) = {mes_idx}")
            except: pass
        if pesquisa: where_clauses.append(f"p.CEProDes LIKE '%{pesquisa}%'")

        where_str = " AND ".join(where_clauses)
        order_venda = "SUM((c4.CRMov4Qtd * c4.CRMov4VlrU) - ((c4.CRMov4Qtd * c4.CRMov4PreT) * (c4.CRMov4PerD / 100.0)) - c4.CRMov4VlDs - c4.CRMov4VDV - c4.CRMov4VAF)"
        order_custo = "SUM(c4.CRMov4CusP * c4.CRMov4Qtd)"
        
        if metrica == "Valor Venda": order_by = f"{order_venda} DESC"
        elif metrica == "Lucro": order_by = f"({order_venda} - {order_custo}) DESC"
        else: order_by = "SUM(c4.CRMov4Qtd) DESC"

        return f"""SELECT TOP {int(limite)} c4.CEProCod, p.CEProDes, 
    SUM(c4.CRMov4Qtd) as TotalQtd,
    {order_venda} as TotalVenda,
    {order_custo} as TotalCusto
FROM crmov4 c4
INNER JOIN crmov2 c2 ON c4.CMEmpCod = c2.CMEmpCod AND c4.CMFilCod = c2.CMFilCod AND c4.CRMovDta = c2.CRMovDta AND c4.CRMovSeq = c2.CRMovSeq
INNER JOIN cepro p ON c4.CEProCod = p.CEProCod
WHERE {where_str}
GROUP BY c4.CEProCod, p.CEProDes 
ORDER BY {order_by}"""

    def carregar_anos(self):

        server = self.config_db.get("servidor")
        database = self.config_db.get("banco")
        username = self.config_db.get("usuario_bd")
        password = self.config_db.get("senha_bd")
        
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
        try:
            import pyodbc, pandas as pd
            conn = pyodbc.connect(conn_str)
            df_anos = pd.read_sql("SELECT DISTINCT YEAR(CRMovDta) as Ano FROM crmov4 WHERE CRMovDta IS NOT NULL AND YEAR(CRMovDta) != 9999", conn)
            conn.close()
            
            anos = sorted(df_anos['Ano'].dropna().unique().tolist(), reverse=True)
            self.combo_ano.configure(values=["Todos"] + [str(int(a)) for a in anos])
        except: pass

    def carregar_dados(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        server = self.config_db.get("servidor"); database = self.config_db.get("banco"); username = self.config_db.get("usuario_bd"); password = self.config_db.get("senha_bd")
        
        ano = self.combo_ano.get(); mes = self.combo_mes.get(); limite = self.combo_limite.get().replace("Top ", ""); metrica = self.combo_metrica.get()
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
        
        try:
            import pyodbc, pandas as pd
            conn = pyodbc.connect(conn_str)
            where_clauses = ["c4.CRMovDta IS NOT NULL", "YEAR(c4.CRMovDta) != 9999", "c2.CRMov2Flag IN ('A', 'F')"]
            if ano != "Todos": where_clauses.append(f"YEAR(c4.CRMovDta) = {int(ano)}")
            if mes != "Todos": 
                try: mes_idx = self.meses_ext.index(mes); where_clauses.append(f"MONTH(c4.CRMovDta) = {mes_idx}")
                except: pass
            
            pesquisa = self.txt_pesquisa.get().strip()
            if pesquisa: where_clauses.append(f"p.CEProDes LIKE '%{pesquisa}%'")

            where_str = " AND ".join(where_clauses)

            order_venda = "SUM((c4.CRMov4Qtd * c4.CRMov4VlrU) - ((c4.CRMov4Qtd * c4.CRMov4PreT) * (c4.CRMov4PerD / 100.0)) - c4.CRMov4VlDs - c4.CRMov4VDV - c4.CRMov4VAF)"
            order_custo = "SUM(c4.CRMov4CusP * c4.CRMov4Qtd)"
            
            if metrica == "Valor Venda": order_by = f"{order_venda} DESC"
            elif metrica == "Lucro": order_by = f"({order_venda} - {order_custo}) DESC"
            else: order_by = "SUM(c4.CRMov4Qtd) DESC"

            query = f"""
                SELECT TOP {int(limite)} c4.CEProCod, p.CEProDes, 
                       SUM(c4.CRMov4Qtd) as TotalQtd,
                       {order_venda} as TotalVenda,
                       {order_custo} as TotalCusto
                FROM crmov4 c4
                INNER JOIN crmov2 c2 ON c4.CMEmpCod = c2.CMEmpCod AND c4.CMFilCod = c2.CMFilCod AND c4.CRMovDta = c2.CRMovDta AND c4.CRMovSeq = c2.CRMovSeq
                INNER JOIN cepro p ON c4.CEProCod = p.CEProCod
                WHERE {where_str}
                GROUP BY c4.CEProCod, p.CEProDes ORDER BY {order_by}
            """
            self.ultima_query_bruta = query # Auditoria SQL Bruta
            df = pd.read_sql(query, conn); conn.close()

            for index, row in df.iterrows():
                try: cod_pro = str(int(float(row['CEProCod'])))
                except: cod_pro = str(row['CEProCod']).strip()

                try: qtd_int = int(float(row['TotalQtd'])); qtd = f"{qtd_int:,}".replace(",", ".")
                except: qtd = str(row['TotalQtd'])

                try:
                    venda_val = float(row.get('TotalVenda', 0))
                    custo_val = float(row.get('TotalCusto', 0))
                    lucro_val = venda_val - custo_val
                    
                    def format_money(val): return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    
                    venda_str = format_money(venda_val)
                    lucro_str = format_money(lucro_val)
                except:
                    venda_str = lucro_str = "R$ 0,00"

                self.tree.insert("", "end", values=(index + 1, cod_pro, row['CEProDes'].strip(), qtd, venda_str, lucro_str), tags=('even' if index % 2 == 0 else 'odd',))
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro de Filtro", f"Não foi possível carregar os produtos:\n\n{str(e)}")

    def exportar_excel(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        try:
             columns = ["Rank", "Cod", "Desc", "Qtd", "Venda", "Lucro"]
             cols_titles = ["🏆 #", "🔑 Código", "📦 Produto", "🛒 Qtd", "💰 Total Venda", "📈 Lucro"]
             data = [self.tree.item(k)['values'] for k in items]
             import pandas as pd, tempfile, os
             df = pd.DataFrame(data, columns=cols_titles)
             
             # Totalização
             try:
                  sum_qtd = sum([float(str(r[3]).replace('.', '').replace(',', '.')) for r in data if len(r) > 3 and str(r[3]).strip()])
                  sum_venda = sum([float(str(r[4]).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) for r in data if len(r) > 4 and str(r[4]).strip()])
                  sum_lucro = sum([float(str(r[5]).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) for r in data if len(r) > 5 and str(r[5]).strip()])
                  def format_v(v): return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                  df.loc[len(df)] = ["TOTAL", f"({len(data)} regs)", "", f"{sum_qtd:,.0f}".replace(",", "."), format_v(sum_venda), format_v(sum_lucro)]
             except: pass

             filepath = os.path.join(tempfile.gettempdir(), "Analise_Produtos.xlsx")
             writer = pd.ExcelWriter(filepath, engine='openpyxl')
             df.to_excel(writer, index=False, sheet_name='Sheet1')
             worksheet = writer.sheets['Sheet1']
             for i, col in enumerate(df.columns):
                  max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                  from openpyxl.utils import get_column_letter
                  worksheet.column_dimensions[get_column_letter(i+1)].width = max_len
             writer.close()
             os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro Excel", str(e))

    def exportar_pdf(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items: messagebox.showwarning("Aviso", "Nenhum dado para exportar."); return
        try:
             columns = ["Rank", "Cod", "Desc", "Qtd", "Venda", "Lucro"]
             cols_titles = ["🏆 #", "🔑 Código", "📦 Produto", "🛒 Qtd", "💰 Total Venda", "📈 Lucro"]
             data = [self.tree.item(k)['values'] for k in items]
             import tempfile, os
             filepath = os.path.join(tempfile.gettempdir(), "Analise_Produtos.pdf")
             import sys, subprocess
             try: import fpdf
             except: subprocess.run([sys.executable, "-m", "pip", "install", "fpdf"], capture_output=True); import fpdf
             pdf = fpdf.FPDF(); pdf.add_page(); pdf.set_font("Arial", 'B', 12); pdf.cell(190, 10, "ANALISE DE PRODUTOS", 0, 1, 'C'); pdf.ln(5)
             pdf.set_font("Arial", 'B', 7); num_cols = len(cols_titles); w_cols = [12, 18, 55, 18, 43, 44]
             for i, c in enumerate(cols_titles): 
                  c_str = str(c).encode('latin-1', 'replace').decode('latin-1')
                  pdf.cell(w_cols[i], 8, c_str, 1, 0, 'C')
             pdf.ln(8); pdf.set_font("Arial", '', 7)
             for r in data:
                  for i, val in enumerate(r): 
                       v_str = str(val)[:25].encode('latin-1', 'replace').decode('latin-1')
                       pdf.cell(w_cols[i], 7, v_str, 1, 0, 'C' if i != 2 else 'L')
                  pdf.ln(7)
             
             # Totalização PDF
             try:
                  sum_qtd = sum([float(str(r[3]).replace('.', '').replace(',', '.')) for r in data if len(r) > 3 and str(r[3]).strip()])
                  sum_venda = sum([float(str(r[4]).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) for r in data if len(r) > 4 and str(r[4]).strip()])
                  sum_lucro = sum([float(str(r[5]).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')) for r in data if len(r) > 5 and str(r[5]).strip()])
                  def format_v(v): return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                  pdf.ln(2); pdf.set_font("Arial", 'B', 7)
                  pdf.cell(w_cols[0]+w_cols[1]+w_cols[2], 8, f"TOTAL ({len(data)} Registros)", 1, 0, 'R')
                  pdf.cell(w_cols[3], 8, f"{sum_qtd:,.0f}".replace(",", "."), 1, 0, 'C')
                  pdf.cell(w_cols[4], 8, format_v(sum_venda), 1, 0, 'C')
                  pdf.cell(w_cols[5], 8, format_v(sum_lucro), 1, 1, 'C')
             except: pass

             pdf.ln(5); pdf.output(filepath); os.startfile(filepath)
        except Exception as e: messagebox.showerror("Erro PDF", str(e))

    def fechar_tela(self):
        if hasattr(self, 'hub') and self.hub:
             self.hub.fechar_modulo_atual()

class AnaliseVendasQuantidadeWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Análise de Vendas: Produtos x Qtde", "VENDAS_QTDE")
        self.config_db = config
        self.rows_cache = []
        self.anos_comp = 3 

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=5)

        # Início dos controles
        ctk.CTkLabel(self.filter_frame, text="Comparar:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_comp = ctk.CTkComboBox(self.filter_frame, values=["2 Anos", "3 Anos", "4 Anos", "5 Anos"], width=100, command=self.mudar_comp)
        self.combo_comp.pack(side="left", padx=5); self.combo_comp.set("3 Anos")

        ctk.CTkLabel(self.filter_frame, text="Grupo:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_grupo = ctk.CTkComboBox(self.filter_frame, width=140, command=self.selecionar_grupo)
        self.combo_grupo.pack(side="left", padx=5)

        ctk.CTkLabel(self.filter_frame, text="Sub-Grupo:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_subgrupo = ctk.CTkComboBox(self.filter_frame, width=140, command=self.selecionar_subgrupo)
        self.combo_subgrupo.pack(side="left", padx=5)

        ctk.CTkLabel(self.filter_frame, text="Produto:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_produto = ctk.CTkComboBox(self.filter_frame, width=180)
        self.combo_produto.pack(side="left", padx=5); self.combo_produto.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Mês:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.combo_mes = ctk.CTkComboBox(self.filter_frame, values=self.meses_ext, width=120)
        self.combo_mes.pack(side="left", padx=5); self.combo_mes.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Gráfico:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_tipo_grafico = ctk.CTkComboBox(self.filter_frame, values=["Barras", "Linhas"], width=100)
        self.combo_tipo_grafico.pack(side="left", padx=5); self.combo_tipo_grafico.set("Barras")

        self.var_detalhar = ctk.BooleanVar(value=False)
        self.check_detalhar = ctk.CTkCheckBox(self.filter_frame, text="Detalhar por Mês", variable=self.var_detalhar, font=("Arial", 11, "bold"))
        self.check_detalhar.pack(side="left", padx=10)

        btn_go = ctk.CTkButton(self.filter_frame, text="📊 Analisar", width=100, fg_color="#1E88E5", command=self.carregar_dados)
        btn_go.pack(side="left", padx=10)

        # Ocultar a Treeview da BaseWindow para focar só em gráficos
        if hasattr(self, 'tree'): self.tree.pack_forget()
        
        # Container de Gráficos (Sempre visível)
        self.charts_frame = ctk.CTkFrame(self.grid_frame, fg_color="white")
        self.charts_frame.pack(fill="both", expand=True)
        
        self.after(200, self.carregar_combos_filtros)

    def mudar_comp(self, val):
        self.anos_comp = int(val.split(" ")[0])

    def carregar_combos_filtros(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor()
                cur.execute("SELECT CEGrpCod, CEGrpDes FROM CEGrp ORDER BY CEGrpDes")
                self.combo_grupo.configure(values=["Todos"] + [f"{int(r[0])} - {str(r[1]).strip()}" for r in cur.fetchall() if r[1]])
                self.combo_grupo.set("Todos")
                self.combo_subgrupo.configure(values=["Todos"])
                self.combo_subgrupo.set("Todos")
                self.atualizar_combo_produtos()
        except: pass

    def selecionar_grupo(self, val):
        if val == "Todos":
            self.combo_subgrupo.configure(values=["Todos"])
            self.combo_subgrupo.set("Todos")
            self.atualizar_combo_produtos()
            return
        
        gid = val.split(" - ")[0]
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor()
                cur.execute(f"SELECT CESgrCod, CESgrDes FROM CESgr WHERE CEGrpCod={gid} ORDER BY CESgrDes")
                subs = ["Todos"] + [f"{int(r[0])} - {str(r[1]).strip()}" for r in cur.fetchall() if r[1]]
                self.combo_subgrupo.configure(values=subs)
                self.combo_subgrupo.set("Todos")
            self.atualizar_combo_produtos()
        except: pass

    def selecionar_subgrupo(self, val):
        self.atualizar_combo_produtos()

    def atualizar_combo_produtos(self):
        try:
            import pyodbc
            grp_v = self.combo_grupo.get()
            sub_v = self.combo_subgrupo.get()
            
            where_grp = f"AND CEGrpCod = {grp_v.split(' - ')[0]}" if grp_v != "Todos" else ""
            where_sub = f"AND CESgrCod = {sub_v.split(' - ')[0]}" if sub_v != "Todos" else ""
            
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor()
                cur.execute(f"SELECT CEProCod, CEProDes FROM CEPro WHERE 1=1 {where_grp} {where_sub} ORDER BY CEProDes")
                prods = ["Todos"] + [f"{int(r[0])} - {str(r[1]).strip().upper()}" for r in cur.fetchall()]
                self.combo_produto.configure(values=prods[:500])
                self.combo_produto.set("Todos")
        except: pass

    def carregar_dados(self):
        try:
            import pyodbc
            ano_corrente = date.today().year 
            mes_sel = self.combo_mes.get()
            grp_v = self.combo_grupo.get()
            sub_v = self.combo_subgrupo.get()
            prod_v = self.combo_produto.get()
            
            necessita_cepro = (grp_v != "Todos" or sub_v != "Todos" or prod_v != "Todos")
            join_sql = "JOIN cepro p ON c4.ceprocod = p.ceprocod" if necessita_cepro else ""
            where_mes = f"AND MONTH(c2.CRMovDta) = {self.meses_ext.index(mes_sel)}" if mes_sel != "Todos" else ""
            where_grp = f"AND p.CEGrpCod = {grp_v.split(' - ')[0]}" if grp_v != "Todos" else ""
            where_sub = f"AND p.CESgrCod = {sub_v.split(' - ')[0]}" if sub_v != "Todos" else ""
            where_prod = f"AND p.CEProCod = {prod_v.split(' - ')[0]}" if prod_v != "Todos" else ""

            self.anos_analisados = sorted([ano_corrente - i for i in range(self.anos_comp + 1)])
            
            sql = f"""
                SELECT YEAR(c2.CRMovDta) as Ano, MONTH(c2.CRMovDta) as Mes, SUM(c4.CRMov4Qtd) as Qtd
                FROM crmov4 c4
                JOIN crmov2 c2 ON c4.cmempcod = c2.cmempcod AND c4.cmfilcod = c2.cmfilcod AND c4.crmovdta = c2.crmovdta AND c4.crmovseq = c2.crmovseq
                {join_sql}
                WHERE c2.CRMov2Flag IN ('A', 'F') AND YEAR(c2.CRMovDta) BETWEEN {min(self.anos_analisados)} AND {ano_corrente}
                {where_mes} {where_grp} {where_sub} {where_prod}
                GROUP BY YEAR(c2.CRMovDta), MONTH(c2.CRMovDta)
                ORDER BY Ano, Mes
            """
            
            self.ultima_query_bruta = sql
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor(); cur.execute(sql)
                # rows = [(2023, 1, 100), (2023, 2, 120)...]
                self.rows_cache = [list(r) for r in cur.fetchall()]
            
            self.renderizar_graficos()
        except Exception as e: messagebox.showerror("Erro BI", str(e))

    def renderizar_graficos(self):
        for w in self.charts_frame.winfo_children(): w.destroy()
        if not self.rows_cache: return
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            
            fig, ax = plt.subplots(figsize=(10, 5))
            detalhar_mes = self.var_detalhar.get()
            tipo_grafico = self.combo_tipo_grafico.get()
            
            def format_val(val): return f"{int(val):,}".replace(",", ".")

            elements_list = []
            labels_info = []
            num_anos = len(self.anos_analisados)

            hoje = date.today()
            ano_atual = hoje.year
            mes_limite = hoje.month - 1 # Só meses fechados para comparação % do ano atual
            if mes_limite == 0: mes_limite = 1 

            # Agregação manual dos dados vindos por mês
            ano_totals = []
            ano_labels_rich = []
            
            for i, ano in enumerate(self.anos_analisados):
                # Total real acumulado (usado na legeda e gráfico anual)
                total_real = sum(r[2] for r in self.rows_cache if r[0] == ano)
                ano_totals.append(total_real)
                
                perc_str = ""
                if i > 0:
                    prev_ano = self.anos_analisados[i-1]
                    # Se estamos calculando a variação do ano atual (ou comparando com ele)
                    if ano == ano_atual:
                        v_atual_comp = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] <= mes_limite)
                        v_prev_comp = sum(r[2] for r in self.rows_cache if r[0] == prev_ano and r[1] <= mes_limite)
                        if v_prev_comp > 0:
                            diff = ((v_atual_comp - v_prev_comp) / v_prev_comp) * 100
                            perc_str = f" ({'+' if diff >= 0 else ''}{diff:.1f}%)"
                    else:
                        # Comparação normal de anos cheios
                        v_atual_full = total_real
                        v_prev_full = sum(r[2] for r in self.rows_cache if r[0] == prev_ano)
                        if v_prev_full > 0:
                            diff = ((v_atual_full - v_prev_full) / v_prev_full) * 100
                            perc_str = f" ({'+' if diff >= 0 else ''}{diff:.1f}%)"
                
                ano_labels_rich.append(f"{ano}: {format_val(total_real)}{perc_str}")

            if detalhar_mes:
                # Eixo X fixo de Janeiro a Dezembro (ou de acordo com os meses_ext)
                meses_labels = self.meses_ext[1:13] 
                x = np.arange(len(meses_labels))
                width = 0.8 / num_anos
                
                for i, ano in enumerate(self.anos_analisados):
                    # Como agora o cache é (Ano, Mes, Qtd), filtramos por ano e mes (1-12)
                    qtds_meses = []
                    for m_idx in range(1, 13):
                        v = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] == m_idx)
                        qtds_meses.append(v)
                    
                    label_l = ano_labels_rich[i]
                    if tipo_grafico == "Barras":
                        offset = (i - (num_anos-1)/2) * width
                        elem = ax.bar(x + offset, qtds_meses, width, label=label_l)
                    else:
                        elem = ax.plot(x, qtds_meses, marker='o', linewidth=2, label=label_l)
                    
                    elements_list.append(elem)
                
                ax.set_xticks(x)
                ax.set_xticklabels(meses_labels, rotation=45, fontsize=9)
                ax.legend(fontsize=11, loc='upper left', bbox_to_anchor=(1, 1))
                labels_info = meses_labels
            else:
                # Visão Anual: Usamos os totais já calculados
                anos_labels = [str(a) for a in self.anos_analisados]
                totais = ano_totals
                
                if tipo_grafico == "Barras":
                    # Cores vibrantes para cada coluna
                    cores_vibrantes = ["#1E88E5", "#D81B60", "#43A047", "#FFB300", "#8E24AA", "#00ACC1", "#F4511E", "#7B1FA2"]
                    cores_plot = [cores_vibrantes[i % len(cores_vibrantes)] for i in range(len(anos_labels))]
                    elem = ax.bar(anos_labels, totais, color=cores_plot, width=0.5)
                    for rect in elem:
                        val = rect.get_height()
                        ax.annotate(format_val(val), xy=(rect.get_x() + rect.get_width() / 2, val),
                                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=9, fontweight="bold")
                    ax.set_xticklabels(ano_labels_rich, rotation=15, fontsize=9)
                else:
                    elem = ax.plot(anos_labels, totais, marker='o', color="#1E88E5", linewidth=3)
                    for i, v in enumerate(totais):
                        ax.annotate(format_val(v), (anos_labels[i], v), xytext=(0, 10), textcoords="offset points", ha='center', fontsize=9, fontweight="bold")
                    ax.set_xticklabels(ano_labels_rich, rotation=15, fontsize=9)
                
                elements_list.append(elem)
                labels_info = anos_labels

            ax.set_title(f"Evolução de Vendas ({tipo_grafico})", fontsize=12, fontweight="bold")
            
            # --- INTERATIVIDADE (Hover) ---
            annot = ax.annotate("", xy=(0,0), xytext=(20,20), textcoords="offset points",
                               bbox=dict(boxstyle="round", fc="white", ec="gray", alpha=0.9),
                               arrowprops=dict(arrowstyle="->"))
            annot.set_visible(False)

            def update_annot(x, y, txt):
                annot.xy = (x, y)
                annot.set_text(txt)
                annot.get_bbox_patch().set_alpha(0.8)

            def hover(event):
                if event.inaxes == ax:
                    found = False
                    for idx_serie, elem in enumerate(elements_list):
                        if tipo_grafico == "Barras":
                            for idx_bar, bar in enumerate(elem):
                                cont, _ = bar.contains(event)
                                if cont:
                                    val = bar.get_height()
                                    txt = f"Ano: {self.anos_analisados[idx_serie] if detalhar_mes else self.anos_analisados[idx_bar]}\nQtde: {format_val(val)}"
                                    if detalhar_mes: txt = f"Mês: {labels_info[idx_bar]}\n" + txt
                                    update_annot(bar.get_x() + bar.get_width()/2, val, txt)
                                    annot.set_visible(True); fig.canvas.draw_idle()
                                    found = True; break
                        else:
                            # Para Linhas, usamos o picker ou proximidade
                            # plot retorna lista de Line2D. elements_list[idx_serie] é uma lista de 1 elemento
                            line = elem[0]
                            cont, ind = line.contains(event)
                            if cont:
                                idx_pt = ind['ind'][0]
                                xdata, ydata = line.get_data()
                                val = ydata[idx_pt]
                                txt = f"Ano: {self.anos_analisados[idx_serie] if detalhar_mes else self.anos_analisados[idx_pt]}\nQtde: {format_val(val)}"
                                if detalhar_mes: txt = f"Mês: {labels_info[idx_pt]}\n" + txt
                                update_annot(xdata[idx_pt], ydata[idx_pt], txt)
                                annot.set_visible(True); fig.canvas.draw_idle()
                                found = True; break
                        if found: break
                    
                    if not found and annot.get_visible():
                        annot.set_visible(False); fig.canvas.draw_idle()

            fig.canvas.mpl_connect("motion_notify_event", hover)
            ax.spines['right'].set_visible(False); ax.spines['top'].set_visible(False)
            fig.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, master=self.charts_frame)
            canvas.draw(); canvas.get_tk_widget().pack(fill="both", expand=True)
            plt.close(fig)
        except Exception as e: print(f"Erro Chart: {e}")





# [Bloco __main__ movido para o final definitivo do arquivo]

class VolumeVendasTicketWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Volume Vendas & Ticket Médio", "VOL_VENDAS_TICKET")
        self.config_db = config
        self.rows_cache = []
        self.anos_comp = 3 

        # --- BARRA DE FILTROS ---
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=5)

        ctk.CTkLabel(self.filter_frame, text="Comparar:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_comp = ctk.CTkComboBox(self.filter_frame, values=["2 Anos", "3 Anos", "4 Anos", "5 Anos"], width=100, command=self.mudar_comp)
        self.combo_comp.pack(side="left", padx=5); self.combo_comp.set("3 Anos")

        ctk.CTkLabel(self.filter_frame, text="Mês:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.combo_mes = ctk.CTkComboBox(self.filter_frame, values=self.meses_ext, width=120)
        self.combo_mes.pack(side="left", padx=5); self.combo_mes.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Métrica:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_metrica = ctk.CTkComboBox(self.filter_frame, values=["Quantidade de Vendas", "Ticket Médio"], width=160, command=lambda _: self.carregar_dados())
        self.combo_metrica.pack(side="left", padx=5); self.combo_metrica.set("Quantidade de Vendas")

        ctk.CTkLabel(self.filter_frame, text="Gráfico:", font=("Arial", 12, "bold")).pack(side="left", padx=5)
        self.combo_tipo_grafico = ctk.CTkComboBox(self.filter_frame, values=["Barras", "Linhas"], width=100, command=lambda _: self.carregar_dados())
        self.combo_tipo_grafico.pack(side="left", padx=5); self.combo_tipo_grafico.set("Barras")

        self.var_detalhar = ctk.BooleanVar(value=False)
        self.check_detalhar = ctk.CTkCheckBox(self.filter_frame, text="Detalhar por Mês", variable=self.var_detalhar, font=("Arial", 11, "bold"), command=self.carregar_dados)
        self.check_detalhar.pack(side="left", padx=10)

        btn_go = ctk.CTkButton(self.filter_frame, text="📊 Analisar", width=100, fg_color="#1E88E5", command=self.carregar_dados)
        btn_go.pack(side="left", padx=10)

        if hasattr(self, 'tree'): self.tree.pack_forget()
        
        # Container Gráficos
        self.charts_frame = ctk.CTkFrame(self.grid_frame, fg_color="white")
        self.charts_frame.pack(fill="both", expand=True)
        self.after(200, self.carregar_dados)

    def mudar_comp(self, val):
        self.anos_comp = int(val.split(" ")[0])

    def carregar_dados(self):
        try:
            import pyodbc
            from datetime import date
            ano_corrente = date.today().year 
            mes_sel = self.combo_mes.get()
            
            where_mes = f"AND MONTH(c2.CRMovDta) = {self.meses_ext.index(mes_sel)}" if mes_sel != "Todos" else ""
            self.anos_analisados = sorted([ano_corrente - i for i in range(self.anos_comp + 1)])
            
            # Condição de ignorar dados absurdos e agrupamento direto em CRMov2
            sql = f"""
                SELECT YEAR(c2.CRMovDta) as Ano, MONTH(c2.CRMovDta) as Mes, COUNT(c2.CRMovSeq) as QtdVendas, SUM(c2.CRMov2VlrO) as VlrTotal
                FROM CRMov2 c2
                WHERE c2.CRMov2Flag IN ('A', 'F') 
                  AND YEAR(c2.CRMovDta) BETWEEN {min(self.anos_analisados)} AND {ano_corrente}
                  AND c2.CRMov2VlrO <= 1000
                  {where_mes}
                GROUP BY YEAR(c2.CRMovDta), MONTH(c2.CRMovDta)
                ORDER BY Ano, Mes
            """
            
            self.ultima_query_bruta = sql
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cur = conn.cursor(); cur.execute(sql)
                self.rows_cache = [list(r) for r in cur.fetchall()]
            
            self.renderizar_graficos()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro BI", str(e))

    def renderizar_graficos(self):
        for w in self.charts_frame.winfo_children(): w.destroy()
        if not self.rows_cache: 
            ctk.CTkLabel(self.charts_frame, text="🔍 Nenhum dado encontrado para os filtros selecionados.", font=("Arial", 14, "bold")).pack(pady=50)
            return
        try:
            import matplotlib.pyplot as plt
            import numpy as np
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from datetime import date
            
            fig, ax = plt.subplots(figsize=(10, 5))
            detalhar_mes = self.var_detalhar.get()
            tipo_grafico = self.combo_tipo_grafico.get()
            metrica_sel = self.combo_metrica.get()
            
            # Cores vibrantes premium
            cores_vibrantes = ["#1E88E5", "#D81B60", "#43A047", "#FFB300", "#8E24AA", "#00ACC1", "#F4511E", "#7B1FA2", "#2E7D32", "#C62828"]

            def format_val(val): 
                if metrica_sel == "Ticket Médio": return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                return f"{int(val):,}".replace(",", ".")

            def format_int(val): return f"{int(val):,}".replace(",", ".")
            def format_moeda(val): return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            elements_list = []
            labels_info = []
            num_anos = len(self.anos_analisados)

            hoje = date.today()
            ano_atual = hoje.year
            mes_limite = hoje.month - 1
            if mes_limite <= 0: mes_limite = 1 

            ano_totals = []
            ano_labels_rich = []
            
            for i, ano in enumerate(self.anos_analisados):
                if metrica_sel == "Ticket Médio":
                    total_vol = sum(r[2] for r in self.rows_cache if r[0] == ano)
                    total_vlr = sum(r[3] for r in self.rows_cache if r[0] == ano)
                    total_real = total_vlr / total_vol if total_vol > 0 else 0
                else:
                    total_real = sum(r[2] for r in self.rows_cache if r[0] == ano)
                
                ano_totals.append(total_real)
                
                perc_str = ""
                if i > 0:
                    prev_ano = self.anos_analisados[i-1]
                    # Comparação literal para o delta
                    if metrica_sel == "Ticket Médio":
                        v_atual = total_real
                        v_prev_vol = sum(r[2] for r in self.rows_cache if r[0] == prev_ano)
                        v_prev_vlr = sum(r[3] for r in self.rows_cache if r[0] == prev_ano)
                        v_prev = v_prev_vlr / v_prev_vol if v_prev_vol > 0 else 0
                    else:
                        v_atual = total_real
                        v_prev = sum(r[2] for r in self.rows_cache if r[0] == prev_ano)

                    if ano == ano_atual:
                        # Para o ano atual, comparamos YTD
                        if metrica_sel == "Ticket Médio":
                            v_at_vol = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] <= mes_limite)
                            v_at_vlr = sum(r[3] for r in self.rows_cache if r[0] == ano and r[1] <= mes_limite)
                            v_at_comp = v_at_vlr / v_at_vol if v_at_vol > 0 else 0
                            v_pr_vol = sum(r[2] for r in self.rows_cache if r[0] == prev_ano and r[1] <= mes_limite)
                            v_pr_vlr = sum(r[3] for r in self.rows_cache if r[0] == prev_ano and r[1] <= mes_limite)
                            v_pr_comp = v_pr_vlr / v_pr_vol if v_pr_vol > 0 else 0
                        else:
                            v_at_comp = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] <= mes_limite)
                            v_pr_comp = sum(r[2] for r in self.rows_cache if r[0] == prev_ano and r[1] <= mes_limite)
                        
                        if v_pr_comp > 0:
                            diff = ((v_at_comp - v_pr_comp) / v_pr_comp) * 100
                            perc_str = f" ({'+' if diff >= 0 else ''}{diff:.1f}%)"
                    else:
                        if v_prev > 0:
                            diff = ((v_atual - v_prev) / v_prev) * 100
                            perc_str = f" ({'+' if diff >= 0 else ''}{diff:.1f}%)"
                
                ano_labels_rich.append(f"{ano}: {format_val(total_real)}{perc_str}")

            if detalhar_mes:
                meses_labels = self.meses_ext[1:13] 
                x = np.arange(len(meses_labels))
                width = 0.8 / num_anos
                
                for i, ano in enumerate(self.anos_analisados):
                    data_meses = []
                    for m_idx in range(1, 13):
                        if metrica_sel == "Ticket Médio":
                            q = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] == m_idx)
                            v = sum(r[3] for r in self.rows_cache if r[0] == ano and r[1] == m_idx)
                            val = v / q if q > 0 else 0
                        else:
                            val = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] == m_idx)
                        data_meses.append(val)
                    
                    label_l = ano_labels_rich[i]
                    cor_ano = cores_vibrantes[i % len(cores_vibrantes)]
                    if tipo_grafico == "Barras":
                        offset = (i - (num_anos-1)/2) * width
                        elem = ax.bar(x + offset, data_meses, width, label=label_l, color=cor_ano)
                    else:
                        elem = ax.plot(x, data_meses, marker='o', linewidth=2, label=label_l, color=cor_ano)
                    
                    elements_list.append(elem)
                
                ax.set_xticks(x)
                ax.set_xticklabels(meses_labels, rotation=45, fontsize=9)
                ax.legend(fontsize=11, loc='upper left', bbox_to_anchor=(1, 1))
                labels_info = meses_labels
            else:
                anos_labels = [str(a) for a in self.anos_analisados]
                totais = ano_totals
                
                if tipo_grafico == "Barras":
                    cores_plot = [cores_vibrantes[idx % len(cores_vibrantes)] for idx in range(len(anos_labels))]
                    elem = ax.bar(anos_labels, totais, color=cores_plot, width=0.5)
                    for rect in elem:
                        val = rect.get_height()
                        ax.annotate(format_val(val), xy=(rect.get_x() + rect.get_width() / 2, val),
                                    xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=9, fontweight="bold")
                    ax.set_xticklabels(ano_labels_rich, rotation=15, fontsize=9)
                else:
                    elem = ax.plot(anos_labels, totais, marker='o', color=cores_vibrantes[0], linewidth=3)
                    for i, v in enumerate(totais):
                        ax.annotate(format_val(v), (anos_labels[i], v), xytext=(0, 10), textcoords="offset points", ha='center', fontsize=9, fontweight="bold")
                    ax.set_xticklabels(ano_labels_rich, rotation=15, fontsize=9)
                
                elements_list.append(elem)
                labels_info = anos_labels

            ax.set_title(f"Evolução de {metrica_sel} ({tipo_grafico})", fontsize=12, fontweight="bold")
            
            annot = ax.annotate("", xy=(0,0), xytext=(20,20), textcoords="offset points",
                               bbox=dict(boxstyle="round", fc="white", ec="gray", alpha=0.9),
                               arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=0.3"))
            annot.set_visible(False)

            def get_ticket_data(is_detalhar, ano, extra_idx):
                if is_detalhar:
                    m = extra_idx + 1 
                    q = sum(r[2] for r in self.rows_cache if r[0] == ano and r[1] == m)
                    v = sum(r[3] for r in self.rows_cache if r[0] == ano and r[1] == m)
                else:
                    q = sum(r[2] for r in self.rows_cache if r[0] == ano)
                    v = sum(r[3] for r in self.rows_cache if r[0] == ano)
                tk = v / q if q > 0 else 0
                return q, tk

            def hover(event):
                if event.inaxes == ax:
                    found = False
                    for idx_serie, elem in enumerate(elements_list):
                        if tipo_grafico == "Barras":
                            for idx_bar, bar in enumerate(elem):
                                cont, _ = bar.contains(event)
                                if cont:
                                    ano_pt = self.anos_analisados[idx_serie] if detalhar_mes else self.anos_analisados[idx_bar]
                                    qt, tk = get_ticket_data(detalhar_mes, ano_pt, idx_bar)
                                    txt = f"Ano: {ano_pt}\nQtd Vendas: {format_int(qt)}\nTicket Médio: {format_moeda(tk)}"
                                    if detalhar_mes: txt = f"Mês: {labels_info[idx_bar]}\n" + txt
                                    annot.xy = (bar.get_x() + bar.get_width()/2, bar.get_height())
                                    annot.set_text(txt)
                                    annot.set_visible(True); fig.canvas.draw_idle()
                                    found = True; break
                        else:
                            line = elem[0]
                            cont, ind = line.contains(event)
                            if cont:
                                idx_pt = ind['ind'][0]
                                ano_pt = self.anos_analisados[idx_serie] if detalhar_mes else self.anos_analisados[idx_pt]
                                qt, tk = get_ticket_data(detalhar_mes, ano_pt, idx_pt)
                                xdata, ydata = line.get_data()
                                txt = f"Ano: {ano_pt}\nQtd Vendas: {format_int(qt)}\nTicket Médio: {format_moeda(tk)}"
                                if detalhar_mes: txt = f"Mês: {labels_info[idx_pt]}\n" + txt
                                annot.xy = (xdata[idx_pt], ydata[idx_pt])
                                annot.set_text(txt)
                                annot.set_visible(True); fig.canvas.draw_idle()
                                found = True; break
                        if found: break
                    if not found and annot.get_visible():
                        annot.set_visible(False); fig.canvas.draw_idle()

            fig.canvas.mpl_connect("motion_notify_event", hover)
            ax.spines['right'].set_visible(False); ax.spines['top'].set_visible(False)
            fig.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, master=self.charts_frame)
            canvas.draw(); canvas.get_tk_widget().pack(fill="both", expand=True)
            plt.close(fig)
        except Exception as e: 
            print(f"Erro Chart VolumeVendas: {e}")
            ctk.CTkLabel(self.charts_frame, text=f"❌ Erro ao gerar gráfico: {e}", text_color="red", font=("Arial", 12)).pack(pady=20)





class ResumoClienteWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Análise de Vendas por Cliente", "RESUMO_CLI")
        self.config_db = config

        # Filtros (Alinhados à esquerda e transparentes)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")


        ctk.CTkLabel(self.filter_frame, text="Ano:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_ano = ctk.CTkComboBox(self.filter_frame, values=["Todos"], width=90, font=("Arial", 15))
        self.combo_ano.pack(side="left", padx=5)
        self.combo_ano.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Mês:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.combo_mes = ctk.CTkComboBox(self.filter_frame, values=["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"], width=130, font=("Arial", 15))
        self.combo_mes.pack(side="left", padx=5)
        self.combo_mes.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Listagem:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.combo_top = ctk.CTkComboBox(self.filter_frame, values=["Top 10", "Top 20", "Top 50", "Top 100", "Todos"], width=110, font=("Arial", 15))
        self.combo_top.pack(side="left", padx=5)
        self.combo_top.set("Top 20")

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=120, font=("Arial", 13, "bold"))
        btn_processar.pack(side="right", padx=15)


        # Ajuste Colunas
        self.tree.configure(columns=("Rank", "Cod", "Nome", "Valor", "Percent"))
        self.tree.heading("Rank", text="#"); self.tree.heading("Cod", text="Cód Cliente"); self.tree.heading("Nome", text="Nome do Cliente"); self.tree.heading("Valor", text="Total Faturado"); self.tree.heading("Percent", text="% Particip.")
        self.tree.column("Rank", width=50, anchor="center"); self.tree.column("Cod", width=90, anchor="center"); self.tree.column("Nome", width=380); self.tree.column("Valor", width=160, anchor="e"); self.tree.column("Percent", width=110, anchor="center")

        self.after(200, self.carregar_anos)

    def get_current_query(self):
        """Monta o SQL real que será executado na análise de clientes, refletindo Ano e Mês em tempo real."""
        ano = self.combo_ano.get()
        mes = self.combo_mes.get()
        top = self.combo_top.get()

        limit = ""
        if top.startswith("Top"): limit = f"TOP {top.replace('Top ', '')} "

        where_clauses = ["CRMov2Flag IN ('A', 'F')", "CRMovDta IS NOT NULL", "YEAR(CRMovDta) != 9999", "CRMov2CodC NOT IN (1)"]
        if ano != "Todos": where_clauses.append(f"YEAR(CRMovDta) = {int(ano)}")
        if mes != "Todos": 
            meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            try: mes_idx = meses_ext.index(mes); where_clauses.append(f"MONTH(CRMovDta) = {mes_idx}")
            except: pass

        where_str = " AND ".join(where_clauses)
        
        return f"""SELECT {limit}
    CRMov2CodC, CRMov2DesC, SUM(CRMov2VlrO) as Valor
FROM crmov2
WHERE {where_str}
GROUP BY CRMov2CodC, CRMov2DesC
ORDER BY SUM(CRMov2VlrO) DESC"""

    
    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: RESUMOCLIENTE ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_anos(self):

        try:
            import pyodbc; import pandas as pd
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str)
            df_anos = pd.read_sql("SELECT DISTINCT YEAR(CRMovDta) as Ano FROM crmov2 WHERE CRMovDta IS NOT NULL AND YEAR(CRMovDta) != 9999", conn)
            anos = sorted(df_anos['Ano'].dropna().unique().tolist(), reverse=True)
            self.combo_ano.configure(values=["Todos"] + [str(int(a)) for a in anos])
            self.combo_ano.set("Todos")
        except: pass

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()

            ano = self.combo_ano.get()
            mes = self.combo_mes.get()
            top = self.combo_top.get()

            limit = ""
            if top.startswith("Top"): limit = f"TOP {top.replace('Top ', '')} "

            where_clauses = ["CRMov2Flag IN ('A', 'F')", "CRMovDta IS NOT NULL", "YEAR(CRMovDta) != 9999"]
            if ano != "Todos": where_clauses.append(f"YEAR(CRMovDta) = {int(ano)}")
            if mes != "Todos": 
                meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                try: mes_idx = meses_ext.index(mes); where_clauses.append(f"MONTH(CRMovDta) = {mes_idx}")
                except: pass

            where_str = " AND ".join(where_clauses)

            # Query para calcular % Total Geral (sem limite de TOP)
            cursor.execute(f"SELECT SUM(CRMov2VlrO) FROM crmov2 WHERE {where_str}")
            total_g_v = cursor.fetchone()[0]; total_geral = float(total_g_v) if total_g_v else 0.0

            query = f"""
                SELECT {limit}
                    CRMov2CodC, CRMov2DesC, SUM(CRMov2VlrO) as Valor
                FROM crmov2
                WHERE {where_str} AND CRMov2CodC NOT IN (1)
                GROUP BY CRMov2CodC, CRMov2DesC
                ORDER BY Valor DESC
            """
            self.ultima_query_bruta = query # Auditoria SQL Bruta
            cursor.execute(query)
            rows = cursor.fetchall(); conn.close()

            for item in self.tree.get_children(): self.tree.delete(item)
            if not rows: return

            for idx, r_item in enumerate(rows):
                cod_c = str(int(r_item[0])) if r_item[0] is not None else "0"
                nome = str(r_item[1]).strip() if r_item[1] else "Desconhecido"
                valor = float(r_item[2]) if r_item[2] else 0.0
                percent = (valor / total_geral * 100) if total_geral > 0 else 0

                v_f = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                p_f = f"{percent:.1f} %"

                tag = 'even' if idx % 2 == 0 else 'odd'
                self.tree.insert("", "end", values=(idx + 1, cod_c, nome, v_f, p_f), tags=(tag,))
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar Clientes", f"Não foi possível listar os clientes:\n\n{str(e)}")

class ClientesPararamWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Clientes Inativos", "CLI_INATIVOS")
        self.config_db = config
        self.emp_fil_cache = {} # Cache para armazenar (emp, fil) por item_id da treeview

        # Filtros (Alinhados à esquerda e transparentes)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")


        ctk.CTkLabel(self.filter_frame, text="Inativo Há:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_dias = ctk.CTkComboBox(self.filter_frame, values=["30 dias", "60 dias", "90 dias", "180 dias", "365 dias"], width=120, font=("Arial", 15))
        self.combo_dias.pack(side="left", padx=5); self.combo_dias.set("90 dias")

        self.entry_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Filtrar por nome...", width=220, font=("Arial", 15))
        self.entry_busca.pack(side="left", padx=10)
        self.entry_busca.bind("<KeyRelease>", self.filtrar_grid)

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=120, font=("Arial", 13, "bold"))
        btn_processar.pack(side="right", padx=15)


        # Ajuste Colunas
        self.tree.configure(columns=("Rank", "Cod", "Nome", "Ultima_Compra", "Dias", "Media", "Contatos"))
        self.tree.heading("Rank", text="#", command=lambda: self.sort_column("Rank", False))
        self.tree.heading("Cod", text="Cód Cliente", command=lambda: self.sort_column("Cod", False))
        self.tree.heading("Nome", text="Nome do Cliente", command=lambda: self.sort_column("Nome", False))
        self.tree.heading("Ultima_Compra", text="Última Compra", command=lambda: self.sort_column("Ultima_Compra", False))
        self.tree.heading("Dias", text="Dias Inativo", command=lambda: self.sort_column("Dias", False))
        self.tree.heading("Media", text="Média/Ano", command=lambda: self.sort_column("Media", False))
        self.tree.heading("Contatos", text="Telefones")
        
        self.tree.column("Rank", width=40, anchor="center"); self.tree.column("Cod", width=80, anchor="center"); self.tree.column("Nome", width=280); self.tree.column("Ultima_Compra", width=110, anchor="center"); self.tree.column("Dias", width=100, anchor="center"); self.tree.column("Media", width=130, anchor="e"); self.tree.column("Contatos", width=110, anchor="center")

        # Clique para ver Telefones
        self.tree.bind("<ButtonRelease-1>", self.on_click_item)

        # Botões de Exportação
        self.btn_exp_excel = ctk.CTkButton(self.bottom_bar, text="📊 Excel", command=self.exportar_excel, fg_color="#43A047", hover_color="#2E7D32", width=90)
        self.btn_exp_excel.pack(side="right", padx=(5, 20), pady=5)
        
        self.btn_exp_pdf = ctk.CTkButton(self.bottom_bar, text="📄 PDF", command=self.exportar_pdf, fg_color="#E53935", hover_color="#B71C1C", width=90)
        self.btn_exp_pdf.pack(side="right", padx=5, pady=5)

        self.after(200, self.carregar_dados)

    def on_click_item(self, event):
        item_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if item_id and col_id == "#7": # Coluna 'Contatos' é a index 7
            self.abrir_telefones(item_id)

    def abrir_telefones(self, item_id):
        values = self.tree.item(item_id, "values")
        if not values: return
        cod_c = values[1]
        nome_c = values[2]
        emp, fil = self.emp_fil_cache.get(item_id, (1, 1))

        janela = ctk.CTkToplevel(self)
        janela.title(f"Telefones de do Cliente {cod_c}")
        janela.geometry("380x320")
        centralizar_tela(janela, 380, 320)
        janela.grab_set()

        ctk.CTkLabel(janela, text=f"👤 {nome_c}", font=("Arial", 14, "bold"), text_color="#1E88E5").pack(pady=(15, 10))
        
        cont_frame = ctk.CTkFrame(janela, fg_color="#F8FAFC", corner_radius=10)
        cont_frame.pack(padx=20, pady=10, fill="both", expand=True)

        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()
            
            query = "SELECT CRCliTel1, CRCliTel2, CRCliTel3, CRCliTel4, CRCliCon FROM crcli WHERE cmempcod = ? AND crclicod = ?"
            cursor.execute(query, (emp, cod_c))
            row = cursor.fetchone(); conn.close()

            if row:
                fones = [str(f).strip() for f in row[:4] if f and str(f).strip()]
                contato = str(row[4]).strip() if row[4] else ""

                if fones:
                    ctk.CTkLabel(cont_frame, text="📞 Ligações / WhatsApp:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(pady=(10, 5))
                    for f in fones:
                        ctk.CTkLabel(cont_frame, text=f"📱 {f}", font=("Arial", 13)).pack(pady=2)
                else:
                    ctk.CTkLabel(cont_frame, text="Nenhum número encontrado.", text_color="gray").pack(pady=20)

                if contato:
                    ctk.CTkLabel(cont_frame, text="🙋‍♂️ Pessoa de Contato:", font=("Arial", 12, "bold"), text_color="#1E293B").pack(pady=(15, 2))
                    ctk.CTkLabel(cont_frame, text=contato, font=("Arial", 13, "italic")).pack()
            else:
                ctk.CTkLabel(cont_frame, text="Cliente não encontrado na CrCli.", text_color="red").pack(pady=30)
        except Exception as e:
            ctk.CTkLabel(cont_frame, text=f"Erro de Busca:\n{str(e)}", text_color="red", wraplength=300).pack(pady=20)

    def sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        def try_convert(x):
            if col in ["Nome", "Contatos"]:
                return str(x[0]).casefold()
            try: return float(x[0].replace('R$ ', '').replace('.', '').replace(',', '.'))
            except: 
                try: return int(x[0].replace(' dias', ''))
                except: return str(x[0]).casefold()

        l.sort(key=try_convert, reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
            self.tree.set(k, "Rank", index + 1)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    def filtrar_grid(self, event=None):
        busca = self.entry_busca.get().strip().upper()
        for item in self.tree.get_children(): self.tree.delete(item)
        if not hasattr(self, 'rows_cache'): return

        self.emp_fil_cache.clear()
        idx_count = 1
        for r_item in self.rows_cache:
            nome = str(r_item[1]).strip().upper()
            if busca and busca not in nome: continue

            cod_c = str(int(r_item[0])) if r_item[0] is not None else "0"
            dta_compra = r_item[2].strftime('%d/%m/%Y') if r_item[2] else "S/ Data"
            dias = int(r_item[3]) if r_item[3] is not None else 0
            media = int(round(float(r_item[4]))) if r_item[4] else 0

            v_f = f"R$ {media:,}".replace(",", ".")
            tag = 'even' if idx_count % 2 == 0 else 'odd'
            
            node_id = self.tree.insert("", "end", values=(idx_count, cod_c, r_item[1].strip(), dta_compra, f"{dias} dias", v_f, "📞 Ver"), tags=(tag,))
            
            # Cache de emp, fil associado ao node_id exato (Incluído contato e tels)
            tels = [r_item[7], r_item[8], r_item[9], r_item[10]]
            tels_str = " / ".join([str(t).strip() for t in tels if t and str(t).strip()])
            contato = str(r_item[11]).strip() if r_item[11] else ""

            self.emp_fil_cache[node_id] = (r_item[5], r_item[6], tels_str, contato) # Salva dados de contato
            idx_count += 1

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: CLIENTESPARARAM ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):

        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()

            dias_inativo = int(self.combo_dias.get().replace(" dias", ""))

            query = f"""
                SELECT 
                    c2.CRMov2CodC, c2.CRMov2DesC, MAX(c2.CRMovDta) as UltimaCompra,
                    DATEDIFF(day, MAX(c2.CRMovDta), GETDATE()) as DiasInativo,
                    SUM(CASE WHEN YEAR(c2.CRMovDta) >= YEAR(GETDATE()) - 1 THEN c2.CRMov2VlrO ELSE 0 END) / 2.0 as MediaAno,
                    MAX(c2.cmempcod) as Emp, MAX(c2.cmfilcod) as Fil,
                    MAX(cli.CRCliTel1) as Tel1, MAX(cli.CRCliTel2) as Tel2, MAX(cli.CRCliTel3) as Tel3, MAX(cli.CRCliTel4) as Tel4, MAX(cli.CRCliCon) as Contato
                FROM crmov2 c2
                LEFT JOIN crcli cli ON c2.cmempcod = cli.cmempcod AND c2.CRMov2CodC = cli.crclicod
                WHERE c2.CRMov2Flag IN ('A', 'F') AND c2.CRMov2CodC NOT IN (1)
                GROUP BY c2.CRMov2CodC, c2.CRMov2DesC
                HAVING MAX(c2.CRMovDta) <= DATEADD(day, -{dias_inativo}, GETDATE())
                ORDER BY MediaAno DESC
            """
            self.ultima_query_bruta = query # Auditoria SQL Bruta
            cursor.execute(query)
            self.rows_cache = cursor.fetchall(); conn.close()
            self.filtrar_grid()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar Inativos", f"Não foi possível listar os clientes inativos:\n\n{str(e)}")

    def exportar_excel(self):
        items = self.tree.get_children()
        from tkinter import messagebox, filedialog
        if not items:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
            return

        try:
            cols = ["#", "Cod", "Nome", "Última Compra", "Dias", "Média/Ano", "Telefones", "Contato"]
            data = []
            for k in items:
                 v = self.tree.item(k)['values']
                 cache_item = self.emp_fil_cache.get(k, ("", "", "", ""))
                 tels = cache_item[2] if len(cache_item) > 2 else ""
                 contato = cache_item[3] if len(cache_item) > 3 else ""
                 
                 v_0 = str(v[0]) if len(v) > 0 else ""
                 v_1 = str(v[1]) if len(v) > 1 else ""
                 v_2 = str(v[2]) if len(v) > 2 else ""
                 v_3 = str(v[3]) if len(v) > 3 else ""
                 v_4 = str(v[4]) if len(v) > 4 else ""
                 v_5 = str(v[5]) if len(v) > 5 else ""

                 data.append([v_0, v_1, v_2, v_3, v_4, v_5, tels, contato])
            import pandas as pd
            df = pd.DataFrame(data, columns=cols)
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Salvar Excel", initialfile="Clientes_Inativos.xlsx")
            if filepath:
                 df.to_excel(filepath, index=False)
                 messagebox.showinfo("Sucesso", f"Excel salvo com sucesso em:\n\n{filepath}")
        except Exception as e:
             messagebox.showerror("Erro Excel", str(e))

    def exportar_pdf(self):
        items = self.tree.get_children()
        from tkinter import messagebox
        if not items:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
            return

        try:
             cols = ["#", "Cod", "Nome", "Ult Cmp", "Dias", "Média \\ Anual", "Telefones"]
             data = []
             for k in items:
                  v = self.tree.item(k)['values']
                  cache_item = self.emp_fil_cache.get(k, ("", "", "", ""))
                  tels = cache_item[2] if len(cache_item) > 2 else ""
                  
                  v_0 = str(v[0]) if len(v) > 0 else ""
                  v_1 = str(v[1]) if len(v) > 1 else ""
                  v_2 = str(v[2])[:25] if len(v) > 2 else ""
                  v_3 = str(v[3]) if len(v) > 3 else ""
                  v_4 = str(v[4]) if len(v) > 4 else ""
                  v_5 = str(v[5]) if len(v) > 5 else ""

                  data.append([v_0, v_1, v_2, v_3, v_4, v_5, tels[:35]])
             import tempfile, os
             filepath = os.path.join(tempfile.gettempdir(), "Clientes_Inativos.pdf")

             import sys, subprocess
             try: import fpdf
             except: 
                 subprocess.run([sys.executable, "-m", "pip", "install", "fpdf"], capture_output=True)
                 import fpdf

             pdf = fpdf.FPDF(orientation='L') 
             pdf.add_page()
             pdf.set_font("Arial", 'B', 14)
             pdf.cell(280, 10, "RELATORIO DE CLIENTES INATIVOS", 0, 1, 'C')
             pdf.ln(5)

             pdf.set_font("Arial", 'B', 10)
             w_cols = [10, 15, 55, 25, 25, 40, 110]
             for i, c in enumerate(cols):
                  pdf.cell(w_cols[i], 8, c, 1, 0, 'C')
             pdf.ln(8)

             pdf.set_font("Arial", '', 9)
             for r in data:
                  f_0 = str(r[0]).encode('latin-1', 'replace').decode('latin-1')
                  f_1 = str(r[1]).encode('latin-1', 'replace').decode('latin-1')
                  f_2 = str(r[2]).encode('latin-1', 'replace').decode('latin-1')
                  f_3 = str(r[3]).encode('latin-1', 'replace').decode('latin-1')
                  f_4 = str(r[4]).encode('latin-1', 'replace').decode('latin-1')
                  f_5 = str(r[5]).encode('latin-1', 'replace').decode('latin-1')
                  f_6 = str(r[6]).encode('latin-1', 'replace').decode('latin-1')
                  
                  pdf.cell(w_cols[0], 7, f_0, 1, 0, 'C')
                  pdf.cell(w_cols[1], 7, f_1, 1, 0, 'C')
                  pdf.cell(w_cols[2], 7, f_2, 1, 0, 'L')
                  pdf.cell(w_cols[3], 7, f_3, 1, 0, 'C')
                  pdf.cell(w_cols[4], 7, f_4, 1, 0, 'C')
                  pdf.cell(w_cols[5], 7, f_5, 1, 0, 'R')
                  pdf.cell(w_cols[6], 7, f_6, 1, 1, 'L')

             pdf.output(filepath)
             os.startfile(filepath)
        except Exception as e:
             messagebox.showerror("Erro PDF", str(e))

class SugestaoCompraWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Sugestão de Compra", "SUGESTAO_COMPRA")
        self.config_db = config

        # Barra de Filtros (Alinhada à esquerda e transparente)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=0, anchor="w")




        ctk.CTkLabel(self.filter_frame, text="Demanda para:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_dias = ctk.CTkComboBox(self.filter_frame, values=["10 dias", "20 dias", "30 dias"], width=110, font=("Arial", 14))
        self.combo_dias.pack(side="left", padx=5); self.combo_dias.set("30 dias")

        ctk.CTkLabel(self.filter_frame, text="Grupo:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_grupo = ctk.CTkComboBox(self.filter_frame, width=150, font=("Arial", 14), command=self.selecionar_grupo)
        self.combo_grupo.pack(side="left", padx=5)

        ctk.CTkLabel(self.filter_frame, text="Sub:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_subgrupo = ctk.CTkComboBox(self.filter_frame, width=150, font=("Arial", 14), command=self.selecionar_subgrupo)
        self.combo_subgrupo.pack(side="left", padx=5)

        self.entry_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Filtrar...", width=140, font=("Arial", 14))
        self.entry_busca.pack(side="left", padx=10)
        self.entry_busca.bind("<KeyRelease>", self.filtrar_grid)

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=110, font=("Arial", 13, "bold"))
        btn_processar.pack(side="left", padx=10, pady=2)

        # Linha inferior para filtros adicionais
        self.check_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.check_frame.pack(side="top", fill="x", padx=20, pady=(0, 5), anchor="w")

        self.check_somente_compra = ctk.CTkCheckBox(self.check_frame, text="Mostrar Somente os que Precisam de Repor o Estoque", font=("Arial", 15, "bold"), text_color="#C62828", fg_color="#C62828", border_color="#B71C1C", hover_color="#EF9A9A", command=self.filtrar_grid)
        self.check_somente_compra.pack(side="left", padx=10)


        # Frame de Tags (Filtros Ativos) - Começa com altura 0 para não criar espaço vago
        self.tags_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent", height=0)
        self.tags_frame.pack(side="top", fill="x", padx=30, pady=0)
        
        # Ajuste para o grid colar no cabeçalho sem nenhum espaço vago
        self.grid_frame.grid_configure(pady=0)





        columns = ("Rank", "Cod", "Produto", "Estoque", "Vendas3M", "MediaDia", "Demanda", "Sugestao")
        self.tree.configure(columns=columns)
        self.tree.heading("Rank", text="#")
        self.tree.heading("Cod", text="Código")
        self.tree.heading("Produto", text="Produto")
        self.tree.heading("Estoque", text="Estoque")
        self.tree.heading("Vendas3M", text="Vend 90d")
        self.tree.heading("MediaDia", text="Média/Dia")
        self.tree.heading("Demanda", text="Demanda")
        self.tree.heading("Sugestao", text="Sugestão")

        w_col = [40, 75, 230, 95, 95, 95, 95, 110]
        for i, c in enumerate(columns):
             self.tree.column(c, width=w_col[i], anchor="center" if c != "Produto" else "w")

        self.grupos_selecionados = {}
        self.subgrupos_selecionados = {}
        self.rows_cache = []
        self.after(100, self.carregar_combos_filtros)
        self.after(300, self.carregar_dados)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: SUGESTAOCOMPRA ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):

        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()

            where_grp = f"p.cegrpcod IN ({','.join(self.grupos_selecionados.keys())})" if self.grupos_selecionados else "1=1"
            where_sgr = f"p.cesgrcod IN ({','.join(self.subgrupos_selecionados.keys())})" if self.subgrupos_selecionados else "1=1"

            query = f"""
                SELECT 
                    p.ceprocod, 
                    p.CEProDes, 
                    ISNULL(SUM(f.ceprofqtda), 0) as Estoque, 
                    ISNULL(SUM(c4.CRMov4Qtd), 0) as Vendas3M
                FROM cepro p
                LEFT JOIN ceprof f ON p.ceprocod = f.ceprocod
                LEFT JOIN crmov4 c4 ON p.ceprocod = c4.CEProCod
                LEFT JOIN crmov2 c2 ON c4.CMEmpCod = c2.CMEmpCod AND c4.CMFilCod = c2.CMFilCod AND c4.CRMovDta = c2.CRMovDta AND c4.CRMovSeq = c2.CRMovSeq
                    AND c2.CRMovDta >= DATEADD(day, -90, GETDATE()) AND c2.CRMov2Flag IN ('A', 'F')
                WHERE {where_grp} AND {where_sgr}
                GROUP BY p.ceprocod, p.CEProDes
                HAVING ISNULL(SUM(c4.CRMov4Qtd), 0) > 0 OR ISNULL(SUM(f.ceprofqtda), 0) > 0
                ORDER BY p.CEProDes
            """
            self.ultima_query_bruta = query # Auditoria SQL Bruta
            cursor.execute(query)
            self.rows_cache = cursor.fetchall(); conn.close()
            self.filtrar_grid()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Dados", f"Erro no agrupamento:\n{str(e)}")

    def filtrar_grid(self, event=None):
        search = self.entry_busca.get().casefold()
        for i in self.tree.get_children(): self.tree.delete(i)
        
        dias_filtro = int(self.combo_dias.get().replace(" dias", ""))
        idx_count = 1

        for r in self.rows_cache:
            des = str(r[1]).strip()
            if search and search not in des.casefold() and str(r[0]) not in search: continue

            est = float(r[2]); vendas = float(r[3])
            media = vendas / 90.0
            demanda = media * dias_filtro
            sug = demanda - est
            if sug < 0: sug = 0

            if hasattr(self, 'check_somente_compra') and self.check_somente_compra.get() and sug <= 0:
                continue

            v_est = f"{int(est)}" if est.is_integer() else f"{est:.2f}"
            v_vend = f"{int(vendas)}" if vendas.is_integer() else f"{vendas:.2f}"
            
            zebra = 'even' if idx_count % 2 == 0 else 'odd'
            tags = ('alerta',) if sug > 0 else (zebra,)
            self.tree.insert("", "end", values=(idx_count, r[0], des, v_est, v_vend, f"{media:.2f}", f"{demanda:.1f}", f"{sug:.1f}"), tags=tags)
            idx_count += 1
        self.tree.tag_configure('alerta', background='#FFEBEE')

    def carregar_combos_filtros(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT CEGrpCod, CEGrpDes FROM CEGrp ORDER BY CEGrpDes")
                valores_g = ["Todos"] + [f"{int(r[0])} - {str(r[1]).strip()}" for r in cursor.fetchall() if r[1]]
                self.combo_grupo.configure(values=valores_g)
                self.combo_grupo.set("Todos")
        except: pass

    def selecionar_grupo(self, valor):
        if valor == "Todos": return
        g_id = valor.split(" - ")[0]
        g_nome = valor.split(" - ")[1]
        self.grupos_selecionados[g_id] = g_nome
        self.renderizar_tags()
        self.combo_grupo.set("Todos")
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            with pyodbc.connect(conn_str) as conn:
                cursor = conn.cursor()
                cursor.execute(f"SELECT CESgrCod, CESgrDes FROM CESgr WHERE CEGrpCod = {g_id} ORDER BY CESgrDes")
                valores_s = ["Todos"] + [f"{int(r[0])} - {str(r[1]).strip()}" for r in cursor.fetchall() if r[1]]
                self.combo_subgrupo.configure(values=valores_s)
                self.combo_subgrupo.set("Todos")
        except: pass

    def selecionar_subgrupo(self, valor):
        if valor == "Todos": return
        s_id = valor.split(" - ")[0]
        s_nome = valor.split(" - ")[1]
        self.subgrupos_selecionados[s_id] = s_nome
        self.renderizar_tags()
        self.combo_subgrupo.set("Todos")

    def renderizar_tags(self):
        for w in self.tags_frame.winfo_children(): w.destroy()
        for k, v in self.grupos_selecionados.items():
            f = ctk.CTkFrame(self.tags_frame, fg_color="#E3F2FD", corner_radius=5, height=25)
            f.pack(side="left", padx=3)
            ctk.CTkLabel(f, text=f"📁 {v}", font=("Arial", 11), text_color="#1565C0").pack(side="left", padx=(8, 4))
            ctk.CTkButton(f, text="×", width=18, height=18, fg_color="transparent", text_color="#C62828", hover_color="#FFCDD2", command=lambda key=k: self.remover_grupo(key)).pack(side="left", padx=(0, 4))
        for k, v in self.subgrupos_selecionados.items():
            f = ctk.CTkFrame(self.tags_frame, fg_color="#E8F5E9", corner_radius=5, height=25)
            f.pack(side="left", padx=3)
            ctk.CTkLabel(f, text=f"📄 {v}", font=("Arial", 11), text_color="#2E7D32").pack(side="left", padx=(8, 4))
            ctk.CTkButton(f, text="×", width=18, height=18, fg_color="transparent", text_color="#C62828", hover_color="#FFCDD2", command=lambda key=k: self.remover_subgrupo(key)).pack(side="left", padx=(0, 4))

    def remover_grupo(self, key):
        if key in self.grupos_selecionados: del self.grupos_selecionados[key]
        self.renderizar_tags()

    def remover_subgrupo(self, key):
        if key in self.subgrupos_selecionados: del self.subgrupos_selecionados[key]
        self.renderizar_tags()

class PosicaoContasReceberWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Posição de Vendas", "POS_VENDAS")
        self.config_db = config

        # Barra de Filtros (Alinhada à esquerda e transparente)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")


        ctk.CTkLabel(self.filter_frame, text="Ano:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_ano = ctk.CTkComboBox(self.filter_frame, values=["Todos"], width=100, font=("Arial", 15))
        self.combo_ano.pack(side="left", padx=5); self.combo_ano.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Mês:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.meses_ext = ["Todos", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        self.combo_mes = ctk.CTkComboBox(self.filter_frame, values=self.meses_ext, width=130, font=("Arial", 15))
        self.combo_mes.pack(side="left", padx=5); self.combo_mes.set("Todos")

        ctk.CTkLabel(self.filter_frame, text="Visualização:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(25, 5))
        self.seg_view = ctk.CTkSegmentedButton(self.filter_frame, values=["Tabela", "Gráfico"], command=self.alternar_visualizacao, width=170, font=("Arial", 13, "bold"))
        self.seg_view.pack(side="left", padx=5)
        self.seg_view.set("Tabela")

        ctk.CTkLabel(self.filter_frame, text="Gráfico:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(15, 5))
        self.combo_grafico = ctk.CTkComboBox(self.filter_frame, values=["Barras", "Pizza", "Linhas"], width=110, font=("Arial", 14))
        self.combo_grafico.pack(side="left", padx=5); self.combo_grafico.set("Barras")

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=120, font=("Arial", 13, "bold"))
        btn_processar.pack(side="right", padx=15)

        # Configuração da Grid usando o Padrão Centralizado
        cols = ("Tipo", "Descricao", "Qtd", "Valor", "Percent")
        heads = ("Cód Tipo", "Tipo de Venda", "Volume Notas", "Total Faturado", "% Particip.")
        wids = (100, 350, 150, 200, 120)
        alns = ("center", "w", "center", "e", "center")
        self.configurar_grid(cols, heads, wids, alns)

        # Container para o Gráfico
        self.chart_frame = ctk.CTkFrame(self.grid_frame, fg_color="#FFFFFF")

        # Legenda
        self.legendas = { 1: "Vista", 2: "Crediário", 5: "Cartão", 13: "Pix" }
 
        # Ajuste de layout: Detalhamento abaixo da barra azul de totais
        self.frame_totais_container.grid(row=2, column=0, sticky="ew", padx=0)
        self.lbl_cartao = ctk.CTkLabel(self, text="", font=("Arial", 13, "bold"), text_color="#1E88E5", anchor="center")
        self.lbl_cartao.grid(row=3, column=0, sticky="ew", pady=5)
        
        # Mover botões para a última linha
        self.bottom_bar.grid(row=4, column=0, sticky="ew")
 
        self.after(200, self.carregar_anos)

    def alternar_visualizacao(self, valor):
        if valor == "Tabela":
            self.chart_frame.pack_forget()
            self.tree.pack(fill="both", expand=True, side="left")
        else:
            self.tree.pack_forget()
            self.chart_frame.pack(fill="both", expand=True, padx=10, pady=10)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: POSICAOCONTASRECEBER ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Retorna o SQL real que será executado para Auditoria."""
        ano = self.combo_ano.get()
        mes = self.combo_mes.get()

        where_clauses = ["c.CRMov2Flag IN ('A', 'F')", "c.CRMovDta IS NOT NULL", "YEAR(c.CRMovDta) != 9999"]
        if ano != "Todos": where_clauses.append(f"YEAR(c.CRMovDta) = {int(ano)}")
        if mes != "Todos": 
            try: mes_idx = self.meses_ext.index(mes); where_clauses.append(f"MONTH(c.CRMovDta) = {mes_idx}")
            except: pass

        where_str = " AND ".join(where_clauses)
        return f"""SELECT 
    c.crmov2sit as Tipo, 
    MAX(r.CRCCrTip) as Categ,
    COUNT(*) as Qtd, 
    SUM(c.CRMov2VlrO) as Valor
FROM crmov2 c
LEFT JOIN crccr r ON c.cmempcod = r.cmempcod 
                AND c.cmfilcod = r.cmfilcod 
                AND c.crccrcod = r.crccrcod
WHERE {where_str}
GROUP BY c.crmov2sit"""

    def carregar_anos(self):

        server = self.config_db.get("servidor")
        database = self.config_db.get("banco")
        username = self.config_db.get("usuario_bd")
        password = self.config_db.get("senha_bd")
        
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
        try:
            import pyodbc, pandas as pd
            conn = pyodbc.connect(conn_str)
            df_anos = pd.read_sql("SELECT DISTINCT YEAR(CRMovDta) as Ano FROM crmov2 WHERE CRMovDta IS NOT NULL AND YEAR(CRMovDta) != 9999", conn)
            conn.close()
            
            anos = sorted(df_anos['Ano'].dropna().unique().tolist(), reverse=True)
            self.combo_ano.configure(values=["Todos"] + [str(int(a)) for a in anos])
            self.combo_ano.set("Todos")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar Anos", f"Não foi possível listar os anos:\n\n{str(e)}")

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()

            query = self.get_current_query()
            self.ultima_query_bruta = query # Auditoria SQL (?)
            
            cursor.execute(query)
            rows = cursor.fetchall(); conn.close()
 
            for item in self.tree.get_children(): self.tree.delete(item)
            for widget in self.chart_frame.winfo_children(): widget.destroy()
 
            if not rows: return
 
            # Agrupamento e Acumuladores em Python
            agrupado = {}
            v_debito = 0.0; v_credito = 0.0
            total_geral = 0.0
 
            for r_item in rows:
                sit = r_item[0] if r_item[0] is not None else 0
                tip = str(r_item[1]).strip().upper() if r_item[1] else ''
                qtd = int(r_item[2])
                vlr = float(r_item[3]) if r_item[3] else 0.0
 
                total_geral += vlr
                if sit not in agrupado: agrupado[sit] = {'qtd': 0, 'vlr': 0.0}
                agrupado[sit]['qtd'] += qtd
                agrupado[sit]['vlr'] += vlr
 
                if sit == 5:
                    if tip == 'D': v_debito += vlr
                    else: v_credito += vlr
 
            # Atualizar Label de Sumário de Cartão
            total_cart = v_debito + v_credito
            if total_cart > 0:
                p_deb = (v_debito / total_cart) * 100
                p_cre = (v_credito / total_cart) * 100
                resumo_txt = f"💳 DETALHAMENTO CARTÃO  ➔  Débito: R$ {int(round(v_debito)):,} ({int(round(p_deb))}%)  |  Crédito: R$ {int(round(v_credito)):,} ({int(round(p_cre))}%)".replace(",", ".")
                self.lbl_cartao.configure(text=resumo_txt)
            else: self.lbl_cartao.configure(text="")
 
            # Preparar linhas ordenadas para Grid e Grafico
            rows_proc = []
            for k, v in agrupado.items(): rows_proc.append((k, v['qtd'], v['vlr']))
            rows_proc = sorted(rows_proc, key=lambda x: x[2], reverse=True)
 
            for idx, row in enumerate(rows_proc):
                sit_cod = row[0]
                qtd = row[1]
                valor = row[2]
                percent = (valor / total_geral * 100) if total_geral > 0 else 0
                
                descricao = self.legendas.get(sit_cod, "Desconhecido")
                valor_f = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                percent_f = f"{percent:.1f} %"
                
                tag = 'even' if idx % 2 == 0 else 'odd'
                self.tree.insert("", "end", values=(str(int(sit_cod)), descricao, f"{qtd:,}".replace(",", "."), valor_f, percent_f), tags=(tag,))

            # Renderizar Gráfico Multinível (BI)
            try:
                 import matplotlib.pyplot as plt
                 from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                 
                 tipo_g = self.combo_grafico.get()
                 labels = [self.legendas.get(row[0], "Desconhecido") for row in rows_proc]
                 valores = [float(row[2]) for row in rows_proc]
                 
                 # Paleta de Cores Premium (Vibrante)
                 colors = ["#1E88E5", "#43A047", "#E53935", "#FB8C00", "#8E24AA", "#00ACC1", "#F4511E", "#3949AB"]
                 # Garantir cores suficientes repetindo a paleta se necessário
                 bar_colors = (colors * (len(labels) // len(colors) + 1))[:len(labels)]

                 fig, ax = plt.subplots(figsize=(8, 4))
                 ax.set_title(f"Faturamento por Tipo de Venda ({tipo_g})", fontsize=12, fontweight="bold", pad=20)

                 if tipo_g == "Barras":
                    bars = ax.bar(labels, valores, color=bar_colors)
                    def format_currency(x, idx): 
                        v_int = int(round(x))
                        p_pct = (x / total_geral * 100) if total_geral > 0 else 0
                        return f"R$ {v_int:,}".replace(",", ".") + f" ({int(p_pct)}%)"
                    ax.bar_label(bars, labels=[format_currency(v, i) for i, v in enumerate(valores)], padding=3, fontsize=8, fontweight="bold")
                    plt.xticks(rotation=15, ha='right', fontsize=9)
                 
                 elif tipo_g == "Pizza":
                    wedges, texts, autotexts = ax.pie(valores, labels=labels, autopct='%1.1f%%', colors=bar_colors, startangle=140, pctdistance=0.85)
                    plt.setp(autotexts, size=8, weight="bold", color="white")
                    plt.setp(texts, size=9)
                    # Adicionar círculo branco no meio para virar Donut (Estética Moderna)
                    centre_circle = plt.Circle((0,0), 0.70, fc='white')
                    fig.gca().add_artist(centre_circle)

                 elif tipo_g == "Linhas":
                    ax.plot(labels, valores, marker='o', markersize=8, linestyle='-', color="#1E88E5", linewidth=3, label="Vendas")
                    ax.fill_between(labels, valores, alpha=0.2, color="#1E88E5")
                    for i, v in enumerate(valores):
                        ax.text(i, v + (max(valores)*0.02), f"R$ {int(v):,}".replace(",", "."), ha='center', fontsize=8, fontweight="bold")
                    plt.xticks(rotation=15, ha='right', fontsize=9)

                 ax.spines['top'].set_visible(False)
                 ax.spines['right'].set_visible(False)
                 fig.tight_layout()

                 canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
                 canvas.draw()
                 canvas.get_tk_widget().pack(fill="both", expand=True)
                 plt.close(fig) 
            except Exception as e:
                print(f"Erro Gráfico: {e}")

            # ATUALIZAR TOTAIS NO RODAPÉ (PADRÃO EXCEL)
            for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
            v_total_f = f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.tree_totais.insert("", "end", values=("", "TOTAL GERAL", f"{int(sum(r[1] for r in rows_proc)):,}".replace(",", "."), v_total_f, "100.0 %"))
            
            # ATUALIZAR CONTADOR DE REGISTROS
            self.lbl_total_regs.configure(text=f"Total: {len(rows_proc)} registros")

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro SQL", f"Erro ao processar dados:\n{e}")

class EstoqueParadoWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Análise de Estoque Parado", "ESTOQUE_PARADO")
        self.config_db = config

        # Filtros (Alinhados à esquerda e transparentes)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", padx=20, pady=(0, 5), anchor="w")


        ctk.CTkLabel(self.filter_frame, text="Sem Giro Há:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_dias = ctk.CTkComboBox(self.filter_frame, values=["30 dias", "60 dias", "90 dias", "180 dias", "365 dias"], width=120, font=("Arial", 15))
        self.combo_dias.pack(side="left", padx=5); self.combo_dias.set("90 dias")

        self.entry_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Filtrar produto...", width=220, font=("Arial", 15))
        self.entry_busca.pack(side="left", padx=15)

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=120, font=("Arial", 13, "bold"))
        btn_processar.pack(side="right", padx=15)



        # --- NOVO PADRAO: Usar configurar_grid ---
        cols = ("Rank", "Cod", "Nome", "Ultima_Venda", "Estoque", "Custo", "Total")
        # Estética Premium nos Cabeçalhos
        heads = ("🏆 #", "🔑 Cód", "📦 Produto", "📅 Última Venda", "📊 Estoque", "💰 Custo", "💵 Total")
        # Larguras otimizadas e stretch
        wids = (50, 90, 480, 155, 140, 140, 180)
        aligns = ("center", "center", "w", "center", "center", "e", "e")
        self.configurar_grid(cols, heads, wids, aligns)

        # Configurações Adicionais para ocupação total
        self.tree.column("Nome", stretch=True)
        self.tree.column("Total", stretch=True)
        
        # Label de Alerta p/ Regras (Apenas se vazio)
        self.lbl_vazio = ctk.CTkLabel(self.grid_frame, text="Use o botão Processar para listar os produtos sem giro.", font=("Arial", 16, "italic"), text_color="gray")
        self.lbl_vazio.place(relx=0.5, rely=0.5, anchor="center")

        # Sincronizar alinhamento e stretch do Rodapé com o Grid Principal
        for c in cols:
            w = self.tree.column(c, 'width')
            anchor = self.tree.column(c, 'anchor')
            stretch = self.tree.column(c, 'stretch')
            self.tree_totais.column(c, width=w, anchor=anchor, stretch=stretch)

    def sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        def try_convert(x):
            try: return float(x[0].replace('R$ ', '').replace('.', '').replace(',', '.'))
            except: 
                try: return int(x[0].replace(' unid', ''))
                except: 
                     if "/" in x[0]: 
                         try: return "".join(reversed(x[0].split("/"))) # yyyymmdd
                         except: pass
                     return str(x[0]).casefold()

        l.sort(key=try_convert, reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
            self.tree.set(k, "Rank", index + 1)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: ESTOQUEPARADO ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Monta o SQL real que será executado na análise de estoque parado, com filtros de dias e busca em tempo real."""
        dias_str = self.combo_dias.get().replace(" dias", "")
        try: dias = int(dias_str)
        except: dias = 90
        
        return f"""SELECT 
    p.CEProCod, 
    p.CEProDes, 
    SUM(f.CEProFQtdA) as Estoque, 
    p.CEProPreCu as Custo,
    MAX(m.CRMovDta) as UltimaVenda
FROM cepro p
INNER JOIN ceprof f ON p.CEProCod = f.CEProCod
LEFT JOIN crmov4 m ON p.CEProCod = m.CEProCod
LEFT JOIN crmov2 c2 ON m.CMEmpCod = c2.CMEmpCod 
                   AND m.CMFilCod = c2.CMFilCod 
                   AND m.CRMovDta = c2.CRMovDta 
                   AND m.CRMovSeq = c2.CRMovSeq
WHERE (c2.CRMov2Flag IS NULL OR c2.CRMov2Flag IN ('A', 'F'))
GROUP BY p.CEProCod, p.CEProDes, p.CEProPreCu
HAVING SUM(f.CEProFQtdA) > 0 
  AND (MAX(m.CRMovDta) <= DATEADD(day, -{dias}, GETDATE()) OR MAX(m.CRMovDta) IS NULL)
ORDER BY (SUM(f.CEProFQtdA) * ISNULL(p.CEProPreCu, 0)) DESC"""

    def carregar_dados(self):

        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cursor = conn.cursor()

            dias = int(self.combo_dias.get().replace(" dias", ""))
            busca = self.entry_busca.get().strip().upper()

            query = f"""
                SELECT 
                    p.CEProCod, 
                    p.CEProDes, 
                    SUM(f.CEProFQtdA) as Estoque, 
                    p.CEProPreCu as Custo,
                    MAX(m.CRMovDta) as UltimaVenda
                FROM cepro p
                INNER JOIN ceprof f ON p.CEProCod = f.CEProCod
                LEFT JOIN crmov4 m ON p.CEProCod = m.CEProCod
                LEFT JOIN crmov2 c2 ON m.CMEmpCod = c2.CMEmpCod 
                                   AND m.CMFilCod = c2.CMFilCod 
                                   AND m.CRMovDta = c2.CRMovDta 
                                   AND m.CRMovSeq = c2.CRMovSeq
                WHERE (c2.CRMov2Flag IS NULL OR c2.CRMov2Flag IN ('A', 'F'))
                GROUP BY p.CEProCod, p.CEProDes, p.CEProPreCu
                HAVING SUM(f.CEProFQtdA) > 0 
                  AND (MAX(m.CRMovDta) <= DATEADD(day, -{dias}, GETDATE()) OR MAX(m.CRMovDta) IS NULL)
                ORDER BY (SUM(f.CEProFQtdA) * ISNULL(p.CEProPreCu, 0)) DESC
            """
            cursor.execute(query)
            rows = cursor.fetchall(); conn.close()

            self.lbl_vazio.place_forget()
            for item in self.tree.get_children(): self.tree.delete(item)

            if not rows: return

            total_parado_geral = 0.0
            total_linhas = 0

            idx_count = 1
            for r_item in rows:
                cod_p = str(int(r_item[0])) if r_item[0] is not None else "0"
                nome = str(r_item[1]).strip() if r_item[1] else "Desconhecido"
                if busca and busca not in nome.upper(): continue

                qtd = float(r_item[2]) if r_item[2] else 0.0
                custo = float(r_item[3]) if r_item[3] else 0.0
                total = qtd * custo
                
                total_parado_geral += total
                total_linhas += 1

                dta_venda = r_item[4].strftime('%d/%m/%Y') if r_item[4] else "Sem Venda"

                custo_int = int(round(custo))
                total_int = int(round(total))

                v_custo = f"R$ {custo_int:,}".replace(",", ".")
                v_total = f"R$ {total_int:,}".replace(",", ".")

                tag = 'even' if idx_count % 2 == 0 else 'odd'
                self.tree.insert("", "end", values=(idx_count, cod_p, nome, dta_venda, f"{int(qtd)} unid", v_custo, v_total), tags=(tag,))
                idx_count += 1

            # Atualizar Novo Rodapé de Totais (Barra Azul Estilo Excel)
            for item in self.tree_totais.get_children():
                self.tree_totais.delete(item)
            
            def fmt_moeda(v): return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            self.tree_totais.insert("", "end", values=(
                "", 
                "TOTAL EM ESTOQUE PARADO NO PERÍODO", 
                "", 
                "", 
                "", 
                "", 
                fmt_moeda(total_parado_geral)
            ))
            
            self.re_zebrar() # Atualiza contador e zebrado

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar Estoque", f"Não foi possível listar o estoque parado:\n\n{str(e)}")

class AnaliseParcelasWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Análise das Parcelas", "ANALISE_PARCELAS")
        self.config_db = config

        # Ocultar a Treeview da classe base (usada apenas no modal)
        self.tree.pack_forget()

        # Botão Processar na Top Bar
        self.btn_processar = ctk.CTkButton(self.top_frame, text="⚙️ Processar Gráfico", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=160, font=("Arial", 13, "bold"))
        self.btn_processar.pack(side="left", padx=20, pady=5)

        # Frame para o Gráfico
        self.chart_frame = ctk.CTkFrame(self.grid_frame, fg_color="white")
        self.chart_frame.pack(side="top", fill="both", expand=True, padx=20, pady=20)

        # Configuração do Matplotlib
        self.fig = Figure(figsize=(10, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.chart_frame)
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=True)

        # Evento de clique no gráfico
        self.canvas.mpl_connect('button_press_event', self.on_click)

        self.df_parcelas = None
        self.faixas_info = [] # Armazena range e label de cada barra
        self.bars = None

        self.after(100, self.carregar_dados)

    def get_current_query(self):
        """Retorna o comando SQL real utilizado para alimentar o gráfico de parcelas em aberto."""
        return """SELECT 
    c2.CRMov2CodC as CodC, 
    c2.CRMov2DesC as Nome, 
    c3.CRMov3DtaV as Vencimento, 
    c3.CRMov3VlAb as Valor
FROM crmov3 c3
INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod 
                   AND c3.CMFilCod = c2.CMFilCod 
                   AND c3.CRMovDta = c2.CRMovDta 
                   AND c3.CRMovSeq = c2.CRMovSeq
WHERE c3.CRMov3VlAb > 0 
  AND c3.CRMov3Flag = 'A'
  AND c2.CRMov2CodC <> 1
  AND c2.CRMov2Flag IN ('A', 'F')
"""

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: ANALISEPARCELAS ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):

        try:
            import pyodbc
            import pandas as pd
            from datetime import datetime

            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str)
            
            # Query com JOIN para pegar dados do cliente e filtro de Flag 'A', excluindo CodC = 1
            query = """
                SELECT 
                    c2.CRMov2CodC as CodC, 
                    c2.CRMov2DesC as Nome, 
                    c3.CRMov3DtaV as Vencimento, 
                    c3.CRMov3VlAb as Valor
                FROM crmov3 c3
                INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod 
                                   AND c3.CMFilCod = c2.CMFilCod 
                                   AND c3.CRMovDta = c2.CRMovDta 
                                   AND c3.CRMovSeq = c2.CRMovSeq
                WHERE c3.CRMov3VlAb > 0 
                  AND c3.CRMov3Flag = 'A'
                  AND c2.CRMov2CodC <> 1
                  AND c2.CRMov2Flag IN ('A', 'F')
            """
            self.df_parcelas = pd.read_sql(query, conn)
            conn.close()

            if self.df_parcelas.empty:
                self.ax.clear()
                self.ax.set_title("NENHUMA PARCELA 'ABERTA' ENCONTRADA")
                self.canvas.draw()
                return

            hoje = datetime.now()
            self.df_parcelas['Vencimento'] = pd.to_datetime(self.df_parcelas['Vencimento'])
            self.df_parcelas['dias_diff'] = (self.df_parcelas['Vencimento'] - hoje).dt.days

            # Faixas definidas pelo usuário
            faixas = [
                # VENCIDAS
                (-99999, -360, "+360d", "#B71C1C"),
                (-360, -180, "+180d", "#C62828"),
                (-180, -90, "+90d", "#D32F2F"),
                (-90, -60, "+60d", "#E53935"),
                (-60, -30, "+30d", "#EF5350"),
                (-30, 0, "Até 30d", "#F44336"),
                # A VENCER
                (0, 31, "Até 30d", "#4CAF50"),
                (31, 61, "30d", "#43A047"),
                (61, 91, "60d", "#388E3C"),
                (91, 181, "90d", "#2E7D32"),
                (181, 361, "180d", "#1B5E20"),
                (361, 99999, "+360d", "#004D40")
            ]

            self.faixas_info = faixas
            labels_x = []
            valores_y = []
            cores = []

            for start, end, label, cor in faixas:
                soma = self.df_parcelas[(self.df_parcelas['dias_diff'] >= start) & (self.df_parcelas['dias_diff'] < end)]['Valor'].sum()
                labels_x.append(label)
                valores_y.append(soma)
                cores.append(cor)

            self.ax.clear()
            self.bars = self.ax.bar(labels_x, valores_y, color=cores, width=0.7, picker=True)
            
            self.ax.set_title("ANÁLISE DE VENCIMENTOS (PARCELAS ABERTAS)", fontsize=14, fontweight='bold', pad=25)
            self.ax.set_ylabel("Valor Total (R$)", fontsize=11)
            self.ax.tick_params(axis='x', rotation=0, labelsize=9)
            
            for bar in self.bars:
                height = bar.get_height()
                if height > 0:
                    self.ax.text(bar.get_x() + bar.get_width()/2., height,
                                f'R$ {height:,.0f}'.replace(',', '.'),
                                ha='center', va='bottom', fontsize=8, fontweight='bold')

            from matplotlib.patches import Patch
            legend_elements = [
                Patch(facecolor='#C62828', label='Parcelas Vencidas'),
                Patch(facecolor='#2E7D32', label='Parcelas A Vencer')
            ]
            self.ax.legend(handles=legend_elements, loc='upper center', bbox_to_anchor=(0.5, -0.12), ncol=2, frameon=False)

            self.ax.spines['top'].set_visible(False)
            self.ax.spines['right'].set_visible(False)
            self.fig.tight_layout()
            self.canvas.draw()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar Parcelas", f"Erro no processamento:\n{str(e)}")

    def on_click(self, event):
        if event.inaxes != self.ax: return
        for i, bar in enumerate(self.bars):
            if bar.contains(event)[0]:
                label_txt = self.faixas_info[i][2]
                start, end = self.faixas_info[i][0], self.faixas_info[i][1]
                tipo = "Vencidas" if start < 0 else "A Vencer"
                self.abrir_detalhes(f"{tipo} ({label_txt})", start, end)
                break

    def abrir_detalhes(self, titulo_faixa, start, end):
        if self.df_parcelas is None: return
        
        # Filtrar dados da faixa
        df_sub = self.df_parcelas[(self.df_parcelas['dias_diff'] >= start) & (self.df_parcelas['dias_diff'] < end)].copy()
        if df_sub.empty: return

        # Criar Janela Toplevel
        top = ctk.CTkToplevel(self)
        top.title(f"Detalhes: {titulo_faixa}")
        top.geometry("800x500")
        top.after(10, lambda: top.focus_force())
        top.grab_set() # Modal

        # Centralizar
        x = self.winfo_screenwidth() // 2 - 400
        y = self.winfo_screenheight() // 2 - 250
        top.geometry(f"800x500+{x}+{y}")

        # Label Título
        ctk.CTkLabel(top, text=f"📄 LISTA DE PARCELAS - {titulo_faixa}", font=("Arial", 16, "bold"), text_color="#1565C0").pack(pady=15)

        # Frame para o Grid
        f_grid = ctk.CTkFrame(top)
        f_grid.pack(fill="both", expand=True, padx=20, pady=10)

        # Treeview com Zebra
        cols = ("Cod", "Nome", "Vencimento", "Valor")
        tree = ttk.Treeview(f_grid, columns=cols, show="headings", height=15)
        tree.pack(fill="both", expand=True, side="left")

        # Configuração de Cores para o Zebra
        tree.tag_configure('even', background='#FFFFFF')
        tree.tag_configure('odd', background='#E2E8F0')

        # Scrollbar
        scb = ttk.Scrollbar(f_grid, orient="vertical", command=tree.yview)
        scb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scb.set)

        tree.heading("Cod", text="Cód. Cliente")
        tree.heading("Nome", text="Nome do Cliente")
        tree.heading("Vencimento", text="Vencimento")
        tree.heading("Valor", text="Valor Aberto")

        tree.column("Cod", width=100, anchor="center")
        tree.column("Nome", width=350)
        tree.column("Vencimento", width=120, anchor="center")
        tree.column("Valor", width=120, anchor="e")

        # Inserir dados com Zebra
        total_acumulado = 0
        for i, (_, row) in enumerate(df_sub.sort_values('Vencimento').iterrows()):
            v_ab = float(row['Valor'])
            total_acumulado += v_ab
            tag = 'even' if i % 2 == 0 else 'odd'
            tree.insert("", "end", values=(
                str(int(row['CodC'])),
                str(row['Nome']).strip(),
                row['Vencimento'].strftime('%d/%m/%Y'),
                f"R$ {v_ab:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ), tags=(tag,))

        # Rodapé com total
        ctk.CTkLabel(top, text=f"💰 TOTAL DESTA FAIXA: R$ {total_acumulado:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), font=("Arial", 14, "bold")).pack(pady=10)

        # Identificador de Tela (Nome Exclusivo) no Canto Inferior Direito
        f_id = ctk.CTkFrame(top, fg_color="transparent")
        f_id.pack(side="bottom", fill="x", padx=20, pady=(0, 10))

        btn_fechar_det = ctk.CTkButton(f_id, text="✖ Fechar", width=90, height=25, fg_color="transparent", hover_color="#E0E0E0", text_color="black", font=("Arial", 11, "bold"), border_width=1, border_color="black", command=top.destroy)
        btn_fechar_det.pack(side="left")

        btn_copy_det = ctk.CTkButton(f_id, text="📋", width=25, height=25, fg_color="transparent", hover_color="#E0E0E0", text_color="black", font=("Arial", 11, "bold"), 
                                    command=lambda: [top.clipboard_clear(), top.clipboard_append("ANALISE_PARCELAS_DET"), messagebox.showinfo("Sucesso", "NOME DA TELA COPIADO")])
        btn_copy_det.pack(side="right")

        lbl_id_det = ctk.CTkLabel(f_id, text="Tela: ANALISE_PARCELAS_DET", font=("Arial", 11, "bold"), text_color="gray")
        lbl_id_det.pack(side="right", padx=5)

        ToolTip(btn_copy_det, "COPIAR NOME TELA")

    def exportar_pdf(self): pass
    def exportar_excel(self): pass

    def exportar_pdf(self): pass
    def exportar_excel(self): pass

class CobrancaClienteWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Cobrança por Cliente", "COBRANCA_CLIENTE")
        self.config_db = config

        # Barra de Filtros
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=5, anchor="w")

        ctk.CTkLabel(self.filter_frame, text="Filtro Atraso:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.combo_atraso = ctk.CTkComboBox(self.filter_frame, values=["Todos", "Até 30 dias", "Até 60 dias", "Até 90 dias", "Superior a 90 dias"], width=180, font=("Arial", 14), command=lambda _: self.carregar_dados())
        self.combo_atraso.pack(side="left", padx=5); self.combo_atraso.set("Todos")

        self.entry_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Buscar cliente...", width=200, font=("Arial", 14))
        self.entry_busca.pack(side="left", padx=10)
        self.entry_busca.bind("<KeyRelease>", lambda _: self.filtrar_grid())

        btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Atualizar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=110, font=("Arial", 13, "bold"))
        btn_processar.pack(side="left", padx=10)

        # --- PADRAO OURO: Usar configurar_grid p/ Alinhamento e Estilo ---
        self.cols_map = {
            "Cod": {"head": "🔑 Código", "align": "center", "width": 80},
            "Nome": {"head": "👤 Nome do Cliente", "align": "w", "width": 300},
            "QtdAtraso": {"head": "🔢 Qtd. Parcelas", "align": "center", "width": 110},
            "ValorAtraso": {"head": "💰 Vlr. Atraso", "align": "e", "width": 140},
            "ValorAVencer": {"head": "⏳ Vlr. a Vencer", "align": "e", "width": 140},
            "ValorTotal": {"head": "📊 Vlr. Total", "align": "e", "width": 140},
            "DtaMaisAtrasada": {"head": "📅 Dta. Atrasada", "align": "center", "width": 110},
            "DiasAtraso": {"head": "🗓️ Dias", "align": "center", "width": 60}
        }
        
        cols = tuple(self.cols_map.keys())
        heads = tuple(m["head"] for m in self.cols_map.values())
        wids = tuple(m["width"] for m in self.cols_map.values())
        aligns = tuple(m["align"] for m in self.cols_map.values())
        
        self.configurar_grid(cols, heads, wids, aligns)
        
        # FORÇAR ALINHAMENTO MILIMÉTRICO (Sync Heading com Column)
        for c, m in self.cols_map.items():
            self.tree.heading(c, anchor=m["align"])
            self.tree.column(c, anchor=m["align"])

        self.tree.column("Nome", stretch=True)
        self.rows_cache = []
        self.reverse_sort = False

        # CLIQUE SIMPLES: Ver parcelas detalhadas (Conforme Standard Ouro)
        self.tree.bind("<Button-1>", self.on_single_click_cobranca)
        self.tree.bind("<Double-1>", self.on_single_click_cobranca) 

    def on_single_click_cobranca(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id: return
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        vals = self.tree.item(item_id)['values']
        self.hub.abrir_modulo_detalhes_cobranca(vals[0], vals[1])

        self.after(100, self.carregar_dados)

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: COBRANCACLIENTE ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Gera a query SQL de cobrança em tempo real com filtros de atraso dinâmicos."""
        emp = self.config_db.get("empresa", "01")
        f_atr = self.combo_atraso.get()
        having = ""
        if f_atr == "Até 30 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 30"
        elif f_atr == "Até 60 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 60"
        elif f_atr == "Até 90 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 90"
        elif f_atr == "Superior a 90 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) > 90"

        return f"""SELECT 
    c2.CRMov2CodC as CodC, MAX(c2.CRMov2DesC) as Nome, 
    COUNT(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN 1 END) as QtdAtraso,
    SUM(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN c3.CRMov3VlAb ELSE 0 END) as ValorAtraso,
    SUM(CASE WHEN c3.CRMov3DtaV >= CAST(GETDATE() AS DATE) THEN c3.CRMov3VlAb ELSE 0 END) as ValorAVencer,
    SUM(c3.CRMov3VlAb) as ValorTotal,
    MIN(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN c3.CRMov3DtaV END) as DtaAtraso,
    MAX(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN DATEDIFF(day, c3.CRMov3DtaV, GETDATE()) ELSE 0 END) as DiasAtraso
FROM crmov3 c3
INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod AND c3.CMFilCod = c2.CMFilCod AND c3.CRMovDta = c2.CRMovDta AND c3.CRMovSeq = c2.CRMovSeq
WHERE c2.CMEmpCod = '{emp}' AND c3.CRMov3VlAb > 0 AND c3.CRMov3Flag = 'A' AND c2.CRMov2Flag IN ('A','F') AND c2.CRMov2CodC <> 1
GROUP BY c2.CRMov2CodC {having}
ORDER BY ValorAtraso DESC"""

    def ordenar_por(self, col, reverse):
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        def clean_val(v, column):
            if not v: return ""
            s = str(v).strip()
            if column == "Nome":
                try: import unidecode; return unidecode.unidecode(s.upper())
                except: return s.upper()

            if "R$" in s:
                try: return float(s.replace("R$", "").replace(".", "").replace(",", ".").strip())
                except: return 0.0
            if s.isdigit(): return int(s)
            if "/" in s and len(s) == 10:
                try: import datetime; return datetime.datetime.strptime(s, "%d/%m/%Y")
                except: pass
            return s.upper()

        data.sort(key=lambda t: clean_val(t[0], col), reverse=reverse)
        for index, (val, k) in enumerate(data):
            self.tree.move(k, "", index)
        
        # 5. Atualizar Cabeçalhos PRESERVANDO O ANCHOR (Alinhamento)
        for c, m in self.cols_map.items():
            if c == col:
                seta = " ▼" if reverse else " ▲"
                self.tree.heading(c, text=m["head"] + seta, anchor=m["align"], command=lambda _c=c: self.ordenar_por(_c, not reverse))
            else:
                self.tree.heading(c, text=m["head"] + " ↕", anchor=m["align"], command=lambda _c=c: self.ordenar_por(_c, False))
        
        self.re_zebrar()

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str)
            
            f_atr = self.combo_atraso.get()
            having = ""
            if f_atr == "Até 30 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 30"
            elif f_atr == "Até 60 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 60"
            elif f_atr == "Até 90 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) <= 90"
            elif f_atr == "Superior a 90 dias": having = "HAVING MAX(DATEDIFF(day, c3.CRMov3DtaV, GETDATE())) > 90"

            query = f"""
                SELECT 
                    c2.CRMov2CodC as CodC, MAX(c2.CRMov2DesC) as Nome, 
                    COUNT(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN 1 END) as QtdAtraso,
                    SUM(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN c3.CRMov3VlAb ELSE 0 END) as ValorAtraso,
                    SUM(CASE WHEN c3.CRMov3DtaV >= CAST(GETDATE() AS DATE) THEN c3.CRMov3VlAb ELSE 0 END) as ValorAVencer,
                    SUM(c3.CRMov3VlAb) as ValorTotal,
                    MIN(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN c3.CRMov3DtaV END) as DtaAtraso,
                    MAX(CASE WHEN c3.CRMov3DtaV < CAST(GETDATE() AS DATE) THEN DATEDIFF(day, c3.CRMov3DtaV, GETDATE()) ELSE 0 END) as DiasAtraso
                FROM crmov3 c3
                INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod AND c3.CMFilCod = c2.CMFilCod AND c3.CRMovDta = c2.CRMovDta AND c3.CRMovSeq = c2.CRMovSeq
                WHERE c2.CMEmpCod = ? AND c3.CRMov3VlAb > 0 AND c3.CRMov3Flag = 'A' AND c2.CRMov2Flag IN ('A','F') AND c2.CRMov2CodC <> 1
                GROUP BY c2.CRMov2CodC {having}
                ORDER BY ValorAtraso DESC
            """
            cursor = conn.cursor()
            cursor.execute(query, (self.config_db.get("empresa", "01"),))
            self.rows_cache = cursor.fetchall(); conn.close()
            self.filtrar_grid()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Cobrança", f"Erro no processamento:\n{str(e)}")

    def filtrar_grid(self):
        for item in self.tree.get_children(): self.tree.delete(item)
        busca = self.entry_busca.get().strip().upper()

        idx = 0; total_clientes = 0
        s_atraso = 0.0; s_vencer = 0.0; s_total = 0.0

        for row in self.rows_cache:
            cod_c = str(int(row[0]))
            if busca and (busca not in str(row[1]).strip().upper() and busca not in cod_c): continue
            
            total_clientes += 1; idx += 1
            v_atr = float(row[3]); s_atraso += v_atr
            v_ven = float(row[4]); s_vencer += v_ven
            v_tot = float(row[5]); s_total += v_tot

            dta_atr = row[6].strftime("%d/%m/%Y") if row[6] else "-"
            
            tag = 'even' if idx % 2 == 0 else 'odd'
            self.tree.insert("", "end", values=(
                cod_c, row[1].strip(), int(row[2]),
                f"R$ {v_atr:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"R$ {v_ven:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                f"R$ {v_tot:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                dta_atr, int(row[7])
            ), tags=(tag,))

        # --- SINCRONIZAR RODAPÉ (GOLD STANDARD) ---
        fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
        for col in self.tree["columns"]:
            self.tree_totais.column(col, width=self.tree.column(col, "width"))

        self.tree_totais.insert("", "end", values=(
            f"Registros: {total_clientes}", "TOTAL EM ATRASO DO SISTEMA", "",
            fmt(s_atraso), fmt(s_vencer), fmt(s_total), "", ""
        ))


    def exportar_pdf(self): pass
    def exportar_excel(self): pass

class DetalhesParcelasCobrancaWindow(BaseWindow):
    def __init__(self, parent, config, cod_cliente, nome_cliente):
        super().__init__(parent, f"Parcelas do Cliente - {nome_cliente}", "DET_PARCELAS_COBRANCA")
        self.config_db = config
        self.cod_cliente = cod_cliente
        self.nome_cliente = nome_cliente

        # BARRA DE FILTRO (Semaforo Colorido)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=5, anchor="w")

        ctk.CTkLabel(self.filter_frame, text="🚥 Filtrar Parcelas:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.situacao_var = ctk.StringVar(value="ABERTAS")
        self.seg_situacao = ctk.CTkSegmentedButton(self.filter_frame, values=["ABERTAS", "FECHADAS", "AMBAS"], 
                                                   variable=self.situacao_var, height=35, font=("Arial", 13, "bold"),
                                                   command=self.atualizar_visual_semaforo)
        self.seg_situacao.pack(side="left", padx=5)
        
        # Cor Inicial (Verde para Abertas)
        self.seg_situacao.configure(selected_color="#2E7D32") 

        self.configurar_grid(
             columns=("Sit", "DtaV", "VlNo", "VlAb", "DtaP"),
             headings=("🔍 Sit.", "📅 Vencimento", "💰 Valor Original", "💰 Valor em Aberto", "✅ Data Pagamento"),
             widths=(60, 160, 180, 200, 160),
             aligns=("center", "center", "e", "e", "center")
        )
        self.tree.heading("VlNo", anchor="e"); self.tree.heading("VlAb", anchor="e")
        self.grid_frame.grid_configure(padx=0)
        self.tree.column("DtaP", stretch=True)
        
        self.scrollbar = ttk.Scrollbar(self.grid_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        # Zebrado Colorido por Situação (Verdes e Vermelhos)
        self.tree.tag_configure('aberta_even', background='#E8F5E9') # Verde 50
        self.tree.tag_configure('aberta_odd', background='#C8E6C9')  # Verde 100
        self.tree.tag_configure('paga_even', background='#FFEBEE')   # Vermelho 50
        self.tree.tag_configure('paga_odd', background='#FFCDD2')    # Vermelho 100
        self.after(100, self.carregar_dados)

    def fechar_tela(self): self.hub.abrir_cobranca()

    def atualizar_visual_semaforo(self, valor):
        if valor == "ABERTAS": self.seg_situacao.configure(selected_color="#2E7D32", unselected_hover_color="#E8F5E9") 
        elif valor == "FECHADAS": self.seg_situacao.configure(selected_color="#C62828", unselected_hover_color="#FFEBEE") 
        else: self.seg_situacao.configure(selected_color="#F9A825", unselected_hover_color="#FFF9C4") 
        self.carregar_dados()

    def re_zebrar(self):
        """Override p/ zebrado colorido dinâmico por situação (aberta/paga)."""
        for i, item in enumerate(self.tree.get_children()):
            vals = self.tree.item(item)['values']
            prefix = 'aberta' if "🟢" in str(vals[0]) else 'paga'
            suffix = 'even' if i % 2 == 0 else 'odd'
            self.tree.item(item, tags=(f"{prefix}_{suffix}",))

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: DETALHESPARCELASCOBRANCA ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            sit = self.situacao_var.get()
            where_extra = ""
            if sit == "ABERTAS": where_extra = "AND c3.CRMov3Flag = 'A' AND c3.CRMov3VlAb > 0"
            elif sit == "FECHADAS": where_extra = "AND (c3.CRMov3Flag = 'F' OR (c3.CRMov3Flag = 'A' AND c3.CRMov3VlAb = 0))"
            
            sql = f"""
            SELECT c3.CRMov3DtaV, c3.CRMov3VlrO, c3.CRMov3VlAb, c3.CRMov3DtaP 
            FROM crmov3 c3 
            INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod AND c3.CMFilCod = c2.CMFilCod AND c3.CRMovDta = c2.CRMovDta AND c3.CRMovSeq = c2.CRMovSeq 
            WHERE c2.CRMov2CodC = ? AND c2.CMEmpCod = ? AND c2.CRMov2Flag IN ('A','F') {where_extra}
            ORDER BY c3.CRMov3DtaV DESC
            """
            cur.execute(sql, (self.cod_cliente, self.config_db.get("empresa","01")))
            rows = cur.fetchall(); t_nom = 0.0; t_ab = 0.0
            fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            for item in self.tree.get_children(): self.tree.delete(item)
            for idx, r in enumerate(rows):
                dv, v_nom, v_ab, dp = r
                v_ab = float(v_ab or 0); v_nom = float(v_nom or 0)
                t_nom += v_nom; t_ab += v_ab
                
                # Ícone de Situação: 🟢 Aberta | 🔴 Fechada
                sit_icon = "🟢" if v_ab > 0 else "🔴"
                
                # Zebrado dinâmico por cor de situação
                prefix = 'aberta' if v_ab > 0 else 'paga'
                suffix = 'even' if idx % 2 == 0 else 'odd'
                tag = f"{prefix}_{suffix}"
                
                self.tree.insert("", "end", values=(sit_icon, dv.strftime("%d/%m/%Y") if dv else "", fmt(v_nom), fmt(v_ab), dp.strftime("%d/%m/%Y") if dp else "---"), tags=(tag,))
            
            if hasattr(self, 'tree_totais'):
                for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
                self.tree_totais.insert("", "end", values=("", f"Registros: {len(rows)}", fmt(t_nom), fmt(t_ab), ""))
            conn.close()
        except Exception: pass

class ExcluirClientesInativosWindow(BaseWindow):
    def __init__(self, parent, config):
        super().__init__(parent, "Excluir Clientes Inativos (Regra Avançada)", "EXCLUIR_INATIVOS")
        self.config_db = config

        # Barra de Busca e Ações no Header
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="left", fill="x", expand=True, padx=20, pady=(0, 5))

        ctk.CTkLabel(self.filter_frame, text="Buscar:", font=("Arial", 16, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.entry_busca = ctk.CTkEntry(self.filter_frame, placeholder_text="Filtrar por nome ou código...", width=250, font=("Arial", 15))
        self.entry_busca.pack(side="left", padx=5)
        self.entry_busca.bind("<KeyRelease>", self.filtrar_grid)

        self.btn_processar = ctk.CTkButton(self.filter_frame, text="⚙️ Processar", command=self.carregar_dados, fg_color="#1E88E5", hover_color="#1565C0", width=110, font=("Arial", 13, "bold"))
        self.btn_processar.pack(side="left", padx=10)

        self.btn_help_exclusao = ctk.CTkButton(self.filter_frame, text="?", width=30, height=35, fg_color="#546E7A", hover_color="#455A64", font=("Arial", 14, "bold"), command=self.mostrar_ajuda_exclusao)
        self.btn_help_exclusao.pack(side="right", padx=(5, 15))
        ToolTip(self.btn_help_exclusao, "MOSTRAR SEQUÊNCIA DE EXCLUSÃO")

        # Botão de Exclusão (Destaque em Vermelho)
        self.btn_excluir = ctk.CTkButton(self.filter_frame, text="🗑️ EXCLUIR DEFINITIVAMENTE", command=self.confirmar_exclusao, fg_color="#D32F2F", hover_color="#B71C1C", width=220, font=("Arial", 13, "bold"))
        self.btn_excluir.pack(side="right", padx=0)

        # --- PADRAO OURO: Usar configurar_grid ---
        cols = ("Rank", "Cod", "Nome", "Det", "UltVenc", "VlAb", "MedMensal", "Status")
        heads = ("🏆 #", "🔑 Código", "👤 Nome", "📋", "📅 Último Vencimento", "💰 Valor em Aberto", "📈 Média", "🔘")
        wids = (50, 100, 480, 50, 155, 180, 140, 180)
        aligns = ("center", "center", "w", "center", "center", "e", "e", "center")
        self.configurar_grid(cols, heads, wids, aligns)
        
        # 1. Ocultar o tree_totais da Base que está poluindo o rodapé (Solicitado: quadrado vermelho)
        if hasattr(self, 'tree_totais'):
            self.tree_totais.pack_forget()

        # 2. Adicionar barra de rolagem vertical (Scrollbar)
        self.scrollbar = ttk.Scrollbar(self.grid_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")

        # 3. Aproveitar largura total removendo as margens brancas laterais
        self.grid_frame.grid_configure(padx=0)
        self.tree.column("Nome", stretch=True)
        self.tree.column("Status", stretch=True)

        # 1. Clique Simples: Abre as parcelas automaticamente (Standard Ouro)
        self.tree.bind("<Button-1>", self.on_click_tree_standard) 
        self.tree.bind("<Double-1>", self.on_click_tree_standard) 

        # Labels de Totais e Status (Rodapé)
        self.lbl_vlr_geral_parcelas = ctk.CTkLabel(self.bottom_bar, text="", font=("Arial", 13, "bold"), text_color="#1565C0")
        self.lbl_vlr_geral_parcelas.pack(side="left", padx=10)
        
        self.lbl_rodape_sistema = ctk.CTkLabel(self.bottom_bar, text="", font=("Arial", 13, "bold"), text_color="#1565C0")
        self.lbl_rodape_sistema.pack(side="left", padx=10)
        
        self.lbl_cli_geral = ctk.CTkLabel(self.bottom_bar, text="", font=("Arial", 11, "bold"), text_color="gray")
        self.lbl_cli_geral.pack(side="left", padx=5)
        self.lbl_cli_ativos = ctk.CTkLabel(self.bottom_bar, text="", font=("Arial", 11, "bold"), text_color="#43A047")
        self.lbl_cli_ativos.pack(side="left", padx=5)
        self.lbl_cli_inativos_db = ctk.CTkLabel(self.bottom_bar, text="", font=("Arial", 11, "bold"), text_color="#D32F2F")
        self.lbl_cli_inativos_db.pack(side="left", padx=5)

        self.after(200, self.carregar_dados)

    def mostrar_ajuda_exclusao(self):
        msg = (
            "SEQUÊNCIA E CAMPOS DE EXCLUSÃO (CASCATA):\n\n"
            "1. Títulos/Vínculos Secundários (crmov8 a crmov3):\n"
            "   - Campos: CMEmpCod, CRMovDta ou CRMovSeq\n"
            "   - Regra: Remove todos os registros vinculados às vendas do cliente.\n\n"
            "2. Cabeçalho de Vendas (crmov2):\n"
            "   - Campos: CRMov2CodC (ID Cliente), CMEmpCod\n"
            "   - Regra: Remove o histórico de vendas/documentos.\n\n"
            "3. Notas Fiscais (CENFC):\n"
            "   - Campos: CRCliCod (ID Cliente), CMEmpCod\n"
            "   - Regra: Remove vínculos de faturamento.\n\n"
            "4. Cadastro Principal (crcli):\n"
            "   - Campos: CRCliCod (ID Cliente), CMEmpCod\n"
            "   - Regra: Exclusão definitiva do registro do cliente.\n\n"
            "Toda a operação é executada dentro de uma transação SQL (Commit/Rollback)."
        )
        messagebox.showinfo("Detalhes da Exclusão", msg)

    def ordenar_por(self, col, reverse):
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        def clean_val(v, column):
            if not v: return ""
            s = str(v).strip()
            if column == "Nome":
                try: import unidecode; return unidecode.unidecode(s.upper())
                except: return s.upper()
            if column in ["VlAb", "MedMensal"]:
                try: return float(s.replace("R$ ", "").replace(".", "").replace(",", "."))
                except: return 0.0
            if column == "Cod": return int(s) if s.isdigit() else 0
            if column == "UltVenc":
                if s == "N/A" or not s:
                    from datetime import datetime
                    return datetime(1900, 1, 1) # Data mínima para ordenação consistente
                try:
                    from datetime import datetime
                    return datetime.strptime(s, "%d/%m/%Y")
                except: 
                    return s
            if s.isdigit(): return int(s)
            return s.upper()
        data.sort(key=lambda t: clean_val(t[0], col), reverse=reverse)
        for index, (val, k) in enumerate(data):
            self.tree.move(k, "", index)
            self.tree.set(k, "Rank", index + 1)
        for c in self.tree["columns"]:
            if c == col:
                seta = " ▼" if reverse else " ▲"
                self.tree.heading(c, text=self.headers_dict[c] + seta, command=lambda _c=c: self.ordenar_por(_c, not reverse))
            else:
                self.tree.heading(c, text=self.headers_dict[c] + " ↕", command=lambda _c=c: self.ordenar_por(_c, False))

        # Sempre mostrar o 1° registro após ordenar (Auto-Scroll p/ Topo solicitado)
        children = self.tree.get_children()
        if children:
            self.tree.see(children[0])
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: EXCLUIRCLIENTESINATIVOS ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def get_current_query(self):
        """Retorna a query SQL de análise de inatividade (Executada no processamento)."""
        emp = self.config_db.get("empresa", "01")
        return f"""SELECT 
    c.CRCliCod, 
    c.CRCliDes, 
    MAX(c3.CRMov3DtaV) as UltimaDataGeral,
    c.CRCliAti,
    SUM(ISNULL(c3.CRMov3VlAb, 0)) as TotalAberto,
    MAX(CASE WHEN ISNULL(c3.CRMov3VlAb, 0) > 0 THEN c3.CRMov3DtaV ELSE NULL END) as UltimaDataAberta
FROM crcli c
LEFT JOIN crmov2 c2 ON c.CRCliCod = c2.CRMov2CodC AND c.CMEmpCod = c2.CMEmpCod AND c2.CRMov2Flag IN ('A','F')
LEFT JOIN crmov3 c3 ON c2.CMEmpCod = c3.CMEmpCod AND c2.CMFilCod = c3.CMFilCod AND c2.CRMovDta = c3.CRMovDta AND c2.CRMovSeq = c3.CRMovSeq
WHERE c.CMEmpCod = '{emp}'
GROUP BY c.CRCliCod, c.CRCliDes, c.CRCliAti
HAVING 
    c.CRCliAti = 'D'
    OR (
        MAX(c3.CRMov3DtaV) < DATEADD(year, -1, GETDATE())
        AND SUM(ISNULL(c3.CRMov3VlAb, 0)) = 0
        AND MAX(c3.CRMov3DtaV) IS NOT NULL
    )
    OR (
        MAX(CASE WHEN ISNULL(c3.CRMov3VlAb, 0) > 0 THEN c3.CRMov3DtaV ELSE NULL END) < DATEADD(year, -2, GETDATE())
        AND SUM(ISNULL(c3.CRMov3VlAb, 0)) > 0
    )"""

    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            emp_padrao = self.config_db.get("empresa", "01")

            # 1. Buscar Candidatos (Base)
            # Ajustado para considerar apenas parcelas em aberto no filtro de dívidas > 2 anos
            query_base = """
            SELECT 
                c.CRCliCod, 
                c.CRCliDes, 
                MAX(c3.CRMov3DtaV) as UltimaDataGeral,
                c.CRCliAti,
                SUM(ISNULL(c3.CRMov3VlAb, 0)) as TotalAberto,
                MAX(CASE WHEN ISNULL(c3.CRMov3VlAb, 0) > 0 THEN c3.CRMov3DtaV ELSE NULL END) as UltimaDataAberta
            FROM crcli c
            LEFT JOIN crmov2 c2 ON c.CRCliCod = c2.CRMov2CodC AND c.CMEmpCod = c2.CMEmpCod AND c2.CRMov2Flag IN ('A','F')
            LEFT JOIN crmov3 c3 ON c2.CMEmpCod = c3.CMEmpCod AND c2.CMFilCod = c3.CMFilCod AND c2.CRMovDta = c3.CRMovDta AND c2.CRMovSeq = c3.CRMovSeq
            WHERE c.CMEmpCod = ?
            GROUP BY c.CRCliCod, c.CRCliDes, c.CRCliAti
            HAVING 
                c.CRCliAti = 'D'
                OR (
                    MAX(c3.CRMov3DtaV) < DATEADD(year, -1, GETDATE())
                    AND SUM(ISNULL(c3.CRMov3VlAb, 0)) = 0
                    AND MAX(c3.CRMov3DtaV) IS NOT NULL
                )
                OR (
                    MAX(CASE WHEN ISNULL(c3.CRMov3VlAb, 0) > 0 THEN c3.CRMov3DtaV ELSE NULL END) < DATEADD(year, -2, GETDATE())
                    AND SUM(ISNULL(c3.CRMov3VlAb, 0)) > 0
                )
            ORDER BY c.CRCliDes
            """
            cur.execute(query_base, (emp_padrao,))
            self.rows_cache = cur.fetchall()

            # 2. Buscar Histórico Mensal para Cálculo de Média
            query_med = """
            SELECT 
                CRMov2CodC,
                YEAR(CRMovDta) as Ano, 
                MONTH(CRMovDta) as Mes, 
                SUM(CRMov2VlrO) as TotalMes
            FROM crmov2
            WHERE CMEmpCod = ? AND CRMov2Flag IN ('A','F') AND CRMovDta >= DATEADD(month, -48, GETDATE())
            GROUP BY CRMov2CodC, YEAR(CRMovDta), MONTH(CRMovDta)
            """
            cur.execute(query_med, (emp_padrao,))
            history = cur.fetchall()

            # 3. Buscar Soma Total do Sistema (Abertas e Quitadas) e Clientes Geral
            query_totais = """
                SELECT 
                    SUM(ISNULL(c3.CRMov3VlAb, 0)) as SaldoAberto,
                    SUM(ISNULL(c3.CRMov3VlrO, 0) - ISNULL(c3.CRMov3VlAb, 0)) as Pago,
                    SUM(ISNULL(c3.CRMov3VlrO, 0)) as VolumeGeral
                FROM crmov3 c3
                INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod AND c3.CMFilCod = c2.CMFilCod AND c3.CRMovDta = c2.CRMovDta AND c3.CRMovSeq = c2.CRMovSeq
                WHERE c2.CMEmpCod = ? AND c2.CRMov2Flag IN ('A','F')
            """
            cur.execute(query_totais, (emp_padrao,))
            tot_ab, tot_pq, vol_geral = cur.fetchone()
            tot_ab = tot_ab or 0.0; tot_pq = tot_pq or 0.0; vol_geral = vol_geral or 0.0

            cur.execute("""
                SELECT 
                    COUNT(*), 
                    SUM(CASE WHEN CRCliAti<>'D' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN CRCliAti='D' THEN 1 ELSE 0 END)
                FROM crcli WHERE CMEmpCod=?
            """, (emp_padrao,))
            tot_cli_g, tot_cli_a, tot_cli_i = cur.fetchone()
            
            def fmt(v): return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.lbl_rodape_sistema.configure(text=f"🌍 SISTEMA | Abertas: {fmt(tot_ab)} | Quitadas: {fmt(tot_pq)}")
            self.lbl_vlr_geral_parcelas.configure(text=f"Vol. Geral: {fmt(vol_geral)}")
            self.lbl_cli_geral.configure(text=f"Geral: {tot_cli_g}")
            self.lbl_cli_ativos.configure(text=f"Ativos: {tot_cli_a or 0}")
            self.lbl_cli_inativos_db.configure(text=f"Inativos (DB): {tot_cli_i or 0}")

            conn.close()

            # Organizar histórico por cliente
            hist_dict = {}
            for h in history:
                cod_c, ano, mes, vlr = h
                if cod_c not in hist_dict: hist_dict[cod_c] = []
                hist_dict[cod_c].append(float(vlr))

            # Calcular médias com as regras de exclusão de outliers (muito abaixo da média)
            self.medias_cache = {}
            for cod_c, valores in hist_dict.items():
                if not valores: 
                    self.medias_cache[cod_c] = 0.0
                    continue
                
                # Regra 1: Não contar meses sem movimento (já garantido pelo GROUP BY que não traz zeros)
                # Regra 2: Calcular média inicial
                v_filtered = [v for v in valores if v > 0]
                if not v_filtered:
                    self.medias_cache[cod_c] = 0.0
                    continue
                
                avg_inicial = sum(v_filtered) / len(v_filtered)
                
                # Regra 3: Não contar meses em que seja muito abaixo da média (Outliers inferiores)
                # Definimos "muito abaixo" como menos de 10% da média inicial (Solicitado pelo usuário)
                threshold = avg_inicial * 0.10
                v_final = [v for v in v_filtered if v >= threshold]
                
                if not v_final:
                    self.medias_cache[cod_c] = avg_inicial # Fallback caso todos sejam filtrados (improvável)
                else:
                    self.medias_cache[cod_c] = sum(v_final) / len(v_final)

            self.filtrar_grid()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro Carregar", f"Erro no processamento das médias e inativos:\n{str(e)}")

    def filtrar_grid(self, event=None):
        busca = self.entry_busca.get().strip().upper()
        for item in self.tree.get_children(): self.tree.delete(item)
        if not hasattr(self, 'rows_cache'): return

        total = 0
        soma_aberto = 0.0
        def fmt(v): return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        for idx, row in enumerate(self.rows_cache):
            cod_c = str(int(row[0]))
            nome_c = str(row[1]).strip().upper()
            if busca and (busca not in nome_c and busca not in cod_c): continue

            total += 1
            
            # Formatar Data e Valores
            vl_aberto = float(row[4] or 0.0)
            dta_exib = row[5] if vl_aberto > 0 and row[5] else row[2]
            dta_str = dta_exib.strftime("%d/%m/%Y") if dta_exib else "N/A"
            soma_aberto += vl_aberto
            
            vl_txt = fmt(vl_aberto)
            media_v = self.medias_cache.get(row[0], 0.0)
            med_txt = fmt(media_v)

            # Status e Tags
            tags = []
            tags.append('even' if total % 2 == 0 else 'odd')
            
            if row[3] == 'D':
                status_txt = "BLOQUEADO (D)"
                tags.append('status_d')
            elif vl_aberto > 0:
                status_txt = "DÍVIDA > 2 ANOS"
            else:
                status_txt = "INATIVO > 1 ANO"
            
            it = self.tree.insert("", "end", values=(total, cod_c, nome_c, " 🔍 VER ", dta_str, vl_txt, med_txt, status_txt), tags=tuple(tags))
            if total % 10 == 0:
                 self.tree.see(it)
                 self.update_idletasks()
        
        # Atualizar Resumo de Totais
        txt_resumo = f"👥 LISTADOS ({total}) | 💰 TOTAL EM ATRASO: {fmt(soma_aberto)}"
        target_lbl = getattr(self, "lbl_rodape_total", None) or getattr(self, "lbl_rodape_sistema", None)
        if target_lbl:
             target_lbl.configure(text=txt_resumo)

    def on_click_tree_standard(self, event):
        """Gatilho para clique na linha do cliente (Conforme Standard Ouro)."""
        item_id = self.tree.identify_row(event.y)
        if not item_id: return
        
        # Selecionar e focar
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)

        # Abrir tela de detalhes
        self.abrir_detalhes_parcelas(item_id)

    def abrir_detalhes_parcelas(self, item_id=None):
        if not item_id:
            sel = self.tree.selection()
            if not sel: return
            item_id = sel[0]

        item_vals = self.tree.item(item_id)['values']
        cod_cliente = item_vals[1]
        nome_cliente = item_vals[2]
        self.hub.abrir_modulo_detalhes_inativos(cod_cliente, nome_cliente)

    def confirmar_exclusao(self):
        sel = self.tree.selection()
        if not sel:
            from tkinter import messagebox
            messagebox.showwarning("Aviso", "Selecione um cliente para excluir.")
            return
        
        vals = self.tree.item(sel[0])['values']
        cod_c = vals[1]
        nome_c = vals[2]
        
        # Diálogo customizado p/ confirmação de exclusão
        class ConfirmDeleteDialog(ctk.CTkToplevel):
            def __init__(self, master, cod, nome):
                super().__init__(master)
                self.title("🔴 CONFIRMAÇÃO DE EXCLUSÃO")
                self.geometry("550x300")
                self.configure(fg_color="#F8FAFC")
                self.result = False
                self.grab_set(); self.resizable(False, False); self.transient(master)
                
                # Centralizar
                self.update_idletasks()
                x = master.winfo_x() + (master.winfo_width() // 2) - 275
                y = master.winfo_y() + (master.winfo_height() // 2) - 150
                self.geometry(f"+{x}+{y}")
                
                ctk.CTkLabel(self, text="⚠️", font=("Arial", 60)).pack(pady=(20, 5))
                ctk.CTkLabel(self, text="Deseja EXCLUIR DEFINITIVAMENTE o cliente:", font=("Arial", 14), text_color="#1E293B").pack(pady=5)
                ctk.CTkLabel(self, text=f"({cod}) {str(nome).upper()}", font=("Arial", 20, "bold"), text_color="#E11D48", wraplength=500).pack(pady=10, fill="x", padx=20)
                ctk.CTkLabel(self, text="Esta ação não poderá ser desfeita!", font=("Arial", 12, "italic"), text_color="gray").pack(pady=5)
                
                btn_frame = ctk.CTkFrame(self, fg_color="transparent")
                btn_frame.pack(pady=20)
                
                ctk.CTkButton(btn_frame, text="✅ SIM, EXCLUIR", command=self.confirmar, fg_color="#D32F2F", hover_color="#B71C1C", width=160, height=40, font=("Arial", 13, "bold")).pack(side="left", padx=10)
                ctk.CTkButton(btn_frame, text="✖ CANCELAR", command=self.destroy, fg_color="#DEE2E6", text_color="black", width=140, height=40, font=("Arial", 13, "bold")).pack(side="left", padx=10)
                
            def confirmar(self):
                self.result = True
                self.destroy()

        dialog = ConfirmDeleteDialog(self, cod_c, nome_c)
        self.wait_window(dialog)
        
        if dialog.result:
            self.executar_exclusao(cod_c, nome_c)

    def executar_exclusao(self, cod, nome):
        # 1. Dialog customizado para senha (MAIOR)
        class SenhaDialog(ctk.CTkToplevel):
            def __init__(self, master):
                super().__init__(master)
                self.title("Autenticação Necessária")
                self.geometry("450x250")
                self.configure(fg_color="#F8FAFC")
                self.result = None
                self.grab_set() # Modal
                self.resizable(False, False)
                
                # Centralizar
                self.transient(master)
                self.update_idletasks()
                x = master.winfo_x() + (master.winfo_width() // 2) - 225
                y = master.winfo_y() + (master.winfo_height() // 2) - 125
                self.geometry(f"+{x}+{y}")

                ctk.CTkLabel(self, text="⚠️ EXCLUSÃO DEFINITIVA", font=("Arial", 16, "bold"), text_color="#E11D48").pack(pady=(20, 10))
                ctk.CTkLabel(self, text=f"Informe a senha para excluir:\n{nome}", font=("Arial", 13)).pack(pady=5)
                
                self.entry = ctk.CTkEntry(self, show="*", width=300, height=40, font=("Arial", 16), placeholder_text="Senha master...")
                self.entry.pack(pady=10)
                self.entry.focus_set()
                self.entry.bind("<Return>", lambda e: self.confirmar())

                btn_frame = ctk.CTkFrame(self, fg_color="transparent")
                btn_frame.pack(pady=10)
                
                ctk.CTkButton(btn_frame, text="Confirmar", command=self.confirmar, fg_color="#1E293B", width=120, height=35).pack(side="left", padx=5)
                ctk.CTkButton(btn_frame, text="Cancelar", command=self.destroy, fg_color="#DEE2E6", text_color="black", width=120, height=35).pack(side="left", padx=5)

            def confirmar(self):
                self.result = self.entry.get()
                self.destroy()

        if not getattr(self, "senha_autenticada", False):
            dialog = SenhaDialog(self)
            self.wait_window(dialog)
            if dialog.result == "Mabelu@2011":
                self.senha_autenticada = True
            elif dialog.result is None:
                return
            else:
                from tkinter import messagebox
                messagebox.showerror("Erro", "Senha incorreta.")
                return

        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            emp = self.config_db.get("empresa", "01")
            conn.autocommit = False
            try:
                # 2. Ordem de Exclusão Robusta para Evitar Conflitos de FK
                # Verifica colunas existentes para construir os DELETEs sem erro de "coluna inválida"
                for tab in ["crmov8", "crmov7", "crmov6", "crmov5", "crmov4", "crmov3"]:
                    try:
                        # Testar se a coluna CRMovDta existe na tabela
                        cur.execute(f"SELECT TOP 0 CRMovDta FROM {tab}")
                        # Se não deu erro, executa o DELETE por data e seq
                        cur.execute(f"""
                            DELETE FROM {tab} 
                            WHERE CMEmpCod=? AND CRMovDta IN (SELECT CRMovDta FROM crmov2 WHERE CRMov2CodC=? AND CMEmpCod=?)
                        """, (emp, cod, emp))
                    except:
                        # Se CRMovDta não existir, tenta por CRMovSeq se possível
                        try:
                            cur.execute(f"SELECT TOP 0 CRMovSeq FROM {tab}")
                            cur.execute(f"""
                                DELETE FROM {tab} 
                                WHERE CMEmpCod=? AND CRMovSeq IN (SELECT CRMovSeq FROM crmov2 WHERE CRMov2CodC=? AND CMEmpCod=?)
                            """, (emp, cod, emp))
                        except:
                            pass # Tabela não segue o padrão ou não existe

                # Deletar cabeçalho
                cur.execute("DELETE FROM crmov2 WHERE CRMov2CodC=? AND CMEmpCod=?", (cod, emp))
                
                # Deletar Vínculos de Notas Fiscais (Children of CENFC)
                # CENFD (Duplicatas), CENFP (Produtos), CENFO1 (Observações), etc.
                for child in ["CENFD", "CENFP", "CENFO1", "CENFR", "CENFV", "CENFM", "CENFK"]:
                    try:
                        cur.execute(f"""
                            DELETE FROM {child} 
                            WHERE CMEmpCod=? AND CENFCSeq IN (SELECT CENFCSeq FROM CENFC WHERE CRCliCod=? AND CMEmpCod=?)
                        """, (emp, cod, emp))
                    except:
                        pass

                # Deletar Notas Fiscais (CENFC)
                cur.execute("DELETE FROM CENFC WHERE CRCliCod=? AND CMEmpCod=?", (cod, emp))

                # Deletar cliente
                cur.execute("DELETE FROM crcli WHERE CRCliCod=? AND CMEmpCod=?", (cod, emp))
                
                conn.commit()
                class SuccessDeleteDialog(ctk.CTkToplevel):
                    def __init__(self, master, nome):
                        super().__init__(master)
                        self.title("✅ EXCLUSÃO CONCLUÍDA")
                        self.geometry("500x320")
                        self.configure(fg_color="#FFFFFF")
                        self.grab_set() 
                        self.resizable(False, False)
                        self.transient(master)
                        
                        # Centralizar
                        self.update_idletasks()
                        x = master.winfo_x() + (master.winfo_width() // 2) - 250
                        y = master.winfo_y() + (master.winfo_height() // 2) - 160
                        self.geometry(f"+{x}+{y}")

                        try:
                            img_path = os.path.join(os.path.dirname(__file__), "success_check.png")
                            success_img_pil = Image.open(img_path)
                            self.success_img = ctk.CTkImage(success_img_pil, size=(90, 90))
                            ctk.CTkLabel(self, text="", image=self.success_img).pack(pady=(25, 10))
                        except Exception:
                            ctk.CTkLabel(self, text="✅", font=("Arial", 60)).pack(pady=(25, 10))

                        ctk.CTkLabel(self, text="Operação Finalizada!", font=("Arial", 20, "bold"), text_color="#15803D").pack(pady=5)
                        
                        msg = f"O cliente {nome.upper()}\ne seus vínculos foram excluídos com sucesso."
                        ctk.CTkLabel(self, text=msg, font=("Arial", 14), justify="center").pack(pady=(10, 20))
                        
                        ctk.CTkButton(self, text="OK", command=self.destroy, fg_color="#1E293B", hover_color="#334155", width=140, height=40, font=("Arial", 13, "bold")).pack(side="bottom", pady=25)

                SuccessDeleteDialog(self, nome)
                self.carregar_dados()
            except Exception as e_sql:
                conn.rollback(); raise e_sql
            finally: conn.close()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro", f"Erro na exclusão: {e}")

class DetalhesParcelasInativosWindow(BaseWindow):
    def __init__(self, parent, config, cod_cliente, nome_cliente):
        super().__init__(parent, f"Histórico de Parcelas - {nome_cliente}", "DET_PARCELAS_INATIVOS")
        self.config_db = config
        self.cod_cliente = cod_cliente
        self.nome_cliente = nome_cliente

        # BARRA DE FILTRO (Semaforo Colorido)
        self.filter_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        self.filter_frame.pack(side="top", fill="x", padx=20, pady=5, anchor="w")

        ctk.CTkLabel(self.filter_frame, text="🚥 Filtrar Parcelas:", font=("Arial", 14, "bold"), text_color="#1E293B").pack(side="left", padx=(10, 5))
        self.situacao_var = ctk.StringVar(value="ABERTAS")
        self.seg_situacao = ctk.CTkSegmentedButton(self.filter_frame, values=["ABERTAS", "FECHADAS", "AMBAS"], 
                                                   variable=self.situacao_var, height=35, font=("Arial", 13, "bold"),
                                                   command=self.atualizar_visual_semaforo)
        self.seg_situacao.pack(side="left", padx=5)
        self.seg_situacao.configure(selected_color="#2E7D32") 

        self.configurar_grid(
             columns=("Sit", "DtaV", "VlNo", "VlAb", "DtaP"),
             headings=("🔍 Sit.", "📅 Vencimento", "💰 Valor Original", "💰 Valor em Aberto", "✅ Data Pagamento"),
             widths=(60, 160, 180, 200, 160),
             aligns=("center", "center", "e", "e", "center")
        )
        self.tree.heading("VlNo", anchor="e"); self.tree.heading("VlAb", anchor="e")
        self.grid_frame.grid_configure(padx=0)
        self.tree.column("DtaP", stretch=True)
        
        self.scrollbar = ttk.Scrollbar(self.grid_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side="right", fill="y")
        # Zebrado Colorido por Situação (Verdes e Vermelhos)
        self.tree.tag_configure('aberta_even', background='#E8F5E9') # Verde 50
        self.tree.tag_configure('aberta_odd', background='#C8E6C9')  # Verde 100
        self.tree.tag_configure('paga_even', background='#FFEBEE')   # Vermelho 50
        self.tree.tag_configure('paga_odd', background='#FFCDD2')    # Vermelho 100
        self.after(100, self.carregar_dados)

    def fechar_tela(self): self.hub.abrir_excluir_inativos()

    def atualizar_visual_semaforo(self, valor):
        if valor == "ABERTAS": self.seg_situacao.configure(selected_color="#2E7D32", unselected_hover_color="#E8F5E9") 
        elif valor == "FECHADAS": self.seg_situacao.configure(selected_color="#C62828", unselected_hover_color="#FFEBEE") 
        else: self.seg_situacao.configure(selected_color="#F9A825", unselected_hover_color="#FFF9C4") 
        self.carregar_dados()

    def re_zebrar(self):
        """Override p/ zebrado colorido dinâmico por situação (aberta/paga)."""
        for i, item in enumerate(self.tree.get_children()):
            vals = self.tree.item(item)['values']
            prefix = 'aberta' if "🟢" in str(vals[0]) else 'paga'
            suffix = 'even' if i % 2 == 0 else 'odd'
            self.tree.item(item, tags=(f"{prefix}_{suffix}",))

    def get_sql_summary(self):
        """Resumo Técnico da Estrutura SQL - Padrão Ouro"""
        return (
            f"--- ESTRUTURA SQL: DETALHESPARCELASINATIVOS ---\n"
            "TABELAS PRINCIPAIS: Dinâmica por Módulo / Contexto da Tela\n"
            "RELACIONAMENTOS:\n"
            "- Lógicas baseadas em carregar_dados e metadados contextuais.\n\n"
            "MÉTRICAS / CAMPOS EXIBIDOS:\n"
            "- Colunas vinculadas à lógica da visualização atual.\n\n"
            "FILTROS APLICADOS:\n"
            "- Filtros customizados e controle de ID Posso selecionados.\n\n"
            "ORDENAÇÃO:\n"
            "- Ordenação decrescente por Padrão de Chave / Data"
        )
    def carregar_dados(self):
        try:
            import pyodbc
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={self.config_db['servidor']};DATABASE={self.config_db['banco']};UID={self.config_db['usuario_bd']};PWD={self.config_db['senha_bd']}"
            conn = pyodbc.connect(conn_str); cur = conn.cursor()
            
            sit = self.situacao_var.get()
            where_extra = ""
            if sit == "ABERTAS": where_extra = "AND c3.CRMov3Flag = 'A' AND c3.CRMov3VlAb > 0"
            elif sit == "FECHADAS": where_extra = "AND (c3.CRMov3Flag = 'F' OR (c3.CRMov3Flag = 'A' AND c3.CRMov3VlAb = 0))"
            
            sql = f"""
            SELECT c3.CRMov3DtaV, c3.CRMov3VlrO, c3.CRMov3VlAb, c3.CRMov3DtaP 
            FROM crmov3 c3 
            INNER JOIN crmov2 c2 ON c3.CMEmpCod = c2.CMEmpCod AND c3.CMFilCod = c2.CMFilCod AND c3.CRMovDta = c2.CRMovDta AND c3.CRMovSeq = c2.CRMovSeq 
            WHERE c2.CRMov2CodC = ? AND c2.CMEmpCod = ? AND c2.CRMov2Flag IN ('A','F') {where_extra}
            ORDER BY c3.CRMov3DtaV DESC
            """
            cur.execute(sql, (self.cod_cliente, self.config_db.get("empresa","01")))
            rows = cur.fetchall(); t_nom = 0.0; t_ab = 0.0
            fmt = lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            for item in self.tree.get_children(): self.tree.delete(item)
            for idx, r in enumerate(rows):
                dv, v_nom, v_ab, dp = r
                v_ab = float(v_ab or 0); v_nom = float(v_nom or 0)
                t_nom += v_nom; t_ab += v_ab
                
                # Ícone de Situação: 🟢 Aberta | 🔴 Fechada
                sit_icon = "🟢" if v_ab > 0 else "🔴"
                
                # Zebrado dinâmico por cor de situação
                prefix = 'aberta' if v_ab > 0 else 'paga'
                suffix = 'even' if idx % 2 == 0 else 'odd'
                tag = f"{prefix}_{suffix}"
                
                self.tree.insert("", "end", values=(sit_icon, dv.strftime("%d/%m/%Y") if dv else "", fmt(v_nom), fmt(v_ab), dp.strftime("%d/%m/%Y") if dp else "---"), tags=(tag,))
            
            if hasattr(self, 'tree_totais'):
                for item in self.tree_totais.get_children(): self.tree_totais.delete(item)
                self.tree_totais.insert("", "end", values=("", f"Registros: {len(rows)}", fmt(t_nom), fmt(t_ab), ""))
            conn.close()
        except Exception: pass



if __name__ == '__main__':
    import sys, argparse
    parser = argparse.ArgumentParser(description="Módulo de Análise e Atualização")
    parser.add_argument("--tela", type=str, choices=["ANALISE", "CEP"], help="Abre uma tela específica diretamente")
    parser.add_argument("--servidor", type=str, help="Servidor")
    parser.add_argument("--banco", type=str, help="Banco")
    parser.add_argument("--usuario", type=str, help="Usuário")
    parser.add_argument("--senha", type=str, help="Senha")
    
    args = parser.parse_args()
    config = {"nome_usuario": "Operador", "servidor": r"servidor\sqlexpress", "banco": "msinfor", "usuario_bd": "sa", "senha_bd": "Mabelu2011"}
    
    # Carregar configuração persistente (CFG_ACESSO) se disponível
    config_path = os.path.join(os.path.dirname(__file__), "config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config.update(json.load(f))
        except: pass

    if args.servidor: config["servidor"] = args.servidor
    if args.banco: config["banco"] = args.banco
    if args.usuario: config["usuario_bd"] = args.usuario
    if args.senha: config["senha_bd"] = args.senha

    if args.tela:
        root = ctk.CTk(); root.withdraw()
        if args.tela.upper() == "ANALISE": app_win = AnaliseVendasWindow(root, config)
        elif args.tela.upper() == "CEP": app_win = App(root, config)
        root.mainloop()
    else:
        app = MainHub()
        if args.servidor: app.config["servidor"] = args.servidor
        if args.banco: app.config["banco"] = args.banco
        if args.usuario: app.config["usuario_bd"] = args.usuario
        if args.senha: app.config["senha_bd"] = args.senha
        app.mainloop()

